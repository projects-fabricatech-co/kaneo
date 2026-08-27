# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.comments import Comment
import datetime as dt

OUT = "/home/user/kaneo/docs/gestao-pessoas/Avaliacao_Desempenho_Squads_v1.xlsx"
F = "Arial"

# ---------- paleta ----------
C_TITLE   = "1F3864"
C_HEAD    = "2E5395"
C_SUB     = "D9E2F3"
C_INPUT   = "FFF2CC"   # preencher
C_CALC    = "F2F2F2"   # calculado
C_WARN    = "F8CBAD"
C_BAD     = "FFC7CE"
C_GOOD    = "C6EFCE"
BLUE      = "0000FF"

thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

def font(sz=10, b=False, color="000000", it=False):
    return Font(name=F, size=sz, bold=b, color=color, italic=it)

wb = openpyxl.Workbook()

def sheet(name):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    return ws

def title(ws, text, sub=None, span=12):
    ws["A1"] = text
    ws["A1"].font = Font(name=F, size=14, bold=True, color=C_TITLE)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name=F, size=9, italic=True, color="595959")
    ws.row_dimensions[1].height = 22

def header_row(ws, row, values, start=1, fill=C_HEAD):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start + i, value=v)
        c.font = Font(name=F, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 30

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w

def band(ws, row, text, span=14):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=F, size=10, bold=True, color=C_TITLE)
    c.fill = PatternFill("solid", fgColor=C_SUB)
    for j in range(2, span + 1):
        ws.cell(row=row, column=j).fill = PatternFill("solid", fgColor=C_SUB)

def mark(ws, rng, kind):
    fill = PatternFill("solid", fgColor={"in": C_INPUT, "calc": C_CALC}[kind])
    for r in ws[rng]:
        for c in r:
            c.fill = fill
            c.font = font(9, color=BLUE if kind == "in" else "000000")
            c.border = BOX
            c.alignment = Alignment(vertical="top", wrap_text=True)

# =========================================================
# DADOS DO MODELO
# =========================================================
EIXO_A = "A — Entrega e Resultado"
EIXO_B = "B — Técnico / Ofício"
EIXO_C = "C — Comportamental / Colaboração"
EIXO_D = "D — Autonomia / Senioridade"
EIXOS = [EIXO_A, EIXO_B, EIXO_C, EIXO_D]

COMP = [
 ("C01", EIXO_A, "Previsibilidade", "Cumpre o que combinou no prazo combinado, ou avisa cedo quando não vai cumprir.", "GP", "Sim"),
 ("C02", EIXO_A, "Qualidade da entrega", "O que entrega volta pouco — de QA, de review ou de produção.", "Líder Técnico", "Sim"),
 ("C03", EIXO_A, "Foco em valor", "Entende o problema do usuário/negócio antes de codar e questiona escopo que não resolve nada.", "Compartilhado", "Sim"),
 ("C04", EIXO_B, "Domínio técnico", "Conhece a stack e o domínio do produto no nível esperado da senioridade.", "Líder Técnico", "Sim"),
 ("C05", EIXO_B, "Design de solução", "Estrutura código, teste e dados de forma que sobreviva à próxima mudança.", "Líder Técnico", "Sim"),
 ("C06", EIXO_B, "Rigor operacional", "Cuida do que roda: observabilidade, deploy, incidente, dado sensível, segurança.", "Líder Técnico", "Sim"),
 ("C07", EIXO_C, "Comunicação", "Escreve e fala de forma que quem não estava no contexto entende e decide.", "GP", "Sim"),
 ("C08", EIXO_C, "Colaboração e review", "Melhora o trabalho dos outros e aceita ter o seu melhorado.", "Líder Técnico", "Sim"),
 ("C09", EIXO_C, "Confiabilidade e conduta", "É previsível como colega: presente nos rituais, cumpre acordos, respeita as pessoas.", "GP", "Sim"),
 ("C10", EIXO_D, "Autonomia e ownership", "Leva o problema até o fim sem ser puxado, e sabe quando escalar.", "Compartilhado", "Sim"),
 ("C11", EIXO_D, "Desenvolvimento de outros", "Faz o time render mais que a soma: ensina, documenta, melhora o processo.", "Líder Técnico", "Não"),
]

ANC = {
"Previsibilidade": [
 "Prazos estouram sem aviso; o time descobre na review. Compromisso assumido não é cumprido nem renegociado.",
 "Entrega, mas quase sempre depois; o aviso de atraso vem tarde demais para o PO replanejar.",
 "Cumpre o combinado na maioria das sprints; quando vai atrasar, avisa a tempo de o time reagir, com nova data.",
 "Estimativa confiável inclusive em item ambíguo; identifica risco antes de começar e negocia escopo em vez de estourar.",
 "É a referência de previsibilidade da squad; o roadmap é planejado sobre a estimativa dela sem colchão."],
"Qualidade da entrega": [
 "Entrega volta de QA ou de produção com frequência; deixa retrabalho recorrente para o próximo.",
 "Cobre o caminho feliz; falha em erro, borda e dado real. PR volta 2+ rodadas com frequência.",
 "PR passa em 1–2 rodadas; teste junto com o código, cobrindo erro e borda. Reprovação em QA é exceção.",
 "Escolhe o nível certo de teste e testa o risco; seus testes já pegaram regressão de terceiros.",
 "Melhorou a testabilidade do sistema (reduziu flaky, criou fixture/harness que o time inteiro usa)."],
"Foco em valor": [
 "Entrega o card literalmente, mesmo quando o card está errado; não pergunta o porquê.",
 "Entende a feature que está fazendo, mas só questiona requisito quando já está codando.",
 "Questiona requisito incoerente no refinamento, antes de codar.",
 "Propõe solução mais barata que atende a mesma dor; conhece o usuário real.",
 "Influencia o roadmap com argumento técnico-econômico; evita construção inteira desnecessária."],
"Domínio técnico": [
 "Erros conceituais recorrentes na stack principal; a mesma explicação precisa ser repetida a cada sprint.",
 "Resolve o caminho feliz; trava em erro, concorrência, performance ou dado real e depende de terceiro.",
 "Resolve sozinho problemas típicos do produto no nível dele; sabe onde o sistema quebra e onde procurar o resto.",
 "Resolve problema fora da zona conhecida; explica trade-off (custo, risco, prazo) e sustenta a escolha com dado.",
 "Referência técnica consultada por outras squads; decisões dele viram padrão ou ADR na empresa."],
"Design de solução": [
 "Cria estrutura nova sem perguntar e ignora fronteiras existentes; gera retrabalho.",
 "Copia padrão sem entender; escolhe por familiaridade, não por trade-off.",
 "Escolhe entre alternativas conhecidas e sabe dizer o que ganha e o que perde.",
 "Antecipa acoplamento e custo de reverter; escreve a decisão antes de codar.",
 "Decisão dele mudou o rumo de mais de um time e envelheceu bem."],
"Rigor operacional": [
 "Já comitou segredo ou expôs dado; deixa build vermelho; não considera autorização ao mexer em rota.",
 "Segue o processo quando lembrado; não instrumenta o que sobe.",
 "Valida entrada e autorização por padrão; segue o processo de release; consegue corrigir o próprio build.",
 "Pensa modelo de ameaça (enumeração, corrida, escalada); instrumenta antes de precisar; atua em rollback com segurança.",
 "Define e revisa a prática de segurança e observabilidade do time; reduziu tempo de build/deploy."],
"Comunicação": [
 "Status enganoso ou ausente; outras pessoas descobrem problemas tarde por falta de informação.",
 "Comunica, mas exige garimpo: card sem contexto, status 'quase pronto' que não significa nada.",
 "Card, PR e mensagem que quem não estava no contexto entende; status honesto, inclusive das más notícias.",
 "Adapta a mensagem ao público (PO, cliente, dev); documenta decisão e evita a mesma discussão duas vezes.",
 "A comunicação dela destrava decisão de outros times; os escritos viram referência de como o time comunica."],
"Colaboração e review": [
 "Aprova PR sem ler ou só comenta estilo; vira gargalo (PR parado 2+ dias sem sinal).",
 "Comenta razoavelmente, mas mistura nit com bloqueante e trava PR por gosto pessoal.",
 "Revisa em até 1 dia útil, pega problema real e separa 'bloqueante' de 'sugestão'.",
 "Comentário ensina e propõe alternativa; melhora o PR sem assumir o teclado.",
 "É referência de review no time; medeia discordância técnica sem escalar."],
"Confiabilidade e conduta": [
 "Acordos quebrados de forma recorrente; ausência sem aviso; cria clima no time.",
 "Presença irregular nos rituais; responde fora do SLA combinado sem avisar.",
 "Presente nos rituais, cumpre acordos, responde no SLA do time; comportamento consistente com todos.",
 "É o colega em quem o time se apoia quando algo depende de alguém aparecer.",
 "Sustenta a cultura do time na prática; comportamento é citado como o padrão esperado."],
"Autonomia e ownership": [
 "Só avança com direcionamento passo a passo; entrega a tarefa e ignora que o problema continua de pé.",
 "Precisa de acompanhamento acima do normal para o nível; escala tarde ou escala tudo, sem tentativa própria.",
 "Toca as próprias demandas com direcionamento normal; escala no momento certo, com hipótese e opções.",
 "Assume problema mal definido e devolve solução; puxa as pessoas necessárias sem precisar do gestor no meio.",
 "Assume tema crítico de ponta a ponta (técnico + stakeholder) e o gestor deixa de precisar acompanhar."],
"Desenvolvimento de outros": [
 "Retém conhecimento; não documenta; discordância técnica vira questão de ego.",
 "Ajuda quando pedido, mas assume o teclado em vez de ensinar.",
 "Pareia e faz onboarding quando necessário; deixa registro do que decidiu.",
 "Review dele ensina; escreve doc/ADR que o time usa; conduz melhoria da retro até acontecer.",
 "Eleva o nível técnico do time de forma visível; outros citam ele como referência de desenvolvimento."],
}

PESOS = {  # eixo -> (Dev/QA, Líder Técnico, PO/BA, UX)
 EIXO_A: (0.35, 0.25, 0.40, 0.35),
 EIXO_B: (0.30, 0.30, 0.15, 0.25),
 EIXO_C: (0.20, 0.25, 0.30, 0.25),
 EIXO_D: (0.15, 0.20, 0.15, 0.15),
}
SENIOR = [("Júnior", 2.7), ("Pleno", 3.0), ("Sênior", 3.2), ("Especialista", 3.4)]
PAPEIS = ["Dev/QA", "Líder Técnico", "PO/BA", "UX"]
SQUADS = ["Alfa", "Beta", "Delta"]
CONTRATOS = ["Contrato A", "Contrato B"]
TIPOS = ["Reconhecimento", "Positiva", "Ponto de atenção", "Incidente",
         "Ausência / Disponibilidade", "Contexto atenuante", "Conduta → encaminhado ao RH"]
IMPACTO = ["Baixo", "Médio", "Alto"]
CONF = ["Squad", "Gestão", "RH-restrito"]
SIMNAO = ["Sim", "Não", "Parcial"]
AUTOR = ["GP", "Líder Técnico"]
STPDI = ["Não iniciado", "Em andamento", "Concluído", "Cancelado"]
NOTAS = ["1", "2", "3", "4", "5", "N/O"]

CICLO = "2026-Q3"
COMP_NAMES = [c[2] for c in COMP]

COLABS = [
 (1, "Ana Ribeiro",   "M-1042", "Dev/QA", "Pleno",  "Alfa", "Contrato A", "Portal do Cliente", 100, dt.date(2025,3,10), None, "Marcos Lima", "Rafael Souza", "CLT"),
 (2, "Bruno Tavares", "M-1188", "Dev/QA", "Júnior", "Alfa", "Contrato A", "Portal do Cliente", 100, dt.date(2026,6,15), None, "Marcos Lima", "Rafael Souza", "CLT"),
 (3, "Carla Nunes",   "M-0977", "Líder Técnico", "Sênior", "Beta", "Contrato B", "Motor de Cobrança", 50, dt.date(2024,8,1), None, "Marcos Lima", "Carla Nunes", "PJ"),
]

# =========================================================
# 1. INSTRUÇÕES
# =========================================================
ws = sheet("Instruções")
title(ws, "Avaliação de Desempenho por Squad — GP + Líder Técnico",
      "Versão 1 · construída com as visões de RH, Gerência de Projetos e Liderança Técnica · revisar e ajustar antes de usar em produção")
widths(ws, {"A": 42, "B": 92})

rows = [
 ("band", "COMO ESTA PLANILHA FUNCIONA"),
 ("p", "", "O registro contínuo é o produto; a nota é subproduto. Se você só conseguir manter uma aba viva, mantenha 'Ocorrências'. Uma nota sem ocorrência vinculada não é avaliação, é impressão."),
 ("band", "RITMO DE USO"),
 ("kv", "Quando acontecer (60 segundos)", "Lance uma ocorrência. Meta realista: 2 a 4 por colaborador por mês, com ao menos 1 positiva a cada 2 registros."),
 ("kv", "Sexta-feira (15 min / squad)", "Abra o 'Painel'. Olhe quem está sem registro e sem 1:1. Não avalie nada."),
 ("kv", "Fim de ciclo (20 min / pessoa)", "Preencha 'Avaliação'. Você não escreve do zero: lê o que já foi acumulado."),
 ("kv", "Antes da devolutiva", "Calibração GP × Líder Técnico apenas nos itens divergentes. Depois, 1:1 com a aba 'Pauta 1a1'."),
 ("band", "QUEM AVALIA O QUÊ"),
 ("kv", "Líder Técnico (palavra final no técnico)", "Qualidade da entrega, domínio técnico, design de solução, rigor operacional, colaboração e review, desenvolvimento de outros."),
 ("kv", "GP (palavra final na entrega)", "Previsibilidade, comunicação, confiabilidade e conduta."),
 ("kv", "Compartilhado", "Foco em valor, autonomia e ownership. Quem não tem evidência deixa em branco — em branco é resposta legítima e melhor que chute."),
 ("band", "AS SEIS REGRAS DURAS"),
 ("n", "1", "3 é 'atende plenamente o esperado PARA A SENIORIDADE'. Não é nota ruim. Um Júnior 3 e um Sênior 3 são comportamentos diferentes — leia a âncora do nível na aba 'Âncoras'."),
 ("n", "2", "Nota 1, 2 ou 5 exige pelo menos uma ocorrência vinculada àquela competência no ciclo. Sem lastro, a célula acusa e a avaliação não fecha."),
 ("n", "3", "Nunca calcule a média entre a nota do GP e a do Líder Técnico. A divergência é a informação mais valiosa da planilha — ela é a pauta da calibração."),
 ("n", "4", "Ocorrência negativa que nunca foi conversada com a pessoa não entra no ciclo. Surpresa na avaliação é falha de gestão, não do colaborador."),
 ("n", "5", "O campo 'Fato observável' não aceita adjetivo de personalidade. Escreva 'Subiu hotfix sem aprovação na sexta 18h', não 'é descuidado'. O juízo vai em 'Impacto', e tem que ser defensável a partir do fato ao lado."),
 ("n", "6", "Comparação e ranking usam SEMPRE o gap contra a expectativa da senioridade, nunca a nota bruta, e nunca entre squads diferentes."),
 ("band", "LGPD E CONDUTA — LEIA ANTES DE ESCREVER"),
 ("p", "", "Não registre dado sensível: saúde, diagnóstico, religião, orientação sexual, filiação sindical, origem racial, vida familiar. Ausência registra o FATO, nunca o motivo."),
 ("p", "", "Caso de conduta, assédio, discriminação ou segurança da informação NÃO é apurado aqui: registre apenas 'encaminhado ao RH em <data>' e a apuração corre em processo próprio."),
 ("p", "", "O colaborador pode pedir acesso ao que se registrou sobre ele. Regra prática: não escreva nada que você não sustentaria lendo em voz alta para a pessoa."),
 ("p", "", "Arquivo único no drive corporativo, com coautoria. Nunca anexo de e-mail, nunca cópia local. Senha de planilha do Excel não é controle de acesso."),
 ("band", "LEGENDA DE CORES"),
 ("legend_in", "Fundo amarelo, texto azul", "Célula que VOCÊ preenche."),
 ("legend_calc", "Fundo cinza", "Célula calculada. Não digite por cima — você quebra o cálculo do painel."),
 ("p", "", "As abas 'Colaboradores', 'Competências', 'Âncoras' e 'Config' são o cadastro: ajuste-as uma vez, no começo, e não mexa mais no dia a dia."),
 ("band", "LIMITE HONESTO DESTA FERRAMENTA"),
 ("p", "", "Excel não tem controle de acesso por linha nem trilha de auditoria confiável. Ocorrências 'RH-restrito' deveriam viver em arquivo separado, só do RH. Use esta planilha para provar o modelo em 2–3 ciclos; se funcionar, migre para um sistema."),
]
r = 4
for item in rows:
    if item[0] == "band":
        band(ws, r, item[1], span=2)
    elif item[0] == "kv":
        ws.cell(row=r, column=1, value=item[1]).font = font(9, b=True)
        c = ws.cell(row=r, column=2, value=item[2]); c.font = font(9); c.alignment = Alignment(wrap_text=True, vertical="top")
    elif item[0] == "n":
        ws.cell(row=r, column=1, value="Regra " + item[1]).font = font(9, b=True, color=C_TITLE)
        c = ws.cell(row=r, column=2, value=item[2]); c.font = font(9); c.alignment = Alignment(wrap_text=True, vertical="top")
    elif item[0] == "p":
        c = ws.cell(row=r, column=2, value=item[2]); c.font = font(9); c.alignment = Alignment(wrap_text=True, vertical="top")
    elif item[0].startswith("legend"):
        c1 = ws.cell(row=r, column=1, value=item[1])
        c1.fill = PatternFill("solid", fgColor=C_INPUT if item[0]=="legend_in" else C_CALC)
        c1.font = font(9, color=BLUE if item[0]=="legend_in" else "000000"); c1.border = BOX
        c = ws.cell(row=r, column=2, value=item[2]); c.font = font(9)
    ws.row_dimensions[r].height = 28
    r += 1

# =========================================================
# 2. CONFIG
# =========================================================
cf = sheet("Config")
title(cf, "Configuração do modelo", "Ajuste aqui uma vez. Toda a planilha lê desta aba.")
widths(cf, {"A": 30, "B": 20, "C": 62, "D": 18, "E": 18,
            "G": 15, "H": 15, "I": 12, "J": 16, "K": 26, "L": 10, "M": 14, "N": 10, "O": 14, "P": 14, "Q": 8})

cf["A4"] = "Ciclo ativo"; cf["A4"].font = font(10, b=True)
cf["B4"] = CICLO
cf["B4"].fill = PatternFill("solid", fgColor=C_INPUT); cf["B4"].font = font(10, b=True, color=BLUE); cf["B4"].border = BOX
cf["C4"] = "Formato AAAA-Qn. O ciclo de cada ocorrência é derivado automaticamente da data do fato."
cf["C4"].font = font(9, it=True, color="595959")

band(cf, 6, "ESCALA DE AVALIAÇÃO — 1 a 5, sem meio ponto, com N/O obrigatório", span=5)
header_row(cf, 7, ["Nota", "Rótulo", "O que significa", "", ""])
esc = [
 (1, "Abaixo crítico", "Gera dano ou retrabalho recorrente. Exige plano de ação formal."),
 (2, "Abaixo do esperado", "Precisa de suporte acima do normal para o nível."),
 (3, "Atende", "Faz o esperado da senioridade, de forma consistente. É a nota saudável."),
 (4, "Supera", "Entrega acima do nível em situações reais e repetidas."),
 (5, "Referência", "Vira padrão para o time; outros copiam a prática."),
 ("N/O", "Não observado", "Sai do cálculo. Se aparecer 2 ciclos seguidos, ou o papel não exige, ou o gestor não está acompanhando."),
]
for i, (n, rot, sig) in enumerate(esc):
    rr = 8 + i
    cf.cell(row=rr, column=1, value=n).font = font(9, b=True)
    cf.cell(row=rr, column=2, value=rot).font = font(9)
    c = cf.cell(row=rr, column=3, value=sig); c.font = font(9); c.alignment = Alignment(wrap_text=True, vertical="top")
    for j in range(1, 4): cf.cell(row=rr, column=j).border = BOX
    cf.row_dimensions[rr].height = 24

band(cf, 15, "PESOS POR EIXO E PAPEL (a soma de cada coluna deve dar 100%)", span=5)
header_row(cf, 16, ["Eixo"] + PAPEIS)
for i, e in enumerate(EIXOS):
    rr = 17 + i
    cf.cell(row=rr, column=1, value=e).font = font(9, b=True)
    cf.cell(row=rr, column=1).border = BOX
    for j, p in enumerate(PESOS[e]):
        c = cf.cell(row=rr, column=2 + j, value=p)
        c.number_format = "0%"; c.fill = PatternFill("solid", fgColor=C_INPUT)
        c.font = font(9, color=BLUE); c.border = BOX; c.alignment = Alignment(horizontal="center")
cf.cell(row=21, column=1, value="Soma (deve ser 100%)").font = font(9, b=True, it=True)
for j in range(2, 6):
    L = get_column_letter(j)
    c = cf.cell(row=21, column=j, value="=SUM({0}17:{0}20)".format(L))
    c.number_format = "0%"; c.font = font(9, b=True); c.alignment = Alignment(horizontal="center"); c.border = BOX
cf.conditional_formatting.add("B21:E21", CellIsRule(operator="notEqual", formula=["1"], fill=PatternFill("solid", fgColor=C_BAD)))

band(cf, 23, "EXPECTATIVA MÍNIMA POR SENIORIDADE (a comparação e o ranking usam o GAP contra este número)", span=5)
header_row(cf, 24, ["Senioridade", "Expectativa", "", "", ""])
for i, (s, v) in enumerate(SENIOR):
    rr = 25 + i
    cf.cell(row=rr, column=1, value=s).font = font(9, b=True); cf.cell(row=rr, column=1).border = BOX
    c = cf.cell(row=rr, column=2, value=v)
    c.number_format = "0.0"; c.fill = PatternFill("solid", fgColor=C_INPUT); c.font = font(9, color=BLUE)
    c.border = BOX; c.alignment = Alignment(horizontal="center")

band(cf, 30, "PROIBIÇÕES EXPLÍCITAS", span=5)
proib = [
 "Nenhuma célula de nota pode ser fórmula sobre linhas de código, nº de PRs, story points, velocidade ou tickets fechados. Se a métrica existir, ela é contexto — nunca nota.",
 "Sem curva forçada. Distribuição imposta em squad de 6 pessoas é matematicamente absurda e destrói a colaboração em um ciclo.",
 "Sem ranking global entre squads. Comparar Sênior em contrato caótico com Júnior em sustentação estável vaza e o dano à confiança é permanente.",
 "Sem média entre a nota do GP e a do Líder Técnico. Média esconde a divergência, que é a informação mais valiosa aqui.",
]
for i, t in enumerate(proib):
    c = cf.cell(row=31 + i, column=1, value="• " + t)
    c.font = font(9); c.alignment = Alignment(wrap_text=True, vertical="top")
    cf.merge_cells(start_row=31 + i, start_column=1, end_row=31 + i, end_column=5)
    cf.row_dimensions[31 + i].height = 26

LISTS = [("G", "Senioridade", [s for s, _ in SENIOR]), ("H", "Papel", PAPEIS), ("I", "Squad", SQUADS),
         ("J", "Cliente / Contrato", CONTRATOS), ("K", "Tipo de ocorrência", TIPOS), ("L", "Impacto", IMPACTO),
         ("M", "Confidencialidade", CONF), ("N", "Sim/Não", SIMNAO), ("O", "Papel do autor", AUTOR),
         ("P", "Status PDI", STPDI), ("Q", "Notas", NOTAS)]
cf["G6"] = "LISTAS (alimentam os menus suspensos — acrescente valores no fim de cada coluna)"
cf["G6"].font = font(10, b=True, color=C_TITLE)
for col, head, vals in LISTS:
    c = cf[col + "7"]; c.value = head
    c.font = Font(name=F, size=9, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=C_HEAD)
    c.alignment = Alignment(wrap_text=True, vertical="center"); c.border = BOX
    for i, v in enumerate(vals):
        cc = cf[col + str(8 + i)]; cc.value = v
        cc.fill = PatternFill("solid", fgColor=C_INPUT); cc.font = font(9, color=BLUE); cc.border = BOX
cf.freeze_panes = "A5"

def dv(col_letter, first, last):
    return "=Config!${0}$8:${0}${1}".format(col_letter, 8 + last - 1)

DV = {
 "senioridade": "=Config!$G$8:$G$20", "papel": "=Config!$H$8:$H$20", "squad": "=Config!$I$8:$I$30",
 "contrato": "=Config!$J$8:$J$30", "tipo": "=Config!$K$8:$K$20", "impacto": "=Config!$L$8:$L$12",
 "conf": "=Config!$M$8:$M$12", "simnao": "=Config!$N$8:$N$12", "autor": "=Config!$O$8:$O$12",
 "stpdi": "=Config!$P$8:$P$12", "nota": "=Config!$Q$8:$Q$13",
 "colab": "=Colaboradores!$B$2:$B$61", "comp": "=Competências!$C$2:$C$12", "eixo": "=Config!$A$17:$A$20",
}

# =========================================================
# 3. COMPETÊNCIAS
# =========================================================
cp = sheet("Competências")
title(cp, "Modelo de competências", "11 competências em 4 eixos. Resista a passar de 12: acima disso o avaliador preenche no automático e a nota vira ruído.")
widths(cp, {"A": 8, "B": 30, "C": 26, "D": 74, "E": 18, "F": 16})
header_row(cp, 1, ["ID", "Eixo", "Competência", "Definição", "Avaliador dono (palavra final)", "Aplicável a Júnior?"])
cp.insert_rows(1); title(cp, "Modelo de competências", "11 competências em 4 eixos. O 'avaliador dono' é quem decide quando GP e Líder Técnico divergem e não há evidência que desempate.")
header_row(cp, 4, ["ID", "Eixo", "Competência", "Definição", "Avaliador dono (palavra final)", "Aplicável a Júnior?"])
for i, row in enumerate(COMP):
    rr = 5 + i
    for j, v in enumerate(row):
        c = cp.cell(row=rr, column=1 + j, value=v)
        c.font = font(9, b=(j == 2)); c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="top")
    cp.row_dimensions[rr].height = 30
cp.freeze_panes = "A5"

# =========================================================
# 4. ÂNCORAS
# =========================================================
an = sheet("Âncoras")
title(an, "Âncoras comportamentais", "O que cada nota significa, em comportamento observável. Leia a âncora ANTES de dar a nota — é o antídoto contra o halo e contra o 3 automático.")
widths(an, {"A": 26, "B": 46, "C": 46, "D": 46, "E": 46, "F": 46})
an["A3"] = "Lembre: a nota é sempre relativa à senioridade. Júnior 3 e Sênior 3 são comportamentos diferentes — o que muda é o tamanho do problema que a pessoa resolve sozinha, não a escala."
an["A3"].font = font(9, it=True, color="C00000")
header_row(an, 4, ["Competência", "1 — Abaixo crítico", "2 — Abaixo do esperado", "3 — Atende (esperado do nível)", "4 — Supera", "5 — Referência"])
for i, name in enumerate(COMP_NAMES):
    rr = 5 + i
    c = an.cell(row=rr, column=1, value=name); c.font = font(9, b=True); c.border = BOX
    c.alignment = Alignment(wrap_text=True, vertical="top")
    for j, txt in enumerate(ANC[name]):
        cc = an.cell(row=rr, column=2 + j, value=txt)
        cc.font = font(9); cc.border = BOX; cc.alignment = Alignment(wrap_text=True, vertical="top")
    an.row_dimensions[rr].height = 58
an.freeze_panes = "B5"

# =========================================================
# 5. COLABORADORES
# =========================================================
co = sheet("Colaboradores")
title(co, "Cadastro — uma linha por ALOCAÇÃO, não por pessoa",
      "Quem está 50% em dois contratos ocupa duas linhas. Cadastro por pessoa faz o desempenho ser atribuído ao contrato errado.")
COLS_CO = ["ID", "Nome", "Matrícula", "Papel", "Senioridade", "Squad", "Cliente / Contrato", "Projeto",
           "Alocação %", "Data de entrada na squad", "Data de saída (vazio = ativo)", "GP responsável",
           "Líder Técnico responsável", "Contratação", "Situação", "Em rampa?"]
widths(co, {"A": 6, "B": 22, "C": 11, "D": 14, "E": 13, "F": 10, "G": 15, "H": 20, "I": 10, "J": 14,
            "K": 15, "L": 16, "M": 18, "N": 12, "O": 11, "P": 12})
header_row(co, 4, COLS_CO)
LAST_CO = 61
for i, cdata in enumerate(COLABS):
    rr = 5 + i
    for j, v in enumerate(cdata):
        co.cell(row=rr, column=1 + j, value=v)
for rr in range(5, LAST_CO + 1):
    co.cell(row=rr, column=15, value='=IF($B{0}="","",IF($K{0}="","Ativo","Inativo"))'.format(rr))
    co.cell(row=rr, column=16, value='=IF(OR($B{0}="",$J{0}=""),"",IF(TODAY()-$J{0}<=90,"Sim — em rampa","Não"))'.format(rr))
mark(co, "A5:N{}".format(LAST_CO), "in")
mark(co, "O5:P{}".format(LAST_CO), "calc")
for rr in range(5, LAST_CO + 1):
    for col in ("J", "K"):
        co[col + str(rr)].number_format = "DD/MM/YYYY"
    co["I" + str(rr)].number_format = "0%"
for i, cdata in enumerate(COLABS):
    co.cell(row=5 + i, column=9, value=cdata[8] / 100.0)
co["A2"].value = "Preenchimento: as colunas amarelas. 'Situação' e 'Em rampa?' são calculadas. Sem 'Data de saída', quem saiu fica no relatório de pendências para sempre e o relatório perde credibilidade em dois meses."
co["A2"].font = font(9, it=True, color="595959")
co.freeze_panes = "C5"
for spec, key in [("D5:D{}", "papel"), ("E5:E{}", "senioridade"), ("F5:F{}", "squad"), ("G5:G{}", "contrato")]:
    d = DataValidation(type="list", formula1=DV[key], allow_blank=True)
    co.add_data_validation(d); d.add(spec.format(LAST_CO))

# =========================================================
# 6. OCORRÊNCIAS
# =========================================================
oc = sheet("Ocorrências")
title(oc, "Ocorrências — o registro contínuo",
      "Alvo: 60 segundos por lançamento. Registro mais recente no topo. Meta: pelo menos 1 registro positivo a cada 2 negativos.")
COLS_OC = ["Data do fato", "Colaborador", "Registrado por", "Papel do autor", "Tipo", "Competência",
           "Fato observável (sem adjetivo de personalidade)", "Impacto", "Consequência observada",
           "Ação tomada pelo gestor", "Informado ao colaborador?", "Link / evidência", "Confidencialidade",
           "Ciclo", "Registro válido?", "Seq. p/ pauta"]
widths(oc, {"A": 12, "B": 18, "C": 15, "D": 13, "E": 20, "F": 20, "G": 52, "H": 10, "I": 40, "J": 32,
            "K": 13, "L": 22, "M": 15, "N": 10, "O": 15, "P": 11})
oc["A3"] = "Regra de escrita: o FATO não aceita adjetivo. 'Subiu hotfix sem aprovação na sexta 18h' ✔ · 'É descuidado' ✘. O juízo vai em Consequência, e precisa ser defensável a partir do fato ao lado. Nada de saúde, vida pessoal ou motivo de ausência."
oc["A3"].font = font(9, it=True, color="C00000")
header_row(oc, 4, COLS_OC)
LAST_OC = 404
EX_OC = [
 (dt.date(2026,7,9), "Ana Ribeiro", "Rafael Souza", "Líder Técnico", "Positiva", "Colaboração e review",
  "No PR #812 apontou uma condição de corrida no resgate de cupom antes do merge, com o trecho e a sugestão de UPDATE atômico.",
  "Alto", "Evitou um bug de duplicidade de resgate que teria ido para produção no contrato A.",
  "Reconhecido na retro do dia 10/07.", "Sim", "PR #812", "Squad"),
 (dt.date(2026,7,22), "Bruno Tavares", "Marcos Lima", "GP", "Ponto de atenção", "Previsibilidade",
  "Card CB-118 ficou 4 dias sem atualização e o bloqueio (falta de credencial de homologação) só foi comunicado na review.",
  "Médio", "A sprint fechou com o item em carry over e o PO replanejou a demo com o cliente em cima da hora.",
  "Feedback na 1:1 de 23/07; combinado avisar bloqueio na daily do mesmo dia.", "Sim", "CB-118", "Squad"),
 (dt.date(2026,8,4), "Bruno Tavares", "Marcos Lima", "GP", "Positiva", "Comunicação",
  "Na daily de 04/08 avisou no mesmo dia que a integração de pagamento estava fora do ar, com print do erro e a alternativa de mockar.",
  "Médio", "O time replanejou a sprint no mesmo dia, sem perda de entrega.",
  "Reconhecido na daily.", "Sim", "", "Squad"),
 (dt.date(2026,8,12), "Carla Nunes", "Marcos Lima", "GP", "Contexto atenuante", "Previsibilidade",
  "Assumiu o on-call do Contrato B por duas semanas cobrindo férias, além dos 50% de alocação no Motor de Cobrança.",
  "Alto", "Reduziu em ~40% o tempo disponível para o roadmap no período.",
  "Registrado para a calibração; escopo da sprint 14 reduzido.", "Sim", "", "Squad"),
 (dt.date(2026,8,19), "Ana Ribeiro", "Rafael Souza", "Líder Técnico", "Ponto de atenção", "Qualidade da entrega",
  "PR #857 voltou em 3 rodadas: nas duas primeiras faltava teste do caminho de erro na rota de resgate.",
  "Médio", "Atrasou o merge em 2 dias e consumiu 3 rodadas de review do time.",
  "Combinado no 1:1 de 20/08: incluir teste de erro antes de abrir o PR.", "Sim", "PR #857", "Squad"),
 (dt.date(2026,8,25), "Carla Nunes", "Rafael Souza", "Líder Técnico", "Reconhecimento", "Desenvolvimento de outros",
  "Escreveu a ADR-07 sobre o advisory lock no cooldown e conduziu a sessão de 40 min com as squads Alfa e Beta.",
  "Alto", "Duas squads passaram a usar o mesmo padrão; a discussão não voltou nos refinamentos seguintes.",
  "Levado ao comitê de calibração como evidência de atuação de Especialista.", "Sim", "ADR-07", "Gestão"),
]
for i, row in enumerate(EX_OC):
    for j, v in enumerate(row):
        oc.cell(row=5 + i, column=1 + j, value=v)
for rr in range(5, LAST_OC + 1):
    oc.cell(row=rr, column=14, value='=IF($A{0}="","",YEAR($A{0})&"-Q"&ROUNDUP(MONTH($A{0})/3,0))'.format(rr))
    oc.cell(row=rr, column=15, value=('=IF($A{0}="","",IF(OR($B{0}="",$E{0}="",$F{0}="",$G{0}="",$I{0}="",$K{0}=""),'
                                      '"Faltam campos","OK"))').format(rr))
    oc.cell(row=rr, column=16, value=("=IF(OR($B{0}=\"\",$B{0}<>'Pauta 1a1'!$B$4,$N{0}<>'Pauta 1a1'!$B$5),\"\","
                                      "COUNTIFS($B$5:$B{0},'Pauta 1a1'!$B$4,$N$5:$N{0},'Pauta 1a1'!$B$5))").format(rr))
    oc["A" + str(rr)].number_format = "DD/MM/YYYY"
mark(oc, "A5:M{}".format(LAST_OC), "in")
mark(oc, "N5:P{}".format(LAST_OC), "calc")
oc.freeze_panes = "C5"
for spec, key in [("B5:B{}", "colab"), ("D5:D{}", "autor"), ("E5:E{}", "tipo"), ("F5:F{}", "comp"),
                  ("H5:H{}", "impacto"), ("K5:K{}", "simnao"), ("M5:M{}", "conf")]:
    d = DataValidation(type="list", formula1=DV[key], allow_blank=True)
    oc.add_data_validation(d); d.add(spec.format(LAST_OC))
dvd = DataValidation(type="date", operator="lessThanOrEqual", formula1="TODAY()", allow_blank=True,
                     error="Data do fato não pode ser futura.", errorTitle="Data inválida")
oc.add_data_validation(dvd); dvd.add("A5:A{}".format(LAST_OC))
oc.conditional_formatting.add("A5:M{}".format(LAST_OC),
    FormulaRule(formula=['$O5="Faltam campos"'], fill=PatternFill("solid", fgColor=C_BAD)))
oc.conditional_formatting.add("K5:K{}".format(LAST_OC),
    FormulaRule(formula=['AND($K5="Não",OR($E5="Ponto de atenção",$E5="Incidente"))'],
                fill=PatternFill("solid", fgColor=C_WARN)))
oc["G4"].comment = Comment("Fato observável: o que aconteceu, verificável por terceiro.\n\nBom: 'Subiu hotfix sem aprovação na sexta 18h'.\nRuim: 'É descuidado', 'falta postura', 'é proativo'.\n\nFeedback vago sem fato recai desproporcionalmente sobre mulheres e pessoas negras — é por isso que este campo é duro.", "RH")

# =========================================================
# 7. AVALIAÇÃO
# =========================================================
av = sheet("Avaliação")
title(av, "Avaliação e calibração — uma linha por colaborador × competência × ciclo",
      "GP e Líder Técnico preenchem na MESMA aba. Duas planilhas separadas viram duas verdades e ninguém concilia.")
COLS_AV = ["Ciclo", "ID", "Colaborador", "Senioridade", "Papel", "Competência", "Eixo", "Avaliador dono",
           "Nota GP", "Nota LT", "Divergência", "Sugestão", "Nota final", "Ocorr. vinculadas", "Lastro",
           "Justificativa da nota final (1 frase)", "Data da calibração"]
widths(av, {"A": 10, "B": 6, "C": 18, "D": 13, "E": 14, "F": 24, "G": 30, "H": 16, "I": 9, "J": 9,
            "K": 11, "L": 12, "M": 10, "N": 12, "O": 13, "P": 52, "Q": 14})
av["A3"] = "Quem não tem evidência deixa em branco — em branco é resposta legítima e melhor que chute. Divergência de 2+ pontos é obrigatoriamente pauta de calibração, e cada lado traz uma ocorrência que sustenta a nota."
av["A3"].font = font(9, it=True, color="C00000")
header_row(av, 4, COLS_AV)
LAST_AV = 304
seed = []
for cid, nome, *_ in COLABS:
    for comp in COMP:
        seed.append((CICLO, cid, comp[2]))
EXEMPLO = {"Previsibilidade": (3, None, 3), "Qualidade da entrega": (None, 2, 2), "Foco em valor": (3, 3, 3),
           "Domínio técnico": (None, 3, 3), "Design de solução": (None, 3, 3), "Rigor operacional": (None, 3, 3),
           "Comunicação": (4, None, 4), "Colaboração e review": (None, 5, 5), "Confiabilidade e conduta": (4, None, 4),
           "Autonomia e ownership": (3, 4, 4), "Desenvolvimento de outros": (None, 3, 3)}
for i, (cic, cid, comp) in enumerate(seed):
    rr = 5 + i
    av.cell(row=rr, column=1, value=cic)
    av.cell(row=rr, column=2, value=cid)
    if cid == 1:
        g, l, f_ = EXEMPLO[comp]
        if g is not None: av.cell(row=rr, column=9, value=g)
        if l is not None: av.cell(row=rr, column=10, value=l)
        av.cell(row=rr, column=13, value=f_)
    av.cell(row=rr, column=6, value=comp)
JUST = {"Qualidade da entrega": "PR #857 voltou 3 rodadas por teste de erro ausente; é o padrão do ciclo, não um caso isolado.",
        "Colaboração e review": "Review do PR #812 pegou condição de corrida antes do merge — review dela mudou a decisão técnica do time.",
        "Autonomia e ownership": "GP viu 3, LT viu 4; a evidência do LT (assumiu o tema de resgate ponta a ponta) desempatou."}
for comp, txt in JUST.items():
    for i, (cic, cid, c2) in enumerate(seed):
        if cid == 1 and c2 == comp:
            av.cell(row=5 + i, column=16, value=txt)
            av.cell(row=5 + i, column=17, value=dt.date(2026, 9, 30))
for rr in range(5, LAST_AV + 1):
    av.cell(row=rr, column=3, value='=IFERROR(INDEX(Colaboradores!$B:$B,MATCH($B{0},Colaboradores!$A:$A,0)),"")'.format(rr))
    av.cell(row=rr, column=4, value='=IFERROR(INDEX(Colaboradores!$E:$E,MATCH($B{0},Colaboradores!$A:$A,0)),"")'.format(rr))
    av.cell(row=rr, column=5, value='=IFERROR(INDEX(Colaboradores!$D:$D,MATCH($B{0},Colaboradores!$A:$A,0)),"")'.format(rr))
    av.cell(row=rr, column=7, value='=IFERROR(INDEX(Competências!$B:$B,MATCH($F{0},Competências!$C:$C,0)),"")'.format(rr))
    av.cell(row=rr, column=8, value='=IFERROR(INDEX(Competências!$E:$E,MATCH($F{0},Competências!$C:$C,0)),"")'.format(rr))
    av.cell(row=rr, column=11, value='=IF(AND(ISNUMBER($I{0}),ISNUMBER($J{0})),ABS($I{0}-$J{0}),"")'.format(rr))
    av.cell(row=rr, column=12, value=(
        '=IF(AND(NOT(ISNUMBER($I{0})),NOT(ISNUMBER($J{0}))),"",'
        'IF(NOT(ISNUMBER($J{0})),$I{0},'
        'IF(NOT(ISNUMBER($I{0})),$J{0},'
        'IF($K{0}=0,$I{0},'
        'IF($K{0}=1,IF($H{0}="Líder Técnico",$J{0},IF($H{0}="GP",$I{0},"CALIBRAR")),"CALIBRAR")))))').format(rr))
    av.cell(row=rr, column=14, value=('=IF($C{0}="","",COUNTIFS(Ocorrências!$B:$B,$C{0},Ocorrências!$F:$F,$F{0},'
                                      'Ocorrências!$N:$N,$A{0}))').format(rr))
    av.cell(row=rr, column=15, value=('=IF($M{0}="","",IF(OR($M{0}=1,$M{0}=2,$M{0}=5),'
                                      'IF($N{0}>0,"OK","SEM LASTRO"),"OK"))').format(rr))
    av.cell(row=rr, column=17).number_format = "DD/MM/YYYY"
mark(av, "A5:B{}".format(LAST_AV), "in")
mark(av, "F5:F{}".format(LAST_AV), "in")
mark(av, "I5:J{}".format(LAST_AV), "in")
mark(av, "M5:M{}".format(LAST_AV), "in")
mark(av, "P5:Q{}".format(LAST_AV), "in")
mark(av, "C5:E{}".format(LAST_AV), "calc")
mark(av, "G5:H{}".format(LAST_AV), "calc")
mark(av, "K5:L{}".format(LAST_AV), "calc")
mark(av, "N5:O{}".format(LAST_AV), "calc")
for col in ("I", "J", "M", "K", "L", "N"):
    for rr in range(5, LAST_AV + 1):
        av[col + str(rr)].alignment = Alignment(horizontal="center", vertical="center")
av.freeze_panes = "C5"
for spec, key in [("F5:F{}", "comp"), ("I5:I{}", "nota"), ("J5:J{}", "nota"), ("M5:M{}", "nota")]:
    d = DataValidation(type="list", formula1=DV[key], allow_blank=True)
    av.add_data_validation(d); d.add(spec.format(LAST_AV))
av.conditional_formatting.add("O5:O{}".format(LAST_AV),
    CellIsRule(operator="equal", formula=['"SEM LASTRO"'], fill=PatternFill("solid", fgColor=C_BAD),
               font=Font(name=F, size=9, bold=True, color="9C0006")))
av.conditional_formatting.add("K5:L{}".format(LAST_AV),
    FormulaRule(formula=['$K5>=2'], fill=PatternFill("solid", fgColor=C_WARN)))
av["M4"].comment = Comment("A nota final é DECISÃO, não cálculo. A coluna Sugestão só resolve o caso trivial.\n\nDivergência 1: decide o avaliador dono da competência.\nDivergência 2+: cada lado traz uma ocorrência. Se só um tem evidência, a nota dele prevalece. Nunca a média.", "RH")

# =========================================================
# 8. PAINEL
# =========================================================
pa = sheet("Painel")
title(pa, "Painel do gestor — o que abrir na segunda de manhã",
      "Uma linha por alocação, no ciclo ativo definido em Config. Comparação SEMPRE por gap contra a expectativa da senioridade, nunca por nota bruta.")
COLS_PA = ["ID", "Colaborador", "Squad", "Papel", "Senioridade", "Situação", "Em rampa?",
           "Eixo A", "Eixo B", "Eixo C", "Eixo D", "Nota ponderada", "Expectativa", "Gap", "Faixa",
           "Ocorr. no ciclo", "Positivas", "Atenção / Incidente", "Dias s/ ocorrência", "Dias desde 1:1",
           "Divergências ≥2", "Notas sem lastro", "Negativas não informadas", "Amplitude das notas", "Sinal"]
widths(pa, {"A": 5, "B": 19, "C": 8, "D": 13, "E": 12, "F": 9, "G": 12, "H": 8, "I": 8, "J": 8, "K": 8,
            "L": 12, "M": 11, "N": 8, "O": 12, "P": 10, "Q": 10, "R": 12, "S": 12, "T": 12, "U": 11,
            "V": 11, "W": 13, "X": 12, "Y": 26, "AA": 8, "AB": 8, "AC": 8, "AD": 8, "AE": 10, "AF": 10})
pa["A3"] = '=\"Ciclo ativo: \"&Config!$B$4&"   ·   Sinal ⚠ = bloqueio de integridade (nota sem lastro ou ocorrência negativa nunca conversada). Resolva antes de fechar o ciclo."'
pa["A3"].font = font(9, it=True, color="C00000")
header_row(pa, 4, COLS_PA)
LAST_PA = 64
for i, e in enumerate(EIXOS):
    c = pa.cell(row=4, column=8 + i)
    c.comment = Comment(e, "modelo")
for rr in range(5, LAST_PA + 1):
    src = rr  # Colaboradores linha correspondente
    pa.cell(row=rr, column=1, value='=IF(Colaboradores!$B{0}="","",Colaboradores!$A{0})'.format(src))
    for col, srccol in [(2, "B"), (3, "F"), (4, "D"), (5, "E"), (6, "O"), (7, "P")]:
        pa.cell(row=rr, column=col, value='=IF(Colaboradores!$B{0}="","",Colaboradores!${1}{0})'.format(src, srccol))
    for i, e in enumerate(EIXOS):
        pa.cell(row=rr, column=8 + i, value=(
            '=IF($B{0}="","",IFERROR(AVERAGEIFS(Avaliação!$M:$M,Avaliação!$B:$B,$A{0},'
            'Avaliação!$A:$A,Config!$B$4,Avaliação!$G:$G,Config!$A${1}),""))').format(rr, 17 + i))
        pa.cell(row=rr, column=27 + i, value=(
            '=IFERROR(INDEX(Config!$B$17:$E$20,MATCH(Config!$A${1},Config!$A$17:$A$20,0),'
            'MATCH($D{0},Config!$B$16:$E$16,0)),0)').format(rr, 17 + i))
    pa.cell(row=rr, column=31, value=('=IFERROR($H{0}*$AA{0},0)+IFERROR($I{0}*$AB{0},0)+'
                                      'IFERROR($J{0}*$AC{0},0)+IFERROR($K{0}*$AD{0},0)').format(rr))
    pa.cell(row=rr, column=32, value=('=IF(ISNUMBER($H{0}),$AA{0},0)+IF(ISNUMBER($I{0}),$AB{0},0)+'
                                      'IF(ISNUMBER($J{0}),$AC{0},0)+IF(ISNUMBER($K{0}),$AD{0},0)').format(rr))
    pa.cell(row=rr, column=12, value='=IF($AF{0}=0,"",$AE{0}/$AF{0})'.format(rr))
    pa.cell(row=rr, column=13, value='=IFERROR(INDEX(Config!$B$25:$B$28,MATCH($E{0},Config!$A$25:$A$28,0)),"")'.format(rr))
    pa.cell(row=rr, column=14, value='=IF(OR($L{0}="",$M{0}=""),"",$L{0}-$M{0})'.format(rr))
    pa.cell(row=rr, column=15, value=('=IF($L{0}="","",IF($L{0}<2.5,"Abaixo",IF($L{0}<3.5,"Atende",'
                                      'IF($L{0}<=4.2,"Supera","Referência"))))').format(rr))
    pa.cell(row=rr, column=16, value='=IF($B{0}="","",COUNTIFS(Ocorrências!$B:$B,$B{0},Ocorrências!$N:$N,Config!$B$4))'.format(rr))
    pa.cell(row=rr, column=17, value=('=IF($B{0}="","",COUNTIFS(Ocorrências!$B:$B,$B{0},Ocorrências!$N:$N,Config!$B$4,'
                                      'Ocorrências!$E:$E,"Reconhecimento")+COUNTIFS(Ocorrências!$B:$B,$B{0},'
                                      'Ocorrências!$N:$N,Config!$B$4,Ocorrências!$E:$E,"Positiva"))').format(rr))
    pa.cell(row=rr, column=18, value=('=IF($B{0}="","",COUNTIFS(Ocorrências!$B:$B,$B{0},Ocorrências!$N:$N,Config!$B$4,'
                                      'Ocorrências!$E:$E,"Ponto de atenção")+COUNTIFS(Ocorrências!$B:$B,$B{0},'
                                      'Ocorrências!$N:$N,Config!$B$4,Ocorrências!$E:$E,"Incidente"))').format(rr))
    pa.cell(row=rr, column=19, value=('=IF($B{0}="","",IF(COUNTIF(Ocorrências!$B:$B,$B{0})=0,"",'
                                      'TODAY()-_xlfn.MAXIFS(Ocorrências!$A:$A,Ocorrências!$B:$B,$B{0})))').format(rr))
    pa.cell(row=rr, column=20, value=("=IF($B{0}=\"\",\"\",IF(COUNTIF('Reuniões 1a1'!$B:$B,$B{0})=0,\"\","
                                      "TODAY()-_xlfn.MAXIFS('Reuniões 1a1'!$A:$A,'Reuniões 1a1'!$B:$B,$B{0})))").format(rr))
    pa.cell(row=rr, column=21, value=('=IF($B{0}="","",COUNTIFS(Avaliação!$B:$B,$A{0},Avaliação!$A:$A,Config!$B$4,'
                                      'Avaliação!$K:$K,">=2"))').format(rr))
    pa.cell(row=rr, column=22, value=('=IF($B{0}="","",COUNTIFS(Avaliação!$B:$B,$A{0},Avaliação!$A:$A,Config!$B$4,'
                                      'Avaliação!$O:$O,"SEM LASTRO"))').format(rr))
    pa.cell(row=rr, column=23, value=('=IF($B{0}="","",COUNTIFS(Ocorrências!$B:$B,$B{0},Ocorrências!$N:$N,Config!$B$4,'
                                      'Ocorrências!$K:$K,"Não",Ocorrências!$E:$E,"Ponto de atenção")+'
                                      'COUNTIFS(Ocorrências!$B:$B,$B{0},Ocorrências!$N:$N,Config!$B$4,'
                                      'Ocorrências!$K:$K,"Não",Ocorrências!$E:$E,"Incidente"))').format(rr))
    pa.cell(row=rr, column=24, value=('=IF($L{0}="","",_xlfn.MAXIFS(Avaliação!$M:$M,Avaliação!$B:$B,$A{0},'
                                      'Avaliação!$A:$A,Config!$B$4)-_xlfn.MINIFS(Avaliação!$M:$M,Avaliação!$B:$B,$A{0},'
                                      'Avaliação!$A:$A,Config!$B$4))').format(rr))
    pa.cell(row=rr, column=25, value=(
        '=IF($B{0}="","",'
        'IF(OR($V{0}>0,$W{0}>0),"BLOQUEIO",'
        'IF($L{0}="","Sem avaliação",'
        'IF($N{0}<-0.3,"Abaixo da expectativa",'
        'IF($U{0}>0,"Calibrar",'
        'IF($X{0}=0,"Nota chapada",'
        'IF(IF($T{0}="",0,$T{0})>30,"1:1 atrasada",'
        'IF(IF($S{0}="",99,$S{0})>45,"Sem registro","OK"))))))))').format(rr))
mark(pa, "A5:Y{}".format(LAST_PA), "calc")
mark(pa, "AA5:AF{}".format(LAST_PA), "calc")
for rr in range(5, LAST_PA + 1):
    for col in ("H", "I", "J", "K", "L", "M", "N"):
        pa[col + str(rr)].number_format = "0.00;-0.00;-"
        pa[col + str(rr)].alignment = Alignment(horizontal="center")
    for col in ("P", "Q", "R", "S", "T", "U", "V", "W", "X"):
        pa[col + str(rr)].alignment = Alignment(horizontal="center")
pa.cell(row=4, column=27, value="auxiliar: peso A").font = font(8, it=True)
for i, lab in enumerate(["peso A", "peso B", "peso C", "peso D", "numerador", "denominador"]):
    c = pa.cell(row=4, column=27 + i, value="aux " + lab)
    c.font = Font(name=F, size=8, italic=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="A6A6A6")
    c.alignment = Alignment(wrap_text=True)
pa.freeze_panes = "C5"
rng = "A5:Y{}".format(LAST_PA)
pa.conditional_formatting.add(rng, FormulaRule(formula=['$Y5="BLOQUEIO"'], fill=PatternFill("solid", fgColor=C_BAD)))
pa.conditional_formatting.add(rng, FormulaRule(formula=['$Y5="Abaixo da expectativa"'], fill=PatternFill("solid", fgColor=C_WARN)))
pa.conditional_formatting.add(rng, FormulaRule(formula=['$Y5="OK"'], fill=PatternFill("solid", fgColor=C_GOOD)))
pa.conditional_formatting.add("N5:N{}".format(LAST_PA),
    CellIsRule(operator="lessThan", formula=["0"], font=Font(name=F, size=9, bold=True, color="9C0006")))
pa.auto_filter.ref = "A4:Y{}".format(LAST_PA)

# =========================================================
# 9. PAUTA 1:1
# =========================================================
pt = sheet("Pauta 1a1")
title(pt, "Pauta de 1:1 — monta sozinha",
      "Escolha o colaborador em B4. Entre na sala sem preparar nada. Reconhecimentos vêm primeiro, de propósito.")
widths(pt, {"A": 26, "B": 22, "C": 22, "D": 14, "E": 46, "F": 44, "G": 14})
pt["A4"] = "Colaborador"; pt["A4"].font = font(10, b=True)
pt["B4"] = "Ana Ribeiro"
pt["B4"].fill = PatternFill("solid", fgColor=C_INPUT); pt["B4"].font = font(10, b=True, color=BLUE); pt["B4"].border = BOX
d = DataValidation(type="list", formula1=DV["colab"], allow_blank=True); pt.add_data_validation(d); d.add("B4")
pt["A5"] = "Ciclo"; pt["A5"].font = font(10, b=True)
pt["B5"] = "=Config!$B$4"; pt["B5"].fill = PatternFill("solid", fgColor=C_CALC); pt["B5"].border = BOX; pt["B5"].font = font(10, b=True)
pt["C4"] = '=IFERROR("ID "&INDEX(Colaboradores!$A:$A,MATCH($B$4,Colaboradores!$B:$B,0))&" · "&INDEX(Colaboradores!$D:$D,MATCH($B$4,Colaboradores!$B:$B,0))&" "&INDEX(Colaboradores!$E:$E,MATCH($B$4,Colaboradores!$B:$B,0))&" · squad "&INDEX(Colaboradores!$F:$F,MATCH($B$4,Colaboradores!$B:$B,0))&" · "&INDEX(Colaboradores!$G:$G,MATCH($B$4,Colaboradores!$B:$B,0)),"")'
pt["C4"].font = font(9, it=True, color="595959")
pt["C5"] = '=IFERROR("Nota ponderada "&TEXT(INDEX(Painel!$L:$L,MATCH($B$4,Painel!$B:$B,0)),"0.00")&"  ·  expectativa "&TEXT(INDEX(Painel!$M:$M,MATCH($B$4,Painel!$B:$B,0)),"0.0")&"  ·  gap "&TEXT(INDEX(Painel!$N:$N,MATCH($B$4,Painel!$B:$B,0)),"+0.00;-0.00")&"  ·  "&INDEX(Painel!$O:$O,MATCH($B$4,Painel!$B:$B,0)),"Sem avaliação no ciclo")'
pt["C5"].font = font(9, b=True, color=C_TITLE)

band(pt, 7, "1. COMPETÊNCIAS — a conversa é sobre a distância entre a nota e a expectativa, não sobre a nota", span=7)
header_row(pt, 8, ["Competência", "Nota GP", "Nota LT", "Nota final", "Âncora do nível em que ela está hoje", "Âncora do nível seguinte (o que falta)", "Gap vs. expectativa"])
for i, comp in enumerate(COMP_NAMES):
    rr = 9 + i
    pt.cell(row=rr, column=1, value=comp).font = font(9, b=True)
    pt.cell(row=rr, column=2, value=('=IFERROR(IF(SUMPRODUCT((Avaliação!$C$5:$C$304=$B$4)*(Avaliação!$A$5:$A$304=$B$5)*'
                                     '(Avaliação!$F$5:$F$304=$A{0})*ISNUMBER(Avaliação!$I$5:$I$304))=0,"",'
                                     'SUMIFS(Avaliação!$I$5:$I$304,Avaliação!$C$5:$C$304,$B$4,Avaliação!$A$5:$A$304,$B$5,'
                                     'Avaliação!$F$5:$F$304,$A{0})),"")').format(rr))
    pt.cell(row=rr, column=3, value=('=IFERROR(IF(SUMPRODUCT((Avaliação!$C$5:$C$304=$B$4)*(Avaliação!$A$5:$A$304=$B$5)*'
                                     '(Avaliação!$F$5:$F$304=$A{0})*ISNUMBER(Avaliação!$J$5:$J$304))=0,"",'
                                     'SUMIFS(Avaliação!$J$5:$J$304,Avaliação!$C$5:$C$304,$B$4,Avaliação!$A$5:$A$304,$B$5,'
                                     'Avaliação!$F$5:$F$304,$A{0})),"")').format(rr))
    pt.cell(row=rr, column=4, value=('=IFERROR(IF(SUMPRODUCT((Avaliação!$C$5:$C$304=$B$4)*(Avaliação!$A$5:$A$304=$B$5)*'
                                     '(Avaliação!$F$5:$F$304=$A{0})*ISNUMBER(Avaliação!$M$5:$M$304))=0,"",'
                                     'SUMIFS(Avaliação!$M$5:$M$304,Avaliação!$C$5:$C$304,$B$4,Avaliação!$A$5:$A$304,$B$5,'
                                     'Avaliação!$F$5:$F$304,$A{0})),"")').format(rr))
    pt.cell(row=rr, column=5, value='=IFERROR(INDEX(Âncoras!$B$5:$F$15,MATCH($A{0},Âncoras!$A$5:$A$15,0),$D{0}),"")'.format(rr))
    pt.cell(row=rr, column=6, value='=IFERROR(IF($D{0}>=5,"Já é referência.",INDEX(Âncoras!$B$5:$F$15,MATCH($A{0},Âncoras!$A$5:$A$15,0),$D{0}+1)),"")'.format(rr))
    pt.cell(row=rr, column=7, value='=IFERROR(IF($D{0}="","",$D{0}-INDEX(Painel!$M:$M,MATCH($B$4,Painel!$B:$B,0))),"")'.format(rr))
    pt.row_dimensions[rr].height = 46
mark(pt, "A9:G19", "calc")
for rr in range(9, 20):
    for col in ("B", "C", "D", "G"):
        pt[col + str(rr)].alignment = Alignment(horizontal="center", vertical="center")
        pt[col + str(rr)].number_format = "0.0;-0.0;-"
pt.conditional_formatting.add("G9:G19", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=C_WARN)))
pt.conditional_formatting.add("B9:C19", FormulaRule(formula=['AND(ISNUMBER($B9),ISNUMBER($C9),ABS($B9-$C9)>=2)'],
                                                    fill=PatternFill("solid", fgColor=C_BAD)))
pt["C8"].comment = Comment("Célula vermelha = GP e Líder Técnico divergiram em 2+ pontos. Diga isso na cara, na 1:1: 'o que discordamos sobre você'. Esconder a divergência é o que faz a devolutiva soar falsa.", "LT")

band(pt, 21, "2. OCORRÊNCIAS DO CICLO — o antídoto contra o viés de recência (reconhecimentos aparecem primeiro na leitura)", span=7)
header_row(pt, 22, ["Data", "Tipo", "Competência", "Impacto", "Fato observável", "Consequência observada", "Informado?"])
for k in range(1, 21):
    rr = 22 + k
    def pull(col):
        return ('=IFERROR(INDEX(Ocorrências!${0}$5:${0}$404,MATCH({1},Ocorrências!$P$5:$P$404,0)),"")').format(col, k)
    pt.cell(row=rr, column=1, value=pull("A")); pt["A" + str(rr)].number_format = "DD/MM/YYYY"
    pt.cell(row=rr, column=2, value=pull("E"))
    pt.cell(row=rr, column=3, value=pull("F"))
    pt.cell(row=rr, column=4, value=pull("H"))
    pt.cell(row=rr, column=5, value=pull("G"))
    pt.cell(row=rr, column=6, value=pull("I"))
    pt.cell(row=rr, column=7, value=pull("K"))
    pt.row_dimensions[rr].height = 32
mark(pt, "A23:G42", "calc")
pt.conditional_formatting.add("A23:G42", FormulaRule(formula=['OR($B23="Reconhecimento",$B23="Positiva")'],
                                                     fill=PatternFill("solid", fgColor=C_GOOD)))
pt.conditional_formatting.add("A23:G42", FormulaRule(formula=['AND($G23="Não",OR($B23="Ponto de atenção",$B23="Incidente"))'],
                                                     fill=PatternFill("solid", fgColor=C_BAD)))

band(pt, 44, "3. PLANO DE DESENVOLVIMENTO EM ANDAMENTO (vem da aba PDI — no máximo 2 focos)", span=7)
header_row(pt, 45, ["Competência", "Nota atual → alvo", "Ação concreta no trabalho real", "Prazo", "Como saberemos que fechou", "Status", ""])
for k in range(1, 4):
    rr = 45 + k
    m = '=IFERROR(INDEX(PDI!${0}$5:${0}$204,MATCH(1,INDEX((PDI!$B$5:$B$204=$B$4)*(PDI!$N$5:$N$204=' + str(k) + '),0),0)),"")'
    pt.cell(row=rr, column=1, value='=IFERROR(INDEX(PDI!$C$5:$C$204,MATCH({0},PDI!$O$5:$O$204,0)),"")'.format(k))
    pt.cell(row=rr, column=2, value=('=IFERROR(INDEX(PDI!$D$5:$D$204,MATCH({0},PDI!$O$5:$O$204,0))&" → "&'
                                     'INDEX(PDI!$E$5:$E$204,MATCH({0},PDI!$O$5:$O$204,0)),"")').format(k))
    pt.cell(row=rr, column=3, value='=IFERROR(INDEX(PDI!$H$5:$H$204,MATCH({0},PDI!$O$5:$O$204,0)),"")'.format(k))
    pt.cell(row=rr, column=4, value='=IFERROR(INDEX(PDI!$K$5:$K$204,MATCH({0},PDI!$O$5:$O$204,0)),"")'.format(k))
    pt["D" + str(rr)].number_format = "DD/MM/YYYY"
    pt.cell(row=rr, column=5, value='=IFERROR(INDEX(PDI!$I$5:$I$204,MATCH({0},PDI!$O$5:$O$204,0)),"")'.format(k))
    pt.cell(row=rr, column=6, value='=IFERROR(INDEX(PDI!$M$5:$M$204,MATCH({0},PDI!$O$5:$O$204,0)),"")'.format(k))
    pt.row_dimensions[rr].height = 34
mark(pt, "A46:F48", "calc")

band(pt, 50, "4. ESCUTA — as perguntas que o gestor esquece de fazer", span=7)
perg = ["Como está a sua carga? O que está pesando mais do que deveria?",
        "O que te travou neste ciclo e eu não vi?",
        "O que você quer estar fazendo daqui a um ano?",
        "O que eu preciso fazer diferente como gestor?"]
for i, p in enumerate(perg):
    rr = 51 + i
    c = pt.cell(row=rr, column=1, value=p); c.font = font(9, b=True); c.alignment = Alignment(wrap_text=True, vertical="top")
    pt.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=7)
    pt.row_dimensions[rr].height = 26
mark(pt, "B51:B54", "in")

band(pt, 56, "5. COMPROMISSOS DESTA 1:1 — reaparecem como pendência na próxima", span=7)
header_row(pt, 57, ["Compromisso", "Dono", "Prazo", "", "", "", ""])
mark(pt, "A58:C62", "in")
for rr in range(58, 63):
    pt["C" + str(rr)].number_format = "DD/MM/YYYY"
pt["E57"] = "Registre a 1:1 na aba 'Reuniões 1a1' — é o que alimenta o alerta de 1:1 atrasada no Painel."
pt["E57"].font = font(9, it=True, color="C00000")
pt.freeze_panes = "A6"

# =========================================================
# 10. PDI
# =========================================================
pd_ = sheet("PDI")
title(pd_, "Plano de desenvolvimento", "Máximo 2 focos por colaborador por ciclo. Plano com 6 itens não é plano. Todo foco nasce de uma competência do modelo e de pelo menos uma ocorrência.")
COLS_PDI = ["Ciclo", "Colaborador", "Competência-alvo", "Nota atual", "Nota alvo", "Gap",
            "Evidência do gap (ocorrência / PR / card)", "Ação concreta no trabalho real",
            "Como saberemos que fechou (evidência de saída)", "Apoio (quem / o quê)", "Prazo",
            "Último check-in", "Status", "Ordem do foco", "Chave"]
widths(pd_, {"A": 10, "B": 19, "C": 22, "D": 10, "E": 10, "F": 8, "G": 34, "H": 44, "I": 44, "J": 22,
             "K": 12, "L": 13, "M": 14, "N": 12, "O": 10})
pd_["A3"] = "'Melhorar em testes' não é PDI. 'Nos próximos 3 PRs de rota nova, incluir teste de caminho de erro sem que o review precise pedir — verificado por mim em 30/09' é PDI."
pd_["A3"].font = font(9, it=True, color="C00000")
header_row(pd_, 4, COLS_PDI)
LAST_PDI = 204
pd_ex = [
 (CICLO, "Ana Ribeiro", "Qualidade da entrega", 3, "PR #857 (3 rodadas por teste de erro ausente) + reprovação de QA na sprint 12",
  "Ser a dona dos testes da nova rota de resgate — escopo real, com o LT como rede de proteção, não curso.",
  "Nos próximos 3 PRs de rota nova, teste de caminho de erro presente sem o review precisar pedir.",
  "Rafael Souza (LT), pareamento de 1h/semana", dt.date(2026,11,30), dt.date(2026,9,30), "Em andamento", 1),
 (CICLO, "Ana Ribeiro", "Autonomia e ownership", 4, "Divergência GP 3 × LT 4 na calibração de 30/09",
  "Conduzir o refinamento do épico de cupom de ponta a ponta, incluindo a conversa com o PO.",
  "Dois refinamentos conduzidos sem o GP precisar entrar para destravar.",
  "Marcos Lima (GP), observando sem intervir", dt.date(2026,12,15), dt.date(2026,9,30), "Não iniciado", 2),
]
for i, row in enumerate(pd_ex):
    rr = 5 + i
    vals = [row[0], row[1], row[2], None, row[3], None, row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11]]
    for j, v in enumerate(vals):
        if v is not None:
            pd_.cell(row=rr, column=1 + j, value=v)
for rr in range(5, LAST_PDI + 1):
    pd_.cell(row=rr, column=4, value=('=IF($B{0}="","",IFERROR(IF(SUMPRODUCT((Avaliação!$C$5:$C$304=$B{0})*'
                                      '(Avaliação!$A$5:$A$304=$A{0})*(Avaliação!$F$5:$F$304=$C{0})*'
                                      'ISNUMBER(Avaliação!$M$5:$M$304))=0,"",SUMIFS(Avaliação!$M$5:$M$304,'
                                      'Avaliação!$C$5:$C$304,$B{0},Avaliação!$A$5:$A$304,$A{0},'
                                      'Avaliação!$F$5:$F$304,$C{0})),""))').format(rr))
    pd_.cell(row=rr, column=6, value='=IF(OR($D{0}="",$E{0}=""),"",$E{0}-$D{0})'.format(rr))
    pd_.cell(row=rr, column=15, value=("=IF(OR($B{0}=\"\",$B{0}<>'Pauta 1a1'!$B$4,$A{0}<>'Pauta 1a1'!$B$5),\"\","
                                       "COUNTIFS($B$5:$B{0},'Pauta 1a1'!$B$4,$A$5:$A{0},'Pauta 1a1'!$B$5))").format(rr))
    for col in ("K", "L"):
        pd_[col + str(rr)].number_format = "DD/MM/YYYY"
mark(pd_, "A5:C{}".format(LAST_PDI), "in")
mark(pd_, "E5:E{}".format(LAST_PDI), "in")
mark(pd_, "G5:N{}".format(LAST_PDI), "in")
mark(pd_, "D5:D{}".format(LAST_PDI), "calc")
mark(pd_, "F5:F{}".format(LAST_PDI), "calc")
mark(pd_, "O5:O{}".format(LAST_PDI), "calc")
pd_.freeze_panes = "C5"
for spec, key in [("B5:B{}", "colab"), ("C5:C{}", "comp"), ("E5:E{}", "nota"), ("M5:M{}", "stpdi")]:
    d = DataValidation(type="list", formula1=DV[key], allow_blank=True)
    pd_.add_data_validation(d); d.add(spec.format(LAST_PDI))
pd_["N4"].comment = Comment("Ordem do foco: 1 ou 2. Se você precisou escrever 3, o plano já falhou — escolha.", "LT")

# =========================================================
# 11. REUNIÕES 1:1
# =========================================================
rn = sheet("Reuniões 1a1")
title(rn, "Registro de 1:1", "Duas linhas por reunião bastam. É este registro que alimenta o alerta de '1:1 atrasada' no Painel.")
widths(rn, {"A": 12, "B": 20, "C": 18, "D": 50, "E": 50, "F": 12, "G": 14})
header_row(rn, 4, ["Data", "Colaborador", "Gestor", "Temas tratados", "Compromissos assumidos", "Prazo", "Status"])
LAST_RN = 204
rn_ex = [(dt.date(2026,7,23), "Bruno Tavares", "Marcos Lima", "Comunicação de bloqueio; combinado avisar na daily do mesmo dia.",
          "Avisar bloqueio na daily do dia em que ocorre.", dt.date(2026,8,31), "Concluído"),
         (dt.date(2026,8,20), "Ana Ribeiro", "Rafael Souza", "Teste de caminho de erro antes de abrir PR; interesse em assumir refinamento.",
          "Incluir teste de erro nos próximos 3 PRs de rota nova.", dt.date(2026,11,30), "Em andamento")]
for i, row in enumerate(rn_ex):
    for j, v in enumerate(row):
        rn.cell(row=5 + i, column=1 + j, value=v)
mark(rn, "A5:G{}".format(LAST_RN), "in")
for rr in range(5, LAST_RN + 1):
    rn["A" + str(rr)].number_format = "DD/MM/YYYY"
    rn["F" + str(rr)].number_format = "DD/MM/YYYY"
d = DataValidation(type="list", formula1=DV["colab"], allow_blank=True); rn.add_data_validation(d); d.add("B5:B{}".format(LAST_RN))
d2 = DataValidation(type="list", formula1=DV["stpdi"], allow_blank=True); rn.add_data_validation(d2); d2.add("G5:G{}".format(LAST_RN))
rn.freeze_panes = "C5"

# =========================================================
# 12. ALERTAS
# =========================================================
al = sheet("Alertas")
title(al, "Integridade do ciclo — o que precisa estar zerado antes de fechar",
      "Cada linha é um mecanismo contra um viés ou contra uma falha de gestão. Vazio é o problema real, não a nota baixa.")
widths(al, {"A": 4, "B": 46, "C": 12, "D": 88})
header_row(al, 4, ["#", "Verificação", "Ocorrências", "Por que isso importa / o que fazer"])
CHECKS = [
 ("Colaboradores ativos sem avaliação no ciclo", '=COUNTIF(Painel!$Y$5:$Y$64,"Sem avaliação")',
  "É a visão mais valiosa da planilha. Ciclo com metade da squad sem nota não vale para mérito nem para PDI."),
 ("Notas 1, 2 ou 5 sem ocorrência vinculada", '=SUM(Painel!$V$5:$V$64)',
  "Nota sem lastro é impressão, não avaliação. Registre a ocorrência que sustenta a nota ou mude a nota. Não feche o ciclo com isto acima de zero."),
 ("Ocorrências negativas nunca conversadas com a pessoa", '=SUM(Painel!$W$5:$W$64)',
  "Não entram no cálculo do ciclo. Surpresa na devolutiva é falha de gestão e é o que derruba um processo demissional."),
 ("Divergências GP × Líder Técnico de 2+ pontos", '=SUM(Painel!$U$5:$U$64)',
  "Não é erro, é sinal. Cada lado traz uma ocorrência. Se só um tem evidência, a nota dele prevalece — nunca a média."),
 ("Sem nenhuma ocorrência há mais de 45 dias", '=COUNTIF(Painel!$S$5:$S$64,">45")',
  "Ou a pessoa é invisível para a gestão, ou o registro parou. Nos dois casos a avaliação do ciclo vai ser chute."),
 ("Sem 1:1 há mais de 30 dias", '=COUNTIF(Painel!$T$5:$T$64,">30")',
  "Planilha que não vira conversa morre. Esta é a falha que mata a ferramenta — todas as outras são consertáveis."),
 ("Avaliação com nota chapada (mesma nota em tudo)", '=COUNTIFS(Painel!$X$5:$X$64,0)',
  "Bom desempenho é irregular. Nota chapada é halo ou preguiça. Reabra as âncoras e avalie competência por competência."),
 ("Registros de ocorrência com campo obrigatório vazio", '=COUNTIF(Ocorrências!$O$5:$O$404,"Faltam campos")',
  "Sem fato, consequência ou 'informado?', o registro não sustenta uma conversa difícil nem uma revisão do RH."),
 ("Pessoas com 3+ registros negativos e nenhum positivo", '=COUNTIFS(Painel!$R$5:$R$64,">=3",Painel!$Q$5:$Q$64,0)',
  "Ou a situação é grave de verdade (e pede plano formal, não planilha), ou o registro está enviesado. As duas leituras exigem ação."),
 ("Concentração de registros no último mês do ciclo", '=COUNTIFS(Ocorrências!$N$5:$N$404,Config!$B$4,Ocorrências!$A$5:$A$404,">="&DATE(2026,9,1))',
  "Se a maior parte do ciclo foi registrada nas últimas semanas, a avaliação é memória recente disfarçada de dado. Ajuste a data de referência desta linha a cada ciclo."),
]
for i, (nome, form, porq) in enumerate(CHECKS):
    rr = 5 + i
    al.cell(row=rr, column=1, value=i + 1).font = font(9, b=True)
    c = al.cell(row=rr, column=2, value=nome); c.font = font(9, b=True)
    cc = al.cell(row=rr, column=3, value=form); cc.font = font(11, b=True); cc.alignment = Alignment(horizontal="center")
    cd = al.cell(row=rr, column=4, value=porq); cd.font = font(9)
    for j in range(1, 5):
        al.cell(row=rr, column=j).border = BOX
        al.cell(row=rr, column=j).alignment = Alignment(wrap_text=True, vertical="top",
            horizontal="center" if j in (1, 3) else "left")
    al.row_dimensions[rr].height = 34
al.conditional_formatting.add("C5:C14", CellIsRule(operator="greaterThan", formula=["0"],
    fill=PatternFill("solid", fgColor=C_BAD), font=Font(name=F, size=11, bold=True, color="9C0006")))
al.conditional_formatting.add("C5:C14", CellIsRule(operator="equal", formula=["0"],
    fill=PatternFill("solid", fgColor=C_GOOD)))
al["B17"] = "Fontes e premissas"
al["B17"].font = font(10, b=True, color=C_TITLE)
al["B18"] = "Modelo de competências, escala, pesos e expectativas: definidos nesta v1 a partir das visões consolidadas de RH, Gerência de Projetos e Liderança Técnica. Não são benchmark de mercado — são pontos de partida para o gestor calibrar."
al["B19"] = "Limiares (45 dias sem ocorrência, 30 dias sem 1:1, divergência ≥2, gap -0,3) são convenções desta v1, editáveis. Estão aqui e nas fórmulas do Painel; ao mudar um, mude nos dois lugares."
al["B20"] = "Expectativas por senioridade (Júnior 2,7 · Pleno 3,0 · Sênior 3,2 · Especialista 3,4) são premissa, não medição. Ajuste na aba Config depois do primeiro ciclo real."
for r_ in (18, 19, 20):
    al.merge_cells(start_row=r_, start_column=2, end_row=r_, end_column=4)
    al.cell(row=r_, column=2).font = font(9, it=True, color="595959")
    al.cell(row=r_, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    al.row_dimensions[r_].height = 28

# ordem final
del wb["Sheet"]
order = ["Instruções", "Painel", "Ocorrências", "Avaliação", "Pauta 1a1", "PDI", "Alertas",
         "Reuniões 1a1", "Colaboradores", "Competências", "Âncoras", "Config"]
wb._sheets = [wb[n] for n in order]
for name in order:
    for row in wb[name].iter_rows():
        for c in row:
            if c.font is None or c.font.name != F:
                c.font = Font(name=F, size=c.font.size or 10, bold=c.font.bold, italic=c.font.italic,
                              color=c.font.color)
wb.active = 1
wb.save(OUT)
print("OK", OUT)
