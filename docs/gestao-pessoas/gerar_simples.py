# -*- coding: utf-8 -*-
"""Avaliacao semanal de desempenho — versao simples. Um arquivo, 4 abas."""
import re, sys, datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.comments import Comment

OUT = "/home/user/kaneo/docs/gestao-pessoas/Avaliacao_Semanal.xlsx"
F = "Arial"
C_TIT, C_HEAD = "1F3864", "2E5395"
C_IN, C_CALC = "FFF2CC", "F2F2F2"
VERDE, AMAR, VERM = "C6EFCE", "FFEB9C", "FFC7CE"
TVERDE, TAMAR, TVERM = "006100", "9C6500", "9C0006"
BLUE = "0000FF"
thin = Side(style="thin", color="BFBFBF"); BOX = Border(thin, thin, thin, thin)

LAST_CAD, LAST_AV, LAST_OC, LAST_RES = 53, 504, 204, 53

IND = ["Entrega", "Qualidade", "Comunicação", "Colaboração", "Autonomia"]
IND_DESC = [
 "Entregou o que combinou, no prazo combinado. Avisou cedo quando não ia dar.",
 "O que entregou voltou pouco — de review, de QA ou de produção.",
 "Deu status claro e honesto. Quem não estava no contexto entendeu.",
 "Ajudou o time, revisou bem, recebeu crítica sem defensividade.",
 "Tocou o próprio trabalho sem precisar ser puxado. Escalou na hora certa.",
]
PESSOAS = [
 ("Ana Ribeiro", "Alfa", "Dev", "Marcos Lima", "Rafael Souza"),
 ("Bruno Tavares", "Alfa", "Dev", "Marcos Lima", "Rafael Souza"),
 ("Carla Nunes", "Beta", "Tech Lead", "Marcos Lima", "Rafael Souza"),
]
NOMES = [p[0] for p in PESSOAS]
AVALIADORES = ["GP", "Líder Técnico"]
TIPOS_OC = ["Positiva", "Ponto de atenção", "Crítica"]
NOTAS = [1, 2, 3, 4, 5]

# notas: (semana, pessoa, avaliador, quem, [5 notas], comentario)
AVAL = [
 (dt.date(2026,8,7), "Ana Ribeiro", "GP", "Marcos Lima", [4,4,5,4,4],
  "Semana boa. Avisou na segunda que o card de cupom ia atrasar e já trouxe o replanejamento."),
 (dt.date(2026,8,7), "Ana Ribeiro", "Líder Técnico", "Rafael Souza", [3,3,4,4,4],
  "PR #812 excelente na revisão dos outros, mas o dela voltou 2 rodadas por falta de teste de erro."),
 (dt.date(2026,8,7), "Bruno Tavares", "GP", "Marcos Lima", [3,3,3,4,2],
  "Ficou 3 dias travado em credencial de homologação sem avisar. Conversei na sexta."),
 (dt.date(2026,8,7), "Bruno Tavares", "Líder Técnico", "Rafael Souza", [2,3,3,4,2],
  "Ainda depende de mim para fechar caso de erro. Combinamos pareamento 1h por semana."),
 (dt.date(2026,8,7), "Carla Nunes", "GP", "Marcos Lima", [5,5,5,5,5],
  "Segurou o on-call do Contrato B cobrindo férias e ainda entregou a sprint."),
 (dt.date(2026,8,7), "Carla Nunes", "Líder Técnico", "Rafael Souza", [4,5,4,5,5],
  "ADR-07 virou padrão em duas squads. Referência técnica do time."),
 (dt.date(2026,8,14), "Ana Ribeiro", "GP", "Marcos Lima", [4,4,5,5,4],
  "Conduziu o refinamento sozinha, sem eu precisar entrar."),
 (dt.date(2026,8,14), "Ana Ribeiro", "Líder Técnico", "Rafael Souza", [3,2,4,4,4],
  "PR #857 voltou 3 rodadas, de novo por teste de caminho de erro. É padrão, não caso isolado."),
 (dt.date(2026,8,14), "Bruno Tavares", "GP", "Marcos Lima", [3,3,4,4,3],
  "Melhorou: avisou o bloqueio na daily do mesmo dia, como combinamos."),
 (dt.date(2026,8,14), "Bruno Tavares", "Líder Técnico", "Rafael Souza", [3,3,4,4,3],
  "Evoluiu na comunicação técnica. Autonomia ainda é o ponto."),
 (dt.date(2026,8,14), "Carla Nunes", "GP", "Marcos Lima", [5,4,5,5,5],
  "Sem ressalvas. Vale conversar sobre carreira — ela puxou o assunto."),
 (dt.date(2026,8,14), "Carla Nunes", "Líder Técnico", "Rafael Souza", [5,5,4,5,5],
  "Conduziu a sessão de arquitetura com Alfa e Beta. Nível de especialista."),
]
OCORR = [
 (dt.date(2026,8,11), "Ana Ribeiro", "Positiva", "Rafael Souza",
  "No PR #812 pegou uma condição de corrida no resgate de cupom antes do merge. Evitou bug de duplicidade em produção."),
 (dt.date(2026,8,12), "Bruno Tavares", "Ponto de atenção", "Marcos Lima",
  "Card CB-118 ficou 4 dias sem atualização; o bloqueio só apareceu na review. Conversado na 1:1 de 13/08."),
 (dt.date(2026,8,25), "Carla Nunes", "Positiva", "Rafael Souza",
  "Escreveu a ADR-07 do advisory lock e conduziu a sessão de 40 min com as squads Alfa e Beta."),
 (dt.date(2026,8,26), "Carla Nunes", "Ponto de atenção", "Marcos Lima",
  "Disse que quer trilha técnica e não gestão. Aceitou o papel de LT por falta de alternativa."),
]

def fnt(sz=10, b=False, color="000000", it=False):
    return Font(name=F, size=sz, bold=b, color=color, italic=it)

wb = openpyxl.Workbook(); wb.calculation.fullCalcOnLoad = True
def sh(n):
    ws = wb.create_sheet(n); ws.sheet_view.showGridLines = False; return ws
def titulo(ws, t, sub=None):
    ws["A1"] = t; ws["A1"].font = Font(name=F, size=15, bold=True, color=C_TIT)
    ws.row_dimensions[1].height = 24
    if sub:
        ws["A2"] = sub; ws["A2"].font = fnt(9, it=True, color="595959")
def cab(ws, row, vals, alturas=34):
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=1 + i, value=v)
        c.font = Font(name=F, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD); c.border = BOX
        c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    ws.row_dimensions[row] = ws.row_dimensions[row]; ws.row_dimensions[row].height = alturas
def larg(ws, spec):
    for k, v in spec.items(): ws.column_dimensions[k].width = v
def pinta(ws, rng, kind):
    fill = PatternFill("solid", fgColor={"in": C_IN, "calc": C_CALC}[kind])
    for r in ws[rng]:
        for c in r:
            c.fill = fill; c.border = BOX
            c.font = fnt(9, color=BLUE if kind == "in" else "000000")
            c.alignment = Alignment(vertical="top", wrap_text=True)

# =====================================================================
# CADASTRO (pessoas + parâmetros + listas)
# =====================================================================
cd = sh("Cadastro")
titulo(cd, "Cadastro", "Preencha uma vez. Alimenta os menus suspensos das outras abas.")
larg(cd, {"A": 24, "B": 12, "C": 14, "D": 18, "E": 18, "F": 4,
          "G": 22, "H": 12, "I": 4, "J": 18, "K": 20, "L": 10})
cab(cd, 4, ["Colaborador", "Squad", "Papel", "GP", "Líder Técnico"])
for i, p in enumerate(PESSOAS):
    for j, v in enumerate(p): cd.cell(row=5 + i, column=1 + j, value=v)
pinta(cd, "A5:E{}".format(LAST_CAD), "in")

cd["G4"] = "FARÓIS"; cd["G4"].font = fnt(11, b=True, color=C_TIT)
for i, (rot, val, cor, tcor) in enumerate([
        ("Verde a partir de", 4.0, VERDE, TVERDE),
        ("Atenção a partir de", 3.0, AMAR, TAMAR),
        ("Abaixo disso: Crítico", None, VERM, TVERM)]):
    rr = 5 + i
    c = cd.cell(row=rr, column=7, value=rot); c.font = fnt(9, b=True); c.border = BOX
    c.fill = PatternFill("solid", fgColor=cor); c.font = Font(name=F, size=9, bold=True, color=tcor)
    if val is not None:
        v = cd.cell(row=rr, column=8, value=val)
        v.number_format = "0.0"; v.fill = PatternFill("solid", fgColor=C_IN)
        v.font = fnt(10, b=True, color=BLUE); v.border = BOX
        v.alignment = Alignment(horizontal="center")
cd["G9"] = "Ajuste os cortes aqui. Todos os faróis da planilha leem destas duas células."
cd["G9"].font = fnt(9, it=True, color="595959")

cd["J4"] = "LISTAS"; cd["J4"].font = fnt(11, b=True, color=C_TIT)
for col, head, vals in [("J", "Avaliador", AVALIADORES), ("K", "Tipo de ocorrência", TIPOS_OC),
                        ("L", "Notas", NOTAS)]:
    c = cd[col + "5"]; c.value = head
    c.font = Font(name=F, size=9, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_HEAD); c.border = BOX
    c.alignment = Alignment(horizontal="center")
    for i, v in enumerate(vals):
        cc = cd[col + str(6 + i)]; cc.value = v
        cc.fill = PatternFill("solid", fgColor=C_IN); cc.font = fnt(9, color=BLUE); cc.border = BOX
        cc.alignment = Alignment(horizontal="center")

cd["A{}".format(LAST_CAD + 3)] = "O QUE CADA NOTA SIGNIFICA"
cd["A{}".format(LAST_CAD + 3)].font = fnt(11, b=True, color=C_TIT)
for i, (n, txt) in enumerate([
    (5, "Referência. Vira exemplo para o time."),
    (4, "Acima do esperado, de forma consistente."),
    (3, "Fez o esperado para o nível dele. É a nota saudável, não é nota ruim."),
    (2, "Abaixo do esperado. Precisou de suporte acima do normal."),
    (1, "Bem abaixo. Gerou retrabalho ou dano.")]):
    rr = LAST_CAD + 4 + i
    c = cd.cell(row=rr, column=1, value=n)
    c.font = fnt(11, b=True); c.alignment = Alignment(horizontal="center"); c.border = BOX
    c.fill = PatternFill("solid", fgColor=VERDE if n >= 4 else (AMAR if n == 3 else VERM))
    d = cd.cell(row=rr, column=2, value=txt); d.font = fnt(9)
    cd.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=5)
cd["A{}".format(LAST_CAD + 10)] = ("A nota é sempre relativa ao nível da pessoa: um Júnior 3 e um Sênior 3 "
                                   "são comportamentos diferentes, mas os dois estão fazendo o esperado.")
cd["A{}".format(LAST_CAD + 10)].font = fnt(9, it=True, color="C00000")
cd.merge_cells(start_row=LAST_CAD + 10, start_column=1, end_row=LAST_CAD + 10, end_column=5)

DV_PESSOA = "=Cadastro!$A$5:$A${}".format(LAST_CAD)
DV_AVALIADOR = "=Cadastro!$J$6:$J$7"
DV_TIPO = "=Cadastro!$K$6:$K$8"
DV_NOTA = "=Cadastro!$L$6:$L$10"
def dv(ws, formula, rng):
    d = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(d); d.add(rng)

# =====================================================================
# AVALIAÇÃO SEMANAL — a aba de preenchimento
# =====================================================================
av = sh("Avaliação Semanal")
titulo(av, "Avaliação semanal",
       "Sexta-feira, dois minutos por pessoa. GP e Líder Técnico preenchem cada um a sua linha. Notas de 1 a 5 — a legenda está no Cadastro.")
av["A3"] = ("Não sabe avaliar um indicador nesta semana? Deixe em branco. Em branco sai da média — chutar 3 estraga o balizador."
            "   ·   O comentário da última coluna é o que mais serve ao gestor depois.")
av["A3"].font = fnt(9, it=True, color="C00000")
COLS_AV = ["Semana", "Colaborador", "Avaliador", "Quem avaliou"] + IND + ["Média", "Farol", "Comentários da semana", "chave"]
larg(av, {"A": 12, "B": 20, "C": 14, "D": 16, "E": 10, "F": 10, "G": 13, "H": 12, "I": 11,
          "J": 9, "K": 12, "L": 78, "M": 26})
cab(av, 4, COLS_AV, alturas=38)
for i, ind in enumerate(IND):
    av.cell(row=4, column=5 + i).comment = Comment(IND_DESC[i], "legenda")
for i, (data, nome, aval, quem, notas, com) in enumerate(AVAL):
    rr = 5 + i
    av.cell(row=rr, column=1, value=data)
    av.cell(row=rr, column=2, value=nome)
    av.cell(row=rr, column=3, value=aval)
    av.cell(row=rr, column=4, value=quem)
    for j, n in enumerate(notas): av.cell(row=rr, column=5 + j, value=n)
    av.cell(row=rr, column=12, value=com)
for rr in range(5, LAST_AV + 1):
    av.cell(row=rr, column=10, value='=IF(COUNT($E{0}:$I{0})=0,"",ROUND(AVERAGE($E{0}:$I{0}),2))'.format(rr))
    av.cell(row=rr, column=11, value=(
        '=IF($J{0}="","",IF($J{0}>=Cadastro!$H$5,"Verde",IF($J{0}>=Cadastro!$H$6,"Atenção","Crítico")))').format(rr))
    av.cell(row=rr, column=13, value='=IF($B{0}="","",$B{0}&"|"&$C{0}&"|"&TEXT($A{0},"00000"))'.format(rr))
    av["A" + str(rr)].number_format = "DD/MM/YYYY"
    av["J" + str(rr)].number_format = "0.00"
    for col in ("A", "C", "E", "F", "G", "H", "I", "J", "K"):
        av[col + str(rr)].alignment = Alignment(horizontal="center", vertical="center")
pinta(av, "A5:I{}".format(LAST_AV), "in")
pinta(av, "J5:K{}".format(LAST_AV), "calc")
pinta(av, "L5:L{}".format(LAST_AV), "in")
pinta(av, "M5:M{}".format(LAST_AV), "calc")
dv(av, DV_PESSOA, "B5:B{}".format(LAST_AV))
dv(av, DV_AVALIADOR, "C5:C{}".format(LAST_AV))
for i in range(5):
    dv(av, DV_NOTA, "{0}5:{0}{1}".format(chr(ord("E") + i), LAST_AV))
for txt, fill, tcor in [("Verde", VERDE, TVERDE), ("Atenção", AMAR, TAMAR), ("Crítico", VERM, TVERM)]:
    av.conditional_formatting.add("K5:K{}".format(LAST_AV), CellIsRule(
        operator="equal", formula=['"{}"'.format(txt)],
        fill=PatternFill("solid", fgColor=fill), font=Font(name=F, size=9, bold=True, color=tcor)))
av.conditional_formatting.add("E5:I{}".format(LAST_AV), CellIsRule(
    operator="lessThanOrEqual", formula=["2"], fill=PatternFill("solid", fgColor=VERM),
    font=Font(name=F, size=9, bold=True, color=TVERM)))
av.conditional_formatting.add("E5:I{}".format(LAST_AV), CellIsRule(
    operator="greaterThanOrEqual", formula=["5"], fill=PatternFill("solid", fgColor=VERDE),
    font=Font(name=F, size=9, bold=True, color=TVERDE)))
av.column_dimensions["M"].hidden = True
av.freeze_panes = "C5"
av.auto_filter.ref = "A4:L{}".format(LAST_AV)

# =====================================================================
# OCORRÊNCIAS
# =====================================================================
oc = sh("Ocorrências")
titulo(oc, "Ocorrências",
       "O que aconteceu no dia a dia e vale registrar. Não é só o negativo — registre o que foi bom também.")
oc["A3"] = ("Escreva o FATO, não o julgamento: \"subiu hotfix sem aprovação na sexta 18h\", não \"é descuidado\".   ·   "
            "Nada de saúde, vida pessoal ou conduta — conduta vai direto para o RH.")
oc["A3"].font = fnt(9, it=True, color="C00000")
larg(oc, {"A": 12, "B": 20, "C": 18, "D": 18, "E": 96})
cab(oc, 4, ["Data", "Colaborador", "Tipo", "Registrado por", "O que aconteceu"])
for i, o in enumerate(OCORR):
    for j, v in enumerate(o): oc.cell(row=5 + i, column=1 + j, value=v)
for rr in range(5, LAST_OC + 1):
    oc["A" + str(rr)].number_format = "DD/MM/YYYY"
    for col in ("A", "C"): oc[col + str(rr)].alignment = Alignment(horizontal="center", vertical="center")
pinta(oc, "A5:E{}".format(LAST_OC), "in")
dv(oc, DV_PESSOA, "B5:B{}".format(LAST_OC))
dv(oc, DV_TIPO, "C5:C{}".format(LAST_OC))
for txt, fill, tcor in [("Positiva", VERDE, TVERDE), ("Ponto de atenção", AMAR, TAMAR), ("Crítica", VERM, TVERM)]:
    oc.conditional_formatting.add("C5:C{}".format(LAST_OC), CellIsRule(
        operator="equal", formula=['"{}"'.format(txt)],
        fill=PatternFill("solid", fgColor=fill), font=Font(name=F, size=9, bold=True, color=tcor)))
oc.freeze_panes = "C5"
oc.auto_filter.ref = "A4:E{}".format(LAST_OC)

# =====================================================================
# RESUMO — a aba compilada
# =====================================================================
rs = sh("Resumo")
titulo(rs, "Resumo por pessoa",
       "Tudo compilado. Escolha o período abaixo e a tabela inteira se ajusta.")
rs["A3"] = "Período de"; rs["A3"].font = fnt(10, b=True)
rs["B3"] = dt.date(2026, 8, 1)
rs["C3"] = "até"; rs["C3"].font = fnt(10, b=True); rs["C3"].alignment = Alignment(horizontal="center")
rs["D3"] = dt.date(2026, 8, 31)
for cel in ("B3", "D3"):
    rs[cel].number_format = "DD/MM/YYYY"
    rs[cel].fill = PatternFill("solid", fgColor=C_IN)
    rs[cel].font = fnt(10, b=True, color=BLUE); rs[cel].border = BOX
    rs[cel].alignment = Alignment(horizontal="center")
rs["E3"] = "Deixe um período largo para ver o acumulado, ou aperte para uma semana só."
rs["E3"].font = fnt(9, it=True, color="595959")

COLS_RS = ["Colaborador", "Squad", "Papel", "Média GP", "Média LT"] + IND + \
          ["MÉDIA FINAL", "Farol", "Avaliações", "Última avaliação", "Ocorrências",
           "Atenção / Críticas", "Último comentário do GP", "Último comentário do Líder Técnico"]
larg(rs, {"A": 20, "B": 10, "C": 12, "D": 10, "E": 10, "F": 10, "G": 11, "H": 13, "I": 12, "J": 11,
          "K": 13, "L": 12, "M": 11, "N": 14, "O": 12, "P": 15, "Q": 62, "R": 62})
cab(rs, 5, COLS_RS, alturas=42)
A_ = "'Avaliação Semanal'"
O_ = "'Ocorrências'"
for i in range(LAST_RES - 5):
    rr = 6 + i
    src = 5 + i
    rs.cell(row=rr, column=1, value='=IF(Cadastro!$A{0}="","",Cadastro!$A{0})'.format(src))
    rs.cell(row=rr, column=2, value='=IF(Cadastro!$A{0}="","",Cadastro!$B{0})'.format(src))
    rs.cell(row=rr, column=3, value='=IF(Cadastro!$A{0}="","",Cadastro!$C{0})'.format(src))
    for col, quem in [(4, "GP"), (5, "Líder Técnico")]:
        rs.cell(row=rr, column=col, value=(
            '=IF($A{0}="","",IF(COUNTIFS({1}!$B$5:$B${2},$A{0},{1}!$C$5:$C${2},"{3}",'
            '{1}!$A$5:$A${2},">="&$B$3,{1}!$A$5:$A${2},"<="&$D$3,{1}!$J$5:$J${2},">0")=0,"",'
            'ROUND(AVERAGEIFS({1}!$J$5:$J${2},{1}!$B$5:$B${2},$A{0},{1}!$C$5:$C${2},"{3}",'
            '{1}!$A$5:$A${2},">="&$B$3,{1}!$A$5:$A${2},"<="&$D$3),2)))').format(rr, A_, LAST_AV, quem))
    for j in range(5):
        colL = chr(ord("E") + j)
        rs.cell(row=rr, column=6 + j, value=(
            '=IF($A{0}="","",IF(COUNTIFS({1}!$B$5:$B${2},$A{0},{1}!$A$5:$A${2},">="&$B$3,'
            '{1}!$A$5:$A${2},"<="&$D$3,{1}!${3}$5:${3}${2},">0")=0,"",'
            'ROUND(AVERAGEIFS({1}!${3}$5:${3}${2},{1}!$B$5:$B${2},$A{0},'
            '{1}!$A$5:$A${2},">="&$B$3,{1}!$A$5:$A${2},"<="&$D$3),2)))').format(rr, A_, LAST_AV, colL))
    rs.cell(row=rr, column=11, value='=IF(COUNT($F{0}:$J{0})=0,"",ROUND(AVERAGE($F{0}:$J{0}),2))'.format(rr))
    rs.cell(row=rr, column=12, value=(
        '=IF($K{0}="","",IF($K{0}>=Cadastro!$H$5,"Verde",IF($K{0}>=Cadastro!$H$6,"Atenção","Crítico")))').format(rr))
    rs.cell(row=rr, column=13, value=(
        '=IF($A{0}="","",COUNTIFS({1}!$B$5:$B${2},$A{0},{1}!$A$5:$A${2},">="&$B$3,'
        '{1}!$A$5:$A${2},"<="&$D$3,{1}!$J$5:$J${2},">0"))').format(rr, A_, LAST_AV))
    rs.cell(row=rr, column=14, value=(
        '=IF($M{0}=0,"",_xlfn.MAXIFS({1}!$A$5:$A${2},{1}!$B$5:$B${2},$A{0},'
        '{1}!$A$5:$A${2},"<="&$D$3))').format(rr, A_, LAST_AV))
    rs.cell(row=rr, column=15, value=(
        '=IF($A{0}="","",COUNTIFS({1}!$B$5:$B${2},$A{0},{1}!$A$5:$A${2},">="&$B$3,'
        '{1}!$A$5:$A${2},"<="&$D$3))').format(rr, O_, LAST_OC))
    rs.cell(row=rr, column=16, value=(
        '=IF($A{0}="","",COUNTIFS({1}!$B$5:$B${2},$A{0},{1}!$C$5:$C${2},"Ponto de atenção",'
        '{1}!$A$5:$A${2},">="&$B$3,{1}!$A$5:$A${2},"<="&$D$3)'
        '+COUNTIFS({1}!$B$5:$B${2},$A{0},{1}!$C$5:$C${2},"Crítica",'
        '{1}!$A$5:$A${2},">="&$B$3,{1}!$A$5:$A${2},"<="&$D$3))').format(rr, O_, LAST_OC))
    for col, quem in [(17, "GP"), (18, "Líder Técnico")]:
        rs.cell(row=rr, column=col, value=(
            '=IF($A{0}="","",IF(COUNTIFS({1}!$B$5:$B${2},$A{0},{1}!$C$5:$C${2},"{3}",'
            '{1}!$A$5:$A${2},">="&$B$3,{1}!$A$5:$A${2},"<="&$D$3)=0,"",'
            'IFERROR(INDEX({1}!$L$5:$L${2},MATCH($A{0}&"|{3}|"&TEXT(_xlfn.MAXIFS({1}!$A$5:$A${2},'
            '{1}!$B$5:$B${2},$A{0},{1}!$C$5:$C${2},"{3}",{1}!$A$5:$A${2},"<="&$D$3),"00000"),'
            '{1}!$M$5:$M${2},0)),"")))').format(rr, A_, LAST_AV, quem))
    for col in ("D", "E", "F", "G", "H", "I", "J", "K"):
        rs[col + str(rr)].number_format = "0.00"
        rs[col + str(rr)].alignment = Alignment(horizontal="center", vertical="center")
    for col in ("B", "C", "L", "M", "N", "O", "P"):
        rs[col + str(rr)].alignment = Alignment(horizontal="center", vertical="center")
    rs["N" + str(rr)].number_format = "DD/MM/YYYY"
    rs.row_dimensions[rr].height = 30
pinta(rs, "A6:R{}".format(LAST_RES), "calc")
for col in ("K", "L"):
    for rr in range(6, LAST_RES + 1):
        rs[col + str(rr)].font = fnt(11, b=True)
for txt, fill, tcor in [("Verde", VERDE, TVERDE), ("Atenção", AMAR, TAMAR), ("Crítico", VERM, TVERM)]:
    rs.conditional_formatting.add("K6:L{}".format(LAST_RES), FormulaRule(
        formula=['$L6="{}"'.format(txt)],
        fill=PatternFill("solid", fgColor=fill), font=Font(name=F, size=11, bold=True, color=tcor)))
rs.conditional_formatting.add("D6:J{}".format(LAST_RES), CellIsRule(
    operator="lessThan", formula=["3"], fill=PatternFill("solid", fgColor=VERM),
    font=Font(name=F, size=9, bold=True, color=TVERM)))
rs.conditional_formatting.add("D6:J{}".format(LAST_RES), CellIsRule(
    operator="greaterThanOrEqual", formula=["4"], fill=PatternFill("solid", fgColor=VERDE),
    font=Font(name=F, size=9, color=TVERDE)))
rs.conditional_formatting.add("P6:P{}".format(LAST_RES), CellIsRule(
    operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=AMAR),
    font=Font(name=F, size=9, bold=True, color=TAMAR)))
rs.conditional_formatting.add("M6:M{}".format(LAST_RES), CellIsRule(
    operator="equal", formula=["0"], fill=PatternFill("solid", fgColor=VERM),
    font=Font(name=F, size=9, bold=True, color=TVERM)))
rs["D5"].comment = Comment(
 "Média do GP e média do Líder Técnico lado a lado, de propósito.\n\n"
 "Quando as duas divergem muito, isso é informação — não erro. Costuma significar que a pessoa "
 "vai bem em uma dimensão e mal em outra.", "leitura")
rs["K5"].comment = Comment("Média dos cinco indicadores no período. É o balizador, não a sentença.", "leitura")
rs["M5"].comment = Comment("Zero avaliações no período fica vermelho: o problema aí é a falta de registro, não a pessoa.", "leitura")
rs.freeze_panes = "B6"
rs.auto_filter.ref = "A5:R{}".format(LAST_RES)
rs.sheet_view.zoomScale = 90

del wb["Sheet"]
wb._sheets = [wb["Resumo"], wb["Avaliação Semanal"], wb["Ocorrências"], wb["Cadastro"]]
wb.active = 0

# ---- validador estático ----
def desbal(f):
    d, ins = 0, False
    for ch in f:
        if ch == '"': ins = not ins
        elif not ins:
            if ch == "(": d += 1
            elif ch == ")": d -= 1
            if d < 0: return True
    return d != 0
PROIB = ["XLOOKUP", "FILTER(", "UNIQUE(", "SORT(", "SEQUENCE(", "SUMPRODUCT("]
SEMPRE = ["MAXIFS(", "MINIFS(", "TEXTJOIN(", "IFS("]
erros, abas, n_today, n_f = [], set(wb.sheetnames), 0, 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            f = c.value
            if not isinstance(f, str) or not f.startswith("="): continue
            n_f += 1; up = f.upper(); loc = "{}!{}".format(ws.title, c.coordinate)
            if desbal(f): erros.append(loc + ": parênteses desbalanceados")
            for m in re.finditer(r'(?:^|[^0-9A-Z$])\$?([A-Z]{1,3}):\$?([A-Z]{1,3})(?![0-9])', f):
                erros.append(loc + ": coluna inteira")
            for p in PROIB:
                if p in up: erros.append(loc + ": função proibida " + p)
            if re.search(r'INDEX\(\s*\(', f): erros.append(loc + ": INDEX sobre expressão")
            for fn in SEMPRE:
                for m in re.finditer(re.escape(fn), up):
                    if m.start() > 0 and up[m.start() - 1].isalpha(): continue
                    if not up[max(0, m.start() - 6):m.start()].endswith("_XLFN."):
                        erros.append(loc + ": " + fn + " sem _xlfn.")
            if "_XLFN._XLFN." in up: erros.append(loc + ": prefixo duplicado")
            n_today += len(re.findall(r'\bTODAY\(\)', f))
            for m in re.finditer(r"'([^']+)'!", f):
                if m.group(1) not in abas: erros.append(loc + ": aba '" + m.group(1) + "'")
            for m in re.finditer(r"(?:^|[^A-Za-z0-9_'!])([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ0-9_]*)!", f):
                if m.group(1) not in abas: erros.append(loc + ": aba '" + m.group(1) + "'")
if erros:
    print("BUILD REPROVADO — {} problema(s):".format(len(erros)))
    for e in sorted(set(erros))[:25]: print("  ·", e)
    sys.exit(1)
wb.save(OUT)
print("OK  {}".format(OUT))
print("    {} fórmulas · {} abas · validador: 0 problemas".format(n_f, len(wb.sheetnames)))

# ---- modelo-sombra ----
import json, collections
soma = collections.defaultdict(list)
for data, nome, aval, quem, notas, com in AVAL:
    for j, n in enumerate(notas): soma[(nome, IND[j])].append(n)
    soma[(nome, "__" + aval)].append(round(sum(notas) / len(notas), 2))
ESP = {}
linha = {n: 6 + i for i, n in enumerate(NOMES)}
for nome in NOMES:
    inds = [round(sum(soma[(nome, ind)]) / len(soma[(nome, ind)]), 2) for ind in IND]
    for j, col in enumerate("FGHIJ"):
        ESP["Resumo!{}{}".format(col, linha[nome])] = inds[j]
    ESP["Resumo!D{}".format(linha[nome])] = round(sum(soma[(nome, "__GP")]) / len(soma[(nome, "__GP")]), 2)
    ESP["Resumo!E{}".format(linha[nome])] = round(sum(soma[(nome, "__Líder Técnico")]) / len(soma[(nome, "__Líder Técnico")]), 2)
    media = round(sum(inds) / len(inds), 2)
    ESP["Resumo!K{}".format(linha[nome])] = media
    ESP["Resumo!L{}".format(linha[nome])] = "Verde" if media >= 4 else ("Atenção" if media >= 3 else "Crítico")
    ESP["Resumo!M{}".format(linha[nome])] = sum(1 for a in AVAL if a[1] == nome)
    ESP["Resumo!O{}".format(linha[nome])] = sum(1 for o in OCORR if o[1] == nome)
with open("/home/user/kaneo/docs/gestao-pessoas/.sombra_simples.json", "w") as fh:
    json.dump(ESP, fh, indent=1, ensure_ascii=False)
print("    modelo-sombra: {} células esperadas".format(len(ESP)))
for nome in NOMES:
    r = linha[nome]
    print("      {:16s} média {:.2f}  farol {}".format(nome, ESP["Resumo!K%d" % r], ESP["Resumo!L%d" % r]))
