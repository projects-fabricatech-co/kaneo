# -*- coding: utf-8 -*-
"""Gerador da planilha de Avaliacao de Desempenho por Squad — SPEC-v3.
Inclui validador estatico (falha o build) e modelo-sombra para conferencia de valores."""
import re, sys, datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.comments import Comment

OUT = "/home/user/kaneo/docs/gestao-pessoas/Avaliacao_Desempenho_Squad_v3.xlsx"
F = "Arial"
C_TIT, C_HEAD, C_SUB = "1F3864", "2E5395", "D9E2F3"
C_IN, C_CALC, C_WARN, C_BAD, C_GOOD = "FFF2CC", "F2F2F2", "FDE9D9", "FFC7CE", "C6EFCE"
BLUE = "0000FF"
thin = Side(style="thin", color="BFBFBF"); BOX = Border(thin, thin, thin, thin)

# ---- abas (constante unica por aba, sempre com aspas simples) ----
SH_CFG, SH_PES, SH_ALO = "Config", "'Pessoas'", "'Alocações'"
SH_OC, SH_AV, SH_PA = "'Ocorrências'", "'Avaliação'", "'Painel'"
SH_PDI, SH_1A1, SH_AL = "'PDI'", "'1a1'", "'Alertas'"

LAST_PES, LAST_ALO, LAST_OC, LAST_AV, LAST_PDI, LAST_PA = 64, 84, 604, 404, 104, 64
LAST_1A1 = 111

def fnt(sz=10, b=False, color="000000", it=False):
    return Font(name=F, size=sz, bold=b, color=color, italic=it)

wb = openpyxl.Workbook()
wb.calculation.fullCalcOnLoad = True

def sh(name):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False; return ws

def titulo(ws, t, sub=None):
    ws["A1"] = t; ws["A1"].font = Font(name=F, size=14, bold=True, color=C_TIT)
    ws.row_dimensions[1].height = 22
    if sub:
        ws["A2"] = sub; ws["A2"].font = Font(name=F, size=9, italic=True, color="595959")

def cab(ws, row, vals, start=1):
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=start + i, value=v)
        c.font = Font(name=F, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD)
        c.alignment = Alignment(vertical="center", wrap_text=True); c.border = BOX
    ws.row_dimensions[row].height = 32

def larg(ws, spec):
    for k, v in spec.items(): ws.column_dimensions[k].width = v

def faixa(ws, row, txt, span=10):
    c = ws.cell(row=row, column=1, value=txt)
    c.font = Font(name=F, size=10, bold=True, color=C_TIT)
    for j in range(1, span + 1):
        ws.cell(row=row, column=j).fill = PatternFill("solid", fgColor=C_SUB)

def pinta(ws, rng, kind):
    fill = PatternFill("solid", fgColor={"in": C_IN, "calc": C_CALC}[kind])
    for r in ws[rng]:
        for c in r:
            c.fill = fill; c.border = BOX
            c.font = fnt(9, color=BLUE if kind == "in" else "000000")
            c.alignment = Alignment(vertical="top", wrap_text=True)

# =====================================================================
# MODELO
# =====================================================================
EA, EB, EC, ED = "A — Entrega", "B — Técnico", "C — Colaboração", "D — Autonomia"
EIXOS = [EA, EB, EC, ED]
PESO = {EA: 0.35, EB: 0.30, EC: 0.20, ED: 0.15}

COMP = [
 ("C01", EA, "Previsibilidade", "GP", [
  "Prazos estouram sem aviso; o time descobre na review. Compromisso assumido não é cumprido nem renegociado.",
  "Entrega, mas quase sempre depois; o aviso de atraso vem tarde demais para o PO replanejar.",
  "Cumpre o combinado na maioria das sprints; quando vai atrasar, avisa a tempo de o time reagir, com nova data.",
  "Estimativa confiável inclusive em item ambíguo; identifica risco antes de começar e negocia escopo em vez de estourar.",
  "É a referência de previsibilidade da squad; o roadmap é planejado sobre a estimativa dela sem colchão."]),
 ("C02", EA, "Qualidade da entrega", "Líder Técnico", [
  "Entrega volta de QA ou de produção com frequência; deixa retrabalho recorrente para o próximo.",
  "Cobre o caminho feliz; falha em erro, borda e dado real. PR volta 2+ rodadas com frequência.",
  "PR passa em 1–2 rodadas; teste junto com o código, cobrindo erro e borda. Reprovação em QA é exceção.",
  "Escolhe o nível certo de teste e testa o risco; seus testes já pegaram regressão de terceiros.",
  "Melhorou a testabilidade do sistema: reduziu flaky, criou fixture ou harness que o time inteiro usa."]),
 ("C03", EB, "Domínio técnico", "Líder Técnico", [
  "Erros conceituais recorrentes na stack principal; a mesma explicação precisa ser repetida a cada sprint.",
  "Resolve o caminho feliz; trava em erro, concorrência, performance ou dado real e depende de terceiro.",
  "Resolve sozinho problemas típicos do produto no nível dele; sabe onde o sistema quebra e onde procurar o resto.",
  "Resolve problema fora da zona conhecida; explica trade-off de custo, risco e prazo, e sustenta a escolha com dado.",
  "Referência técnica consultada por outras squads; decisões dele viram padrão ou ADR na empresa."]),
 ("C04", EB, "Design de solução", "Líder Técnico", [
  "Cria estrutura nova sem perguntar e ignora fronteiras existentes; gera retrabalho.",
  "Copia padrão sem entender; escolhe por familiaridade, não por trade-off.",
  "Escolhe entre alternativas conhecidas e sabe dizer o que ganha e o que perde.",
  "Antecipa acoplamento e custo de reverter; escreve a decisão antes de codar.",
  "Decisão dele mudou o rumo de mais de um time e envelheceu bem."]),
 ("C05", EC, "Comunicação e colaboração", "Compartilhado", [
  "Status enganoso ou ausente; aprova PR sem ler; discordância técnica vira questão de ego.",
  "Comunica, mas exige garimpo; mistura nit com bloqueante e trava PR por gosto pessoal.",
  "Card, PR e mensagem que quem não estava no contexto entende; revisa em até 1 dia útil e pega problema real.",
  "Adapta a mensagem ao público; o comentário de review ensina e propõe alternativa sem assumir o teclado.",
  "A comunicação dela destrava decisão de outros times; é referência de review e medeia discordância sem escalar."]),
 ("C06", ED, "Autonomia e ownership", "Compartilhado", [
  "Só avança com direcionamento passo a passo; entrega a tarefa e ignora que o problema continua de pé.",
  "Precisa de acompanhamento acima do normal para o nível; escala tarde ou escala tudo, sem tentativa própria.",
  "Toca as próprias demandas com direcionamento normal; escala no momento certo, com hipótese e opções.",
  "Assume problema mal definido e devolve solução; puxa as pessoas necessárias sem precisar do gestor no meio.",
  "Assume tema crítico de ponta a ponta, técnico e stakeholder, e o gestor deixa de precisar acompanhar."]),
]
COMP_NOMES = [c[2] for c in COMP]
SENIOR = [("Júnior", 2.7), ("Pleno", 3.0), ("Sênior", 3.2), ("Especialista", 3.4)]
PAPEIS = ["Dev/QA", "Líder Técnico", "PO/BA", "UX"]
SQUADS = ["Alfa", "Beta", "Delta"]
CONTRATOS = ["Contrato A", "Contrato B", "Contrato C"]
TIPOS = ["Reconhecimento", "Positiva", "Ponto de atenção", "Incidente",
         "Ausência / Disponibilidade", "Contexto atenuante"]
NEGATIVOS = ["Ponto de atenção", "Incidente"]
IMPACTO = ["Cliente", "Equipe", "Entrega", "Nenhum"]
CONF = ["Squad", "Gestão"]
SIMNAO = ["Sim", "Não", "Parcial"]
AUTOR = ["GP", "Líder Técnico"]
BASEOBS = ["Diária", "Cerimônias", "Esporádica", "Insuficiente"]
REACAO = ["Reconheceu", "Discordou", "Trouxe contexto"]
PRAZOS = ["7 dias", "14 dias", "1 ciclo"]
STPDI = ["Não iniciado", "Em andamento", "Concluído", "Cancelado"]
NOTAS = ["1", "2", "3", "4", "5", "N/O"]
CICLOS = ["2026-Q1", "2026-Q2", "2026-Q3", "2026-Q4", "2027-Q1", "2027-Q2"]
CICLO = "2026-Q3"

PESSOAS = [
 ("P001", "Ana Ribeiro", "M-1042", "Pleno", dt.date(2024,2,5), dt.date(2025,7,1), "Marcos Lima", "CLT"),
 ("P002", "Bruno Tavares", "M-1188", "Júnior", dt.date(2026,6,15), dt.date(2026,6,15), "Marcos Lima", "CLT"),
 ("P003", "Carla Nunes", "M-0977", "Sênior", dt.date(2022,9,12), dt.date(2024,3,1), "Marcos Lima", "PJ"),
]
ALOCS = [
 ("A001","P001","Alfa","Contrato A","Portal do Cliente","Dev/QA",1.00,dt.date(2025,3,10),None,"Marcos Lima","Rafael Souza"),
 ("A002","P002","Alfa","Contrato A","Portal do Cliente","Dev/QA",1.00,dt.date(2026,6,15),None,"Marcos Lima","Rafael Souza"),
 ("A003","P003","Beta","Contrato B","Motor de Cobrança","Líder Técnico",0.50,dt.date(2024,8,1),None,"Marcos Lima","Carla Nunes"),
 ("A004","P003","Delta","Contrato C","Faturamento","Dev/QA",0.50,dt.date(2026,4,1),None,"Paula Reis","Rafael Souza"),
]
# notas do exemplo (Ana / A001): comp -> (GP, LT, final)
EX = {"Previsibilidade": (3, None, 3), "Qualidade da entrega": (None, 2, 2),
      "Domínio técnico": (None, 3, 3), "Design de solução": (None, 3, 3),
      "Comunicação e colaboração": (4, 5, 5), "Autonomia e ownership": (3, 4, 4)}

OCORR = [
 (dt.date(2026,7,9), "Ana Ribeiro", "Rafael Souza", "Líder Técnico", "Positiva", "Comunicação e colaboração",
  "No PR #812 apontou uma condição de corrida no resgate de cupom antes do merge, com o trecho e a sugestão de UPDATE atômico.",
  "Entrega", "Squad", "", "", None, "", "", "", None, "", ""),
 (dt.date(2026,8,19), "Ana Ribeiro", "Rafael Souza", "Líder Técnico", "Ponto de atenção", "Qualidade da entrega",
  "PR #857 voltou em 3 rodadas: nas duas primeiras faltava teste do caminho de erro na rota de resgate.",
  "Entrega", "Squad",
  "Atrasou o merge em 2 dias e consumiu 3 rodadas de review do time.", "Sim", dt.date(2026,8,20), "Rafael Souza",
  "Incluir teste de caminho de erro antes de abrir o PR, nos próximos 3 PRs de rota nova.", "1 ciclo",
  "Reconheceu", "Em acompanhamento"),
 (dt.date(2026,7,22), "Bruno Tavares", "Marcos Lima", "GP", "Ponto de atenção", "Previsibilidade",
  "Card CB-118 ficou 4 dias sem atualização e o bloqueio, falta de credencial de homologação, só foi comunicado na review.",
  "Entrega", "Squad",
  "A sprint fechou com o item em carry over e o PO replanejou a demo com o cliente em cima da hora.", "Sim",
  dt.date(2026,7,23), "Marcos Lima", "Comunicar bloqueio na daily do mesmo dia em que ele ocorre.", "14 dias",
  "Reconheceu", "Encerrado — comportamento mudou"),
 (dt.date(2026,8,4), "Bruno Tavares", "Marcos Lima", "GP", "Positiva", "Comunicação e colaboração",
  "Na daily de 04/08 avisou no mesmo dia que a integração de pagamento estava fora do ar, com print do erro e a alternativa de mockar.",
  "Equipe", "Squad", "", "", None, "", "", "", None, "", ""),
 (dt.date(2026,8,25), "Carla Nunes", "Rafael Souza", "Líder Técnico", "Reconhecimento", "Autonomia e ownership",
  "Escreveu a ADR-07 sobre o advisory lock no cooldown e conduziu a sessão de 40 min com as squads Alfa e Beta.",
  "Equipe", "Squad", "", "", None, "", "", "", None, "", ""),
 (dt.date(2026,8,12), "Carla Nunes", "Marcos Lima", "GP", "Contexto atenuante", "Previsibilidade",
  "Assumiu o on-call do Contrato B por duas semanas cobrindo férias, além dos 50% de alocação no Motor de Cobrança.",
  "Entrega", "Squad", "", "", None, "", "", "", None, "", ""),
]

# =====================================================================
# CONFIG
# =====================================================================
cf = sh("Config")
titulo(cf, "Configuração do modelo", "Ajuste aqui uma vez. Toda a planilha lê desta aba.")
larg(cf, {"A": 30, "B": 22, "C": 52, "D": 16, "E": 16, "F": 16, "G": 16, "H": 4,
          "I": 15, "J": 15, "K": 22, "L": 12, "M": 15, "N": 10, "O": 14, "P": 13, "Q": 12,
          "R": 10, "S": 11, "T": 14})
for r, k, v, obs in [
    (4, "Ciclo ativo", CICLO, "Trimestre em avaliação. Formato AAAA-Qn."),
    (5, "Ciclo inicial do uso", CICLO, "Alertas que dependem de histórico ficam suprimidos até haver ciclos suficientes."),
]:
    cf.cell(row=r, column=1, value=k).font = fnt(10, b=True)
    c = cf.cell(row=r, column=2, value=v)
    c.fill = PatternFill("solid", fgColor=C_IN); c.font = fnt(10, b=True, color=BLUE); c.border = BOX
    cf.cell(row=r, column=3, value=obs).font = fnt(9, it=True, color="595959")
cf["A6"] = "Hoje"; cf["A6"].font = fnt(10, b=True)
cf["B6"] = "=TODAY()"; cf["B6"].number_format = "DD/MM/YYYY"
cf["C6"] = "Único TODAY() do arquivo. Toda a planilha lê desta célula — volátil por linha suja a cadeia inteira a cada digitação."
cf["C6"].font = fnt(9, it=True, color="595959")
cf["A7"] = "Índice do ciclo ativo"; cf["A7"].font = fnt(10, b=True)
cf["B7"] = '=VALUE(LEFT($B$4,4))*4+VALUE(RIGHT($B$4,1))'
cf["C7"] = "Inteiro. Nenhuma comparação temporal usa o texto do ciclo."
cf["C7"].font = fnt(9, it=True, color="595959")
cf["A8"] = "Índice do ciclo inicial"; cf["A8"].font = fnt(10, b=True)
cf["B8"] = '=VALUE(LEFT($B$5,4))*4+VALUE(RIGHT($B$5,1))'
cf["A9"] = "Ciclos decorridos"; cf["A9"].font = fnt(10, b=True)
cf["B9"] = '=$B$7-$B$8+1'
for r in (6, 7, 8, 9):
    cf.cell(row=r, column=2).fill = PatternFill("solid", fgColor=C_CALC)
    cf.cell(row=r, column=2).border = BOX; cf.cell(row=r, column=2).font = fnt(10, b=True)

faixa(cf, 11, "ESCALA — 1 a 5, sem meio ponto, com N/O obrigatório", 3)
cab(cf, 12, ["Nota", "Rótulo", "O que significa"])
for i, (n, rot, sig) in enumerate([
    (1, "Abaixo crítico", "Gera dano ou retrabalho recorrente. Exige plano de ação formal."),
    (2, "Abaixo do esperado", "Precisa de suporte acima do normal para o nível."),
    (3, "Atende", "Faz o esperado da senioridade, de forma consistente. É a nota saudável, não é nota ruim."),
    (4, "Supera", "Entrega acima do nível em situações reais e repetidas."),
    (5, "Referência", "Vira padrão para o time; outros copiam a prática."),
    ("N/O", "Não observado", "Sai do cálculo. Deixar em branco é resposta legítima e melhor que chutar.")]):
    rr = 13 + i
    cf.cell(row=rr, column=1, value=n).font = fnt(9, b=True)
    cf.cell(row=rr, column=2, value=rot).font = fnt(9)
    cf.cell(row=rr, column=3, value=sig).font = fnt(9)
    for j in (1, 2, 3):
        cf.cell(row=rr, column=j).border = BOX
        cf.cell(row=rr, column=j).alignment = Alignment(wrap_text=True, vertical="top")
    cf.row_dimensions[rr].height = 22

faixa(cf, 20, "PESOS POR EIXO — únicos. Papel é contexto exibido, nunca multiplicador.", 3)
cab(cf, 21, ["Eixo", "Peso", "Por que o papel não entra"])
for i, e in enumerate(EIXOS):
    rr = 22 + i
    cf.cell(row=rr, column=1, value=e).font = fnt(9, b=True); cf.cell(row=rr, column=1).border = BOX
    c = cf.cell(row=rr, column=2, value=PESO[e]); c.number_format = "0%"
    c.fill = PatternFill("solid", fgColor=C_IN); c.font = fnt(9, color=BLUE)
    c.border = BOX; c.alignment = Alignment(horizontal="center")
cf["C22"] = "Quem é Dev num contrato e Tech Lead em outro não tem UM papel. Peso por papel produziria número diferente conforme a alocação lida primeiro — sem avisar."
cf["C22"].font = fnt(9, it=True, color="C00000"); cf["C22"].alignment = Alignment(wrap_text=True, vertical="top")
cf.merge_cells("C22:C25")
cf.cell(row=26, column=1, value="Soma (deve ser 100%)").font = fnt(9, b=True, it=True)
c = cf.cell(row=26, column=2, value='=SUM($B$22:$B$25)')
c.number_format = "0%"; c.font = fnt(9, b=True); c.border = BOX; c.alignment = Alignment(horizontal="center")
cf.conditional_formatting.add("B26", CellIsRule(operator="notEqual", formula=["1"],
                                                fill=PatternFill("solid", fgColor=C_BAD)))

faixa(cf, 28, "EXPECTATIVA POR SENIORIDADE — premissa v0, NÃO CALIBRADA. Confirme antes de decidir mérito.", 4)
cab(cf, 29, ["Senioridade", "Expectativa", "Confirmado por", "Confirmado em"])
for i, (s, v) in enumerate(SENIOR):
    rr = 30 + i
    cf.cell(row=rr, column=1, value=s).font = fnt(9, b=True); cf.cell(row=rr, column=1).border = BOX
    c = cf.cell(row=rr, column=2, value=v); c.number_format = "0.0"
    c.fill = PatternFill("solid", fgColor=C_IN); c.font = fnt(9, color=BLUE)
    c.border = BOX; c.alignment = Alignment(horizontal="center")
    for j in (3, 4):
        cc = cf.cell(row=rr, column=j); cc.fill = PatternFill("solid", fgColor=C_IN)
        cc.border = BOX; cc.font = fnt(9, color=BLUE)
    cf.cell(row=rr, column=4).number_format = "DD/MM/YYYY"
cf["A35"] = "Estes números são ponto de partida desta versão, não benchmark de mercado. Vieram do modelo, não de medição. Calibre ao fim do primeiro ciclo real e registre quem confirmou."
cf["A35"].font = fnt(9, it=True, color="C00000"); cf.merge_cells("A35:D36")
cf["A35"].alignment = Alignment(wrap_text=True, vertical="top")

faixa(cf, 38, "COMPETÊNCIAS E ÂNCORAS — a nota é sempre relativa à senioridade", 9)
cab(cf, 39, ["ID", "Eixo", "Competência", "Avaliador dono", "1 — Abaixo crítico", "2 — Abaixo",
             "3 — Atende (esperado do nível)", "4 — Supera", "5 — Referência"])
for i, (cid, eixo, nome, dono, anc) in enumerate(COMP):
    rr = 40 + i
    for j, v in enumerate([cid, eixo, nome, dono] + anc):
        c = cf.cell(row=rr, column=1 + j, value=v)
        c.font = fnt(9, b=(j == 2)); c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="top")
    cf.row_dimensions[rr].height = 54
for col, w in {"E": 40, "F": 40, "G": 42, "H": 40, "I": 40}.items():
    cf.column_dimensions[col].width = w

LISTAS = [("K", "Senioridade", [s for s, _ in SENIOR]), ("L", "Papel", PAPEIS), ("M", "Squad", SQUADS),
          ("N", "Contrato", CONTRATOS), ("O", "Tipo", TIPOS), ("P", "Impacto", IMPACTO),
          ("Q", "Confidenc.", CONF), ("R", "Sim/Não", SIMNAO), ("S", "Autor", AUTOR),
          ("T", "Base obs.", BASEOBS), ("U", "Reação", REACAO), ("V", "Prazo", PRAZOS),
          ("W", "Status PDI", STPDI), ("X", "Notas", NOTAS), ("Y", "Competências", COMP_NOMES),
          ("Z", "Ciclos", CICLOS)]
cf["K11"] = "LISTAS — alimentam os menus. Acrescente valores no fim de cada coluna."
cf["K11"].font = fnt(10, b=True, color=C_TIT)
for col, head, vals in LISTAS:
    c = cf[col + "12"]; c.value = head
    c.font = Font(name=F, size=9, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_HEAD); c.border = BOX
    c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, v in enumerate(vals):
        cc = cf[col + str(13 + i)]; cc.value = v
        cc.fill = PatternFill("solid", fgColor=C_IN); cc.font = fnt(9, color=BLUE); cc.border = BOX
    cf.column_dimensions[col].width = 20
cf.freeze_panes = "A4"

def L(col, first, n):  # intervalo exato de lista
    return "=Config!${0}${1}:${0}${2}".format(col, first, first + n - 1)

DV = {
 "senior": L("K", 13, len(SENIOR)), "papel": L("L", 13, len(PAPEIS)), "squad": L("M", 13, len(SQUADS) + 5),
 "contrato": L("N", 13, len(CONTRATOS) + 5), "tipo": L("O", 13, len(TIPOS)), "impacto": L("P", 13, len(IMPACTO)),
 "conf": L("Q", 13, len(CONF)), "simnao": L("R", 13, len(SIMNAO)), "autor": L("S", 13, len(AUTOR)),
 "baseobs": L("T", 13, len(BASEOBS)), "reacao": L("U", 13, len(REACAO)), "prazo": L("V", 13, len(PRAZOS)),
 "stpdi": L("W", 13, len(STPDI)), "nota": L("X", 13, len(NOTAS)), "comp": L("Y", 13, len(COMP)),
 "ciclo": L("Z", 13, len(CICLOS)),
 "pessoa": "=Pessoas!$B$5:$B${}".format(LAST_PES),
}
def dv(ws, key, rng):
    d = DataValidation(type="list", formula1=DV[key], allow_blank=True)
    ws.add_data_validation(d); d.add(rng)

# =====================================================================
# PESSOAS
# =====================================================================
pe = sh("Pessoas")
titulo(pe, "Cadastro de pessoas", "ID_PESSOA é texto, estável e imutável. Avaliação, PDI e histórico pertencem à PESSOA; squad e contrato são contexto do fato.")
COLS_PE = ["ID_PESSOA", "Nome", "Matrícula", "Senioridade", "Admissão", "No nível desde",
           "Gestor de Pessoas", "Contratação", "Situação", "Última 1:1", "Alocações ativas",
           "Papel considerado", "Multi-alocação", "Meses no nível", "Sem registro no ciclo", "seq"]
larg(pe, {"A": 11, "B": 21, "C": 11, "D": 13, "E": 12, "F": 13, "G": 17, "H": 12, "I": 10,
          "J": 12, "K": 11, "L": 15, "M": 13, "N": 12, "O": 12, "P": 6})
cab(pe, 4, COLS_PE)
for i, p in enumerate(PESSOAS):
    for j, v in enumerate(p): pe.cell(row=5 + i, column=1 + j, value=v)
for rr in range(5, LAST_PES + 1):
    pe.cell(row=rr, column=9, value='=IF($B{0}="","","Ativo")'.format(rr))
    pe.cell(row=rr, column=10, value=('=IF($B{0}="","",IF(COUNTIFS({1}!$B$52:$B${2},$A{0},{1}!$G$52:$G${2},"Sim")=0,"",'
                                      '_xlfn.MAXIFS({1}!$A$52:$A${2},{1}!$B$52:$B${2},$A{0},{1}!$G$52:$G${2},"Sim")))'
                                      ).format(rr, SH_1A1, LAST_1A1))
    pe.cell(row=rr, column=11, value='=IF($B{0}="","",COUNTIFS({1}!$B$5:$B${2},$A{0},{1}!$N$5:$N${2},"Sim"))'.format(rr, SH_ALO, LAST_ALO))
    pe.cell(row=rr, column=12, value=('=IF($B{0}="","",IF($K{0}=0,"sem alocação",IF($K{0}=1,'
                                      'INDEX({1}!$F$5:$F${2},MATCH($A{0},{1}!$O$5:$O${2},0)),"Multi")))'
                                      ).format(rr, SH_ALO, LAST_ALO))
    pe.cell(row=rr, column=13, value='=IF($B{0}="","",IF($K{0}>1,"Sim","Não"))'.format(rr))
    pe.cell(row=rr, column=14, value='=IF(OR($B{0}="",$F{0}=""),"",ROUND((Config!$B$6-$F{0})/30.4,0))'.format(rr))
    pe.cell(row=rr, column=15, value=('=IF($B{0}="","",IF(COUNTIFS({1}!$C$5:$C${2},$A{0},{1}!$T$5:$T${2},Config!$B$7)=0,1,0))'
                                      ).format(rr, SH_OC, LAST_OC))
    pe.cell(row=rr, column=16, value='=IF($O{0}=1,COUNTIF($O$5:$O{0},1),"")'.format(rr))
    pe.cell(row=rr, column=10).number_format = "DD/MM/YYYY"
    for col in ("E", "F"): pe[col + str(rr)].number_format = "DD/MM/YYYY"
pinta(pe, "A5:H{}".format(LAST_PES), "in")
pinta(pe, "I5:P{}".format(LAST_PES), "calc")
dv(pe, "senior", "D5:D{}".format(LAST_PES))
pe.freeze_panes = "C5"

# =====================================================================
# ALOCAÇÕES
# =====================================================================
al = sh("Alocações")
titulo(al, "Alocações — uma linha por alocação", "Quem está 50% em dois contratos ocupa duas linhas. É aqui que mora a visão por contrato: filtre por cliente e olhe a coluna Gap.")
COLS_AL = ["ID_ALOC", "ID_PESSOA", "Squad", "Cliente / Contrato", "Projeto", "Papel", "Alocação %",
           "Entrada", "Saída", "GP", "Líder Técnico", "Handover feito em", "Pontos transferidos",
           "Ativa", "chave ativa", "Pessoa", "Gap no ciclo"]
larg(al, {"A": 10, "B": 11, "C": 9, "D": 15, "E": 19, "F": 14, "G": 10, "H": 11, "I": 11,
          "J": 15, "K": 15, "L": 14, "M": 40, "N": 8, "O": 11, "P": 19, "Q": 12})
cab(al, 4, COLS_AL)
for i, a in enumerate(ALOCS):
    for j, v in enumerate(a): al.cell(row=5 + i, column=1 + j, value=v)
    al.cell(row=5 + i, column=7).number_format = "0%"
for rr in range(5, LAST_ALO + 1):
    al.cell(row=rr, column=14, value='=IF($A{0}="","",IF($I{0}="","Sim","Não"))'.format(rr))
    al.cell(row=rr, column=15, value='=IF($N{0}="Sim",$B{0},"")'.format(rr))
    al.cell(row=rr, column=16, value='=IFERROR(INDEX(Pessoas!$B$5:$B${1},MATCH($B{0},Pessoas!$A$5:$A${1},0)),"")'.format(rr, LAST_PES))
    al.cell(row=rr, column=17, value=('=IF($A{0}="","",IFERROR(INDEX({1}!$M$5:$M${2},MATCH($B{0},{1}!$A$5:$A${2},0)),""))'
                                      ).format(rr, SH_PA, LAST_PA))
    al.cell(row=rr, column=17).number_format = "0.00;-0.00;-"
    for col in ("H", "I", "L"): al[col + str(rr)].number_format = "DD/MM/YYYY"
    al["G" + str(rr)].number_format = "0%"
pinta(al, "A5:M{}".format(LAST_ALO), "in")
pinta(al, "N5:Q{}".format(LAST_ALO), "calc")
for k, rng in [("squad", "C"), ("contrato", "D"), ("papel", "F")]:
    dv(al, k, "{0}5:{0}{1}".format(rng, LAST_ALO))
al.freeze_panes = "C5"
al.auto_filter.ref = "A4:Q{}".format(LAST_ALO)

# =====================================================================
# OCORRÊNCIAS
# =====================================================================
oc = sh("Ocorrências")
titulo(oc, "Ocorrências — o registro contínuo",
       "Bloco rápido (A–J): alvo de 45 segundos, um único campo de texto. Bloco de sinalização (K–S): só para Ponto de atenção e Incidente.")
oc["A3"] = ("O FATO não aceita adjetivo de personalidade. \"Subiu hotfix sem aprovação na sexta 18h\" ✔  ·  \"É descuidado\" ✘.  "
            "Nada de saúde, vida pessoal, motivo de ausência ou conduta — conduta vai para o RH, em processo próprio.")
oc["A3"].font = fnt(9, it=True, color="C00000")
COLS_OC = ["Data do fato", "Pessoa", "ID_PESSOA", "Registrado por", "Papel do autor", "Tipo", "Competência",
           "Fato observável", "Impacto", "Confidenc.",
           "Consequência observada", "Comunicado?", "Data da comunicação", "Quem comunicou",
           "Expectativa declarada", "Prazo dado", "Reavaliação em", "Reação da pessoa", "Desfecho",
           "Idx ciclo", "chave ID|ciclo", "chave ID|ciclo|comp", "Registro válido?", "Viaja no handover?",
           "novo avaliador", "Exibição", "negativa comunicada", "seq 1a1"]
larg(oc, {"A": 12, "B": 19, "C": 11, "D": 15, "E": 13, "F": 19, "G": 22, "H": 56, "I": 11, "J": 11,
          "K": 42, "L": 12, "M": 14, "N": 15, "O": 40, "P": 12, "Q": 13, "R": 15, "S": 26,
          "T": 9, "U": 15, "V": 22, "W": 20, "X": 16, "Y": 11, "Z": 11, "AA": 12, "AB": 8})
cab(oc, 4, COLS_OC)
for j in range(1, 11): oc.cell(row=4, column=j).fill = PatternFill("solid", fgColor="1F3864")
for i, o in enumerate(OCORR):
    rr = 5 + i
    vals = list(o[:2]) + [None] + list(o[2:9]) + list(o[9:])
    for j, v in enumerate(vals):
        if v not in (None, ""): oc.cell(row=rr, column=1 + j, value=v)
for rr in range(5, LAST_OC + 1):
    oc.cell(row=rr, column=3, value='=IFERROR(INDEX(Pessoas!$A$5:$A${1},MATCH($B{0},Pessoas!$B$5:$B${1},0)),"")'.format(rr, LAST_PES))
    oc.cell(row=rr, column=17, value=('=IF(OR($M{0}="",$P{0}=""),"",$M{0}+IF($P{0}="7 dias",7,IF($P{0}="14 dias",14,90)))').format(rr))
    oc.cell(row=rr, column=20, value='=IF($A{0}="","",YEAR($A{0})*4+ROUNDUP(MONTH($A{0})/3,0))'.format(rr))
    oc.cell(row=rr, column=21, value='=IF($C{0}="","",$C{0}&"|"&$T{0})'.format(rr))
    oc.cell(row=rr, column=22, value='=IF($C{0}="","",$C{0}&"|"&$T{0}&"|"&$G{0})'.format(rr))
    oc.cell(row=rr, column=23, value=(
        '=IF($A{0}="","",IF(OR($B{0}="",$F{0}="",$G{0}="",$H{0}="",$I{0}=""),"Faltam campos básicos",'
        'IF(AND(OR($F{0}="Ponto de atenção",$F{0}="Incidente"),OR($K{0}="",$L{0}="",$O{0}="",$P{0}="")),'
        '"Faltam campos de sinalização","OK")))').format(rr))
    oc.cell(row=rr, column=24, value=(
        '=IF($A{0}="","",IF(AND(OR($F{0}="Ponto de atenção",$F{0}="Incidente"),$L{0}<>"Sim"),"Não",'
        'IF(AND($S{0}="",$W{0}<>"OK"),"Não","Sim")))').format(rr))
    oc.cell(row=rr, column=25, value='=IF($C{0}="",0,IF(COUNTIFS($C$5:$C{0},$C{0},$D$5:$D{0},$D{0})=1,1,0))'.format(rr))
    oc.cell(row=rr, column=26, value=(
        '=IF($A{0}="","",IF(Config!$B$7-$T{0}>2,"Arquivada",IF(Config!$B$7-$T{0}>0,"Histórico","Ativa")))').format(rr))
    oc.cell(row=rr, column=27, value=(
        '=IF($A{0}="",0,IF(AND(OR($F{0}="Ponto de atenção",$F{0}="Incidente"),$L{0}="Sim"),1,0))').format(rr))
    oc.cell(row=rr, column=28, value=(
        '=IF(OR($C{0}="",$C{0}<>{1}!$B$4,$T{0}<>Config!$B$7),"",COUNTIFS($C$5:$C{0},{1}!$B$4,$T$5:$T{0},Config!$B$7))'
        ).format(rr, SH_1A1))
    for col in ("A", "M", "Q"): oc[col + str(rr)].number_format = "DD/MM/YYYY"
pinta(oc, "A5:B{}".format(LAST_OC), "in")
pinta(oc, "D5:L{}".format(LAST_OC), "in")
pinta(oc, "N5:P{}".format(LAST_OC), "in")
pinta(oc, "R5:S{}".format(LAST_OC), "in")
pinta(oc, "C5:C{}".format(LAST_OC), "calc")
pinta(oc, "M5:M{}".format(LAST_OC), "in")
pinta(oc, "Q5:Q{}".format(LAST_OC), "calc")
pinta(oc, "T5:AB{}".format(LAST_OC), "calc")
for k, col in [("pessoa", "B"), ("autor", "E"), ("tipo", "F"), ("comp", "G"), ("impacto", "I"),
               ("conf", "J"), ("simnao", "L"), ("prazo", "P"), ("reacao", "R")]:
    dv(oc, k, "{0}5:{0}{1}".format(col, LAST_OC))
d = DataValidation(type="date", operator="lessThanOrEqual", formula1="TODAY()", allow_blank=True,
                   error="A data do fato não pode ser futura.", errorTitle="Data inválida")
oc.add_data_validation(d); d.add("A5:A{}".format(LAST_OC))
oc.conditional_formatting.add("A5:S{}".format(LAST_OC),
    FormulaRule(formula=['LEFT($W5,5)="Falta"'], fill=PatternFill("solid", fgColor=C_BAD)))
oc.conditional_formatting.add("K5:S{}".format(LAST_OC),
    FormulaRule(formula=['AND(OR($F5="Ponto de atenção",$F5="Incidente"),K5="")'],
                fill=PatternFill("solid", fgColor=C_WARN)))
oc.conditional_formatting.add("A5:S{}".format(LAST_OC),
    FormulaRule(formula=['$Z5="Arquivada"'], font=Font(name=F, size=9, color="A6A6A6")))
oc["H4"].comment = Comment(
 "Fato observável: o que aconteceu, verificável por terceiro.\n\n"
 "Bom: 'Subiu hotfix sem aprovação na sexta 18h'.\n"
 "Ruim: 'É descuidado', 'falta postura', 'é proativo'.\n\n"
 "Feedback vago sem fato recai desproporcionalmente sobre mulheres e pessoas negras. "
 "É por isso que este campo é duro.", "RH")
oc["L4"].comment = Comment(
 "Ocorrência negativa nunca comunicada NÃO entra no cálculo do ciclo e não viaja no handover.\n\n"
 "Surpresa na devolutiva é falha de gestão, e é o que derruba um processo demissional.", "RH")
oc.freeze_panes = "C5"
oc.auto_filter.ref = "A4:AB{}".format(LAST_OC)

# =====================================================================
# AVALIAÇÃO
# =====================================================================
av = sh("Avaliação")
titulo(av, "Avaliação e calibração",
       "Chave: pessoa + ciclo + competência + ALOCAÇÃO. Sem a alocação na chave, quem tem dois contratos colide na mesma célula e a média silenciosa volta pela porta dos fundos.")
av["A3"] = ("Nunca a média entre avaliadores — nem GP × LT, nem GP × GP. Divergência é informação: cada lado traz uma ocorrência; "
            "se só um tem evidência, a nota dele prevalece.")
av["A3"].font = fnt(9, it=True, color="C00000")
COLS_AV = ["Ciclo", "Idx ciclo", "ID_PESSOA", "Pessoa", "ID_ALOC", "Senioridade", "Competência", "Eixo",
           "Avaliador dono", "Nota GP", "Nota LT", "Base obs. GP", "Base obs. LT", "Divergência",
           "Sugestão", "Nota final", "Decisor", "Origem da nota final", "Motivo do desempate",
           "Papel acumulado", "Ocorr. vinculadas", "Lastro", "Justificativa (1 frase)", "Data calibração",
           "chave ID|ciclo", "chave ID|ciclo|eixo", "chave ID|ciclo|comp"]
larg(av, {"A": 10, "B": 9, "C": 11, "D": 19, "E": 10, "F": 12, "G": 22, "H": 17, "I": 15,
          "J": 9, "K": 9, "L": 12, "M": 12, "N": 11, "O": 11, "P": 10, "Q": 15, "R": 18,
          "S": 42, "T": 13, "U": 12, "V": 13, "W": 46, "X": 13, "Y": 15, "Z": 22, "AA": 24})
cab(av, 4, COLS_AV)
seed = []
for pid, nome, *_ in PESSOAS:
    for a in ALOCS:
        if a[1] != pid: continue
        for c in COMP:
            seed.append((pid, a[0], c[2]))
for i, (pid, aid, comp) in enumerate(seed):
    rr = 5 + i
    av.cell(row=rr, column=1, value=CICLO)
    av.cell(row=rr, column=3, value=pid)
    av.cell(row=rr, column=5, value=aid)
    av.cell(row=rr, column=7, value=comp)
    if pid == "P001":
        g, l, f_ = EX[comp]
        if g is not None: av.cell(row=rr, column=10, value=g)
        if l is not None: av.cell(row=rr, column=11, value=l)
        av.cell(row=rr, column=12, value="Cerimônias" if g is not None else "Insuficiente")
        av.cell(row=rr, column=13, value="Diária" if l is not None else "Esporádica")
        av.cell(row=rr, column=16, value=f_)
        av.cell(row=rr, column=17, value="Marcos Lima")
        av.cell(row=rr, column=20, value="Não")
        av.cell(row=rr, column=24, value=dt.date(2026, 9, 30))
JUST = {
 "Qualidade da entrega": "PR #857 voltou 3 rodadas por teste de erro ausente; é o padrão do ciclo, não um caso isolado.",
 "Comunicação e colaboração": "GP viu 4, LT viu 5. O review do PR #812 pegou uma condição de corrida antes do merge — evidência do LT desempatou.",
 "Autonomia e ownership": "GP viu 3, LT viu 4; a evidência do LT (assumiu o tema de resgate ponta a ponta) desempatou.",
}
for i, (pid, aid, comp) in enumerate(seed):
    if pid == "P001" and comp in JUST:
        av.cell(row=5 + i, column=19, value="Evidência do Líder Técnico, competência de dono técnico.")
        av.cell(row=5 + i, column=23, value=JUST[comp])
for rr in range(5, LAST_AV + 1):
    av.cell(row=rr, column=2, value='=IF($A{0}="","",VALUE(LEFT($A{0},4))*4+VALUE(RIGHT($A{0},1)))'.format(rr))
    av.cell(row=rr, column=4, value='=IFERROR(INDEX(Pessoas!$B$5:$B${1},MATCH($C{0},Pessoas!$A$5:$A${1},0)),"")'.format(rr, LAST_PES))
    av.cell(row=rr, column=6, value='=IFERROR(INDEX(Pessoas!$D$5:$D${1},MATCH($C{0},Pessoas!$A$5:$A${1},0)),"")'.format(rr, LAST_PES))
    av.cell(row=rr, column=8, value='=IFERROR(INDEX(Config!$B$40:$B$45,MATCH($G{0},Config!$C$40:$C$45,0)),"")'.format(rr))
    av.cell(row=rr, column=9, value='=IFERROR(INDEX(Config!$D$40:$D$45,MATCH($G{0},Config!$C$40:$C$45,0)),"")'.format(rr))
    av.cell(row=rr, column=14, value='=IF(AND(ISNUMBER($J{0}),ISNUMBER($K{0})),ABS($J{0}-$K{0}),"")'.format(rr))
    av.cell(row=rr, column=15, value=(
        '=IF(AND(NOT(ISNUMBER($J{0})),NOT(ISNUMBER($K{0}))),"",'
        'IF(NOT(ISNUMBER($K{0})),$J{0},IF(NOT(ISNUMBER($J{0})),$K{0},'
        'IF($N{0}=0,$J{0},IF(AND($N{0}=1,$T{0}<>"Sim"),'
        'IF($I{0}="Líder Técnico",$K{0},IF($I{0}="GP",$J{0},"CALIBRAR")),"CALIBRAR")))))').format(rr))
    av.cell(row=rr, column=18, value=(
        '=IF($P{0}="","",IF(AND(ISNUMBER($J{0}),$P{0}=$J{0},ISNUMBER($K{0}),$P{0}=$K{0}),"= ambos",'
        'IF(AND(ISNUMBER($J{0}),$P{0}=$J{0}),"= GP",IF(AND(ISNUMBER($K{0}),$P{0}=$K{0}),"= LT",'
        '"nenhuma das duas"))))').format(rr))
    av.cell(row=rr, column=21, value=(
        '=IF($C{0}="","",COUNTIFS({1}!$V$5:$V${2},$C{0}&"|"&$B{0}&"|"&$G{0},{1}!$X$5:$X${2},"Sim"))'
        ).format(rr, SH_OC, LAST_OC))
    av.cell(row=rr, column=22, value=(
        '=IF($P{0}="","",IF(OR($P{0}=1,$P{0}=2,$P{0}=5),IF($U{0}>0,"OK","SEM LASTRO"),"OK"))').format(rr))
    av.cell(row=rr, column=25, value='=IF($C{0}="","",$C{0}&"|"&$B{0})'.format(rr))
    av.cell(row=rr, column=26, value='=IF($C{0}="","",$C{0}&"|"&$B{0}&"|"&$H{0})'.format(rr))
    av.cell(row=rr, column=27, value='=IF($C{0}="","",$C{0}&"|"&$B{0}&"|"&$G{0})'.format(rr))
    av.cell(row=rr, column=24).number_format = "DD/MM/YYYY"
    for col in ("J", "K", "N", "O", "P", "U"):
        av[col + str(rr)].alignment = Alignment(horizontal="center", vertical="center")
for rng in ("A5:A{}", "C5:C{}", "E5:E{}", "G5:G{}", "J5:M{}", "P5:Q{}", "S5:T{}", "W5:X{}"):
    pinta(av, rng.format(LAST_AV), "in")
for rng in ("B5:B{}", "D5:D{}", "F5:F{}", "H5:I{}", "N5:O{}", "R5:R{}", "U5:V{}", "Y5:AA{}"):
    pinta(av, rng.format(LAST_AV), "calc")
for k, col in [("ciclo", "A"), ("comp", "G"), ("nota", "J"), ("nota", "K"),
               ("baseobs", "L"), ("baseobs", "M"), ("nota", "P"), ("simnao", "T")]:
    dv(av, k, "{0}5:{0}{1}".format(col, LAST_AV))
av.conditional_formatting.add("V5:V{}".format(LAST_AV),
    CellIsRule(operator="equal", formula=['"SEM LASTRO"'], fill=PatternFill("solid", fgColor=C_BAD),
               font=Font(name=F, size=9, bold=True, color="9C0006")))
av.conditional_formatting.add("N5:O{}".format(LAST_AV), FormulaRule(formula=['$N5>=2'],
    fill=PatternFill("solid", fgColor=C_WARN)))
av.conditional_formatting.add("S5:S{}".format(LAST_AV),
    FormulaRule(formula=['AND($N5>=1,$S5="")'], fill=PatternFill("solid", fgColor=C_BAD)))
av.conditional_formatting.add("L5:M{}".format(LAST_AV),
    CellIsRule(operator="equal", formula=['"Insuficiente"'], fill=PatternFill("solid", fgColor=C_WARN)))
av["P4"].comment = Comment(
 "A nota final é DECISÃO, não cálculo. A Sugestão só resolve o caso trivial.\n\n"
 "Divergência 1: decide o avaliador dono da competência.\n"
 "Divergência 2+: cada lado traz uma ocorrência. Se só um tem evidência, a nota dele prevalece.\n"
 "Nunca a média.", "Gestão de Pessoas")
av["T4"].comment = Comment(
 "Papel acumulado = Sim quando o Líder Técnico É TAMBÉM o Gestor de Pessoas.\n\n"
 "Nesse caso a pessoa avalia, decide e arbitra a própria divergência. O desempate sobe um nível "
 "(gestor do gestor, ou RH) e o Motivo do desempate passa a ser obrigatório em TODA divergência.", "Gestão de Pessoas")
av["L4"].comment = Comment(
 "Base de observação = Insuficiente significa que a nota NÃO deve ser dada: deixe em branco.\n\n"
 "Nota de quem entrou há três semanas não pode pesar igual à de quem acompanha há um ano.", "Gestão de Pessoas")
av.freeze_panes = "D5"
av.auto_filter.ref = "A4:AA{}".format(LAST_AV)

# =====================================================================
# PAINEL
# =====================================================================
pa = sh("Painel")
titulo(pa, "Painel do gestor — o que abrir na segunda de manhã")
pa["A2"] = ('="Ciclo "&Config!$B$4&"  ·  comparação sempre por GAP contra a expectativa da senioridade — nunca por nota bruta, nunca entre squads."')
pa["A2"].font = fnt(9, it=True, color="595959")
pa["A3"] = ('=IF(SUM(Pessoas!$O$5:$O${0})=0,"Todas as pessoas têm registro neste ciclo.",'
            '"Sem nenhum registro neste ciclo: "'
            '&IFERROR(INDEX(Pessoas!$B$5:$B${0},MATCH(1,Pessoas!$P$5:$P${0},0)),"")'
            '&IFERROR(", "&INDEX(Pessoas!$B$5:$B${0},MATCH(2,Pessoas!$P$5:$P${0},0)),"")'
            '&IFERROR(", "&INDEX(Pessoas!$B$5:$B${0},MATCH(3,Pessoas!$P$5:$P${0},0)),"")'
            '&IF(SUM(Pessoas!$O$5:$O${0})>3,", …","")'
            '&" ("&SUM(Pessoas!$O$5:$O${0})&" de "&COUNTIF(Pessoas!$I$5:$I${0},"Ativo")&")")').format(LAST_PES)
pa["A3"].font = Font(name=F, size=11, bold=True, color="9C0006")
pa["A3"].comment = Comment("A única métrica desta planilha que mede o GESTOR, não a equipe. Vazio é o problema real, não a nota baixa.", "Gestão de Pessoas")
pa["A4"] = ('="Ocorrências registradas: "&COUNTIF({0}!$A$5:$A${1},">0")&" de {2} linhas disponíveis"').format(SH_OC, LAST_OC, LAST_OC - 4)
pa["A4"].font = fnt(9, b=True)
pa.conditional_formatting.add("A4", FormulaRule(
    formula=['COUNTIF(\'Ocorrências\'!$A$5:$A${0},">0")>{1}'.format(LAST_OC, int((LAST_OC - 4) * 0.8))],
    fill=PatternFill("solid", fgColor=C_WARN)))
COLS_PA = ["ID", "Pessoa", "Senioridade", "Papel considerado", "Multi", "Eixo A", "Eixo B", "Eixo C",
           "Eixo D", "Nota ponderada", "Expectativa", "Gap", "Faixa", "Ocorr.", "Positivas",
           "Negativas", "Neg. não comunic.", "Dias desde 1:1", "Diverg. 2+", "Sem lastro",
           "Nota s/ omissão", "Sinal", "num", "den"]
larg(pa, {"A": 9, "B": 20, "C": 12, "D": 14, "E": 7, "F": 8, "G": 8, "H": 8, "I": 8, "J": 12,
          "K": 11, "L": 8, "M": 12, "N": 8, "O": 9, "P": 10, "Q": 12, "R": 12, "S": 10, "T": 10,
          "U": 12, "V": 26, "W": 8, "X": 8})
cab(pa, 6, COLS_PA)
for i, lab in enumerate(["num", "den"]):
    c = pa.cell(row=6, column=23 + i)
    c.fill = PatternFill("solid", fgColor="A6A6A6"); c.font = Font(name=F, size=8, italic=True, color="FFFFFF")
EIXO_ROW = {EA: 22, EB: 23, EC: 24, ED: 25}
for rr in range(7, LAST_PA + 1):
    src = rr - 2
    pa.cell(row=rr, column=1, value='=IF(Pessoas!$B{0}="","",Pessoas!$A{0})'.format(src))
    for col, sc in [(2, "B"), (3, "D"), (4, "L"), (5, "M")]:
        pa.cell(row=rr, column=col, value='=IF(Pessoas!$B{0}="","",Pessoas!${1}{0})'.format(src, sc))
    for i, e in enumerate(EIXOS):
        pa.cell(row=rr, column=6 + i, value=(
            '=IF($B{0}="","",IFERROR(AVERAGEIFS({1}!$P$5:$P${2},{1}!$Z$5:$Z${2},'
            '$A{0}&"|"&Config!$B$7&"|"&Config!$A${3}),""))').format(rr, SH_AV, LAST_AV, EIXO_ROW[e]))
    pa.cell(row=rr, column=23, value=(
        '=IF(ISNUMBER($F{0}),$F{0}*Config!$B$22,0)+IF(ISNUMBER($G{0}),$G{0}*Config!$B$23,0)'
        '+IF(ISNUMBER($H{0}),$H{0}*Config!$B$24,0)+IF(ISNUMBER($I{0}),$I{0}*Config!$B$25,0)').format(rr))
    pa.cell(row=rr, column=24, value=(
        '=IF(ISNUMBER($F{0}),Config!$B$22,0)+IF(ISNUMBER($G{0}),Config!$B$23,0)'
        '+IF(ISNUMBER($H{0}),Config!$B$24,0)+IF(ISNUMBER($I{0}),Config!$B$25,0)').format(rr))
    pa.cell(row=rr, column=10, value='=IF($X{0}=0,"",$W{0}/$X{0})'.format(rr))
    pa.cell(row=rr, column=11, value='=IFERROR(INDEX(Config!$B$30:$B$33,MATCH($C{0},Config!$A$30:$A$33,0)),"")'.format(rr))
    pa.cell(row=rr, column=12, value='=IF(OR($J{0}="",$K{0}=""),"",$J{0}-$K{0})'.format(rr))
    pa.cell(row=rr, column=13, value=(
        '=IF($J{0}="","",IF($J{0}<2.5,"Abaixo",IF($J{0}<3.5,"Atende",IF($J{0}<=4.2,"Supera","Referência"))))').format(rr))
    pa.cell(row=rr, column=14, value='=IF($B{0}="","",COUNTIFS({1}!$U$5:$U${2},$A{0}&"|"&Config!$B$7))'.format(rr, SH_OC, LAST_OC))
    pa.cell(row=rr, column=15, value=(
        '=IF($B{0}="","",COUNTIFS({1}!$U$5:$U${2},$A{0}&"|"&Config!$B$7,{1}!$F$5:$F${2},"Reconhecimento")'
        '+COUNTIFS({1}!$U$5:$U${2},$A{0}&"|"&Config!$B$7,{1}!$F$5:$F${2},"Positiva"))').format(rr, SH_OC, LAST_OC))
    pa.cell(row=rr, column=16, value=(
        '=IF($B{0}="","",COUNTIFS({1}!$U$5:$U${2},$A{0}&"|"&Config!$B$7,{1}!$F$5:$F${2},"Ponto de atenção")'
        '+COUNTIFS({1}!$U$5:$U${2},$A{0}&"|"&Config!$B$7,{1}!$F$5:$F${2},"Incidente"))').format(rr, SH_OC, LAST_OC))
    pa.cell(row=rr, column=17, value=(
        '=IF($B{0}="","",COUNTIFS({1}!$U$5:$U${2},$A{0}&"|"&Config!$B$7,{1}!$F$5:$F${2},"Ponto de atenção",'
        '{1}!$L$5:$L${2},"Não")+COUNTIFS({1}!$U$5:$U${2},$A{0}&"|"&Config!$B$7,{1}!$F$5:$F${2},"Incidente",'
        '{1}!$L$5:$L${2},"Não"))').format(rr, SH_OC, LAST_OC))
    pa.cell(row=rr, column=18, value=(
        '=IF($B{0}="","",IF(INDEX(Pessoas!$J$5:$J${1},MATCH($A{0},Pessoas!$A$5:$A${1},0))="","nunca",'
        'Config!$B$6-INDEX(Pessoas!$J$5:$J${1},MATCH($A{0},Pessoas!$A$5:$A${1},0))))').format(rr, LAST_PES))
    pa.cell(row=rr, column=19, value=(
        '=IF($B{0}="","",COUNTIFS({1}!$Y$5:$Y${2},$A{0}&"|"&Config!$B$7,{1}!$N$5:$N${2},">=2"))').format(rr, SH_AV, LAST_AV))
    pa.cell(row=rr, column=20, value=(
        '=IF($B{0}="","",COUNTIFS({1}!$Y$5:$Y${2},$A{0}&"|"&Config!$B$7,{1}!$V$5:$V${2},"SEM LASTRO"))').format(rr, SH_AV, LAST_AV))
    pa.cell(row=rr, column=21, value=(
        '=IF($B{0}="","",IF(AND($Q{0}>0,COUNTIFS({1}!$Y$5:$Y${2},$A{0}&"|"&Config!$B$7,{1}!$P$5:$P${2},">=3")>0),$Q{0},0))'
        ).format(rr, SH_AV, LAST_AV))
    pa.cell(row=rr, column=22, value=(
        '=IF($B{0}="","",'
        'IF($U{0}>0,"Nota sustentada por omissão",'
        'IF($T{0}>0,"Nota sem lastro",'
        'IF($J{0}="","Sem avaliação no ciclo",'
        'IF($S{0}>0,"Calibrar divergência",'
        'IF($L{0}<-0.3,"Abaixo da expectativa",'
        'IF($Q{0}>0,"Negativa não conversada",'
        'IF(OR($R{0}="nunca",$R{0}>30),"1:1 atrasada","OK"))))))))').format(rr))
    for col in ("F", "G", "H", "I", "J", "K", "L", "W", "X"):
        pa[col + str(rr)].number_format = "0.00;-0.00;-"
        pa[col + str(rr)].alignment = Alignment(horizontal="center")
    for col in ("N", "O", "P", "Q", "R", "S", "T", "U"):
        pa[col + str(rr)].alignment = Alignment(horizontal="center")
pinta(pa, "A7:V{}".format(LAST_PA), "calc")
pinta(pa, "W7:X{}".format(LAST_PA), "calc")
rng = "A7:V{}".format(LAST_PA)
for txt, cor in [("Nota sustentada por omissão", C_BAD), ("Nota sem lastro", C_BAD),
                 ("Abaixo da expectativa", C_WARN), ("Negativa não conversada", C_WARN),
                 ("OK", C_GOOD)]:
    pa.conditional_formatting.add(rng, FormulaRule(formula=['$V7="{}"'.format(txt)],
                                                   fill=PatternFill("solid", fgColor=cor)))
pa.conditional_formatting.add("L7:L{}".format(LAST_PA),
    CellIsRule(operator="lessThan", formula=["0"], font=Font(name=F, size=9, bold=True, color="9C0006")))
pa["U6"].comment = Comment(
 "REGRA 9 — a contradição que ela resolve.\n\n"
 "A regra 2 exige ocorrência vinculada para nota 1, 2 ou 5.\n"
 "A regra 4 tira do ciclo a ocorrência negativa não comunicada.\n"
 "Logo, o GP que não deu o feedback era obrigado a dar 3: a planilha convertia a omissão do gestor "
 "em nota melhor para a pessoa, em silêncio.\n\n"
 "Este alerta é endereçado ao GESTOR DO GP. Nunca à pessoa avaliada.", "Qualidade")
pa.freeze_panes = "C7"
pa.auto_filter.ref = "A6:V{}".format(LAST_PA)

# =====================================================================
# 1a1 — folha imprimível
# =====================================================================
um = sh("1a1")
titulo(um, "Pauta de 1:1 — imprima e leve na conversa",
       "Escolha a pessoa em B4. A folha monta sozinha. Reconhecimentos vêm primeiro, de propósito.")
larg(um, {"A": 25, "B": 21, "C": 11, "D": 11, "E": 46, "F": 34, "G": 13})
um["A4"] = "Pessoa"; um["A4"].font = fnt(10, b=True)
um["B4"] = "Ana Ribeiro"
um["B4"].fill = PatternFill("solid", fgColor=C_IN); um["B4"].font = fnt(11, b=True, color=BLUE); um["B4"].border = BOX
dv(um, "pessoa", "B4")
um["C4"] = ('=IFERROR(INDEX(Pessoas!$A$5:$A${0},MATCH($B$4,Pessoas!$B$5:$B${0},0)),"")').format(LAST_PES)
um["C4"].font = fnt(9, color="A6A6A6")
um["D4"] = ('=IFERROR(INDEX(Pessoas!$D$5:$D${0},MATCH($B$4,Pessoas!$B$5:$B${0},0))&"  ·  "'
            '&INDEX(Pessoas!$L$5:$L${0},MATCH($B$4,Pessoas!$B$5:$B${0},0))&"  ·  "'
            '&INDEX(Pessoas!$N$5:$N${0},MATCH($B$4,Pessoas!$B$5:$B${0},0))&" meses no nível","")').format(LAST_PES)
um["D4"].font = fnt(9, it=True, color="595959")
um["A5"] = "Ciclo"; um["A5"].font = fnt(10, b=True)
um["B5"] = "=Config!$B$4"; um["B5"].fill = PatternFill("solid", fgColor=C_CALC); um["B5"].border = BOX; um["B5"].font = fnt(10, b=True)
um["D5"] = ('=IFERROR("Nota ponderada "&TEXT(INDEX({0}!$J$7:$J${1},MATCH($B$4,{0}!$B$7:$B${1},0)),"0.00")'
            '&"   ·   expectativa "&TEXT(INDEX({0}!$K$7:$K${1},MATCH($B$4,{0}!$B$7:$B${1},0)),"0.0")'
            '&"   ·   gap "&TEXT(INDEX({0}!$L$7:$L${1},MATCH($B$4,{0}!$B$7:$B${1},0)),"+0.00;-0.00")'
            '&"   ·   "&INDEX({0}!$M$7:$M${1},MATCH($B$4,{0}!$B$7:$B${1},0)),"Sem avaliação neste ciclo")').format(SH_PA, LAST_PA)
um["D5"].font = fnt(10, b=True, color=C_TIT)

faixa(um, 7, "1.  COMPETÊNCIAS — a conversa é sobre a distância até a expectativa, não sobre a nota", 7)
cab(um, 8, ["Competência", "GP", "LT", "Final", "Onde ela está hoje (âncora do nível)",
            "O que falta para o próximo nível", "Gap"])
for i, comp in enumerate(COMP_NOMES):
    rr = 9 + i
    um.cell(row=rr, column=1, value=comp).font = fnt(9, b=True)
    for col, letra in [(2, "J"), (3, "K"), (4, "P")]:
        um.cell(row=rr, column=col, value=(
            '=IF(COUNTIFS({0}!$AA$5:$AA${1},$C$4&"|"&Config!$B$7&"|"&$A{2},{0}!${3}$5:${3}${1},">0")=0,"",'
            'SUMIFS({0}!${3}$5:${3}${1},{0}!$AA$5:$AA${1},$C$4&"|"&Config!$B$7&"|"&$A{2}))'
            ).format(SH_AV, LAST_AV, rr, letra))
    um.cell(row=rr, column=5, value='=IFERROR(INDEX(Config!$E$40:$I$45,MATCH($A{0},Config!$C$40:$C$45,0),$D{0}),"")'.format(rr))
    um.cell(row=rr, column=6, value=(
        '=IFERROR(IF($D{0}>=5,"Já é referência nesta competência.",'
        'INDEX(Config!$E$40:$I$45,MATCH($A{0},Config!$C$40:$C$45,0),$D{0}+1)),"")').format(rr))
    um.cell(row=rr, column=7, value=(
        '=IFERROR(IF($D{0}="","",$D{0}-INDEX({1}!$K$7:$K${2},MATCH($B$4,{1}!$B$7:$B${2},0))),"")').format(rr, SH_PA, LAST_PA))
    um.row_dimensions[rr].height = 44
pinta(um, "A9:G14", "calc")
for rr in range(9, 15):
    for col in ("B", "C", "D", "G"):
        um[col + str(rr)].alignment = Alignment(horizontal="center", vertical="center")
        um[col + str(rr)].number_format = "0.0;-0.0;-"
um.conditional_formatting.add("G9:G14", CellIsRule(operator="lessThan", formula=["0"],
                                                   fill=PatternFill("solid", fgColor=C_WARN)))
um.conditional_formatting.add("B9:C14", FormulaRule(
    formula=['AND(ISNUMBER($B9),ISNUMBER($C9),ABS($B9-$C9)>=2)'], fill=PatternFill("solid", fgColor=C_BAD)))
um["C8"].comment = Comment(
 "Vermelho = GP e Líder Técnico divergiram em 2+ pontos.\n\n"
 "Diga isso na cara, na 1:1: 'o que discordamos sobre você'.\n"
 "Esconder a divergência é o que faz a devolutiva soar falsa.", "Liderança Técnica")

faixa(um, 16, "2.  O QUE ACONTECEU NESTE CICLO — o antídoto contra o viés de recência", 7)
cab(um, 17, ["Data", "Tipo", "Competência", "Impacto", "Fato observável", "Consequência / desfecho", "Conversado?"])
for k in range(1, 13):
    rr = 17 + k
    for col, src in [(1, "A"), (2, "F"), (3, "G"), (4, "I"), (5, "H"), (6, "K"), (7, "L")]:
        um.cell(row=rr, column=col, value=(
            '=IFERROR(INDEX({0}!${1}$5:${1}${2},MATCH({3},{0}!$AB$5:$AB${2},0)),"")').format(SH_OC, src, LAST_OC, k))
    um["A" + str(rr)].number_format = "DD/MM/YYYY"
    um.row_dimensions[rr].height = 30
pinta(um, "A18:G29", "calc")
um.conditional_formatting.add("A18:G29", FormulaRule(
    formula=['OR($B18="Reconhecimento",$B18="Positiva")'], fill=PatternFill("solid", fgColor=C_GOOD)))
um.conditional_formatting.add("A18:G29", FormulaRule(
    formula=['AND($G18="Não",OR($B18="Ponto de atenção",$B18="Incidente"))'], fill=PatternFill("solid", fgColor=C_BAD)))

faixa(um, 31, "3.  PLANO DE DESENVOLVIMENTO — no máximo 2 focos", 7)
cab(um, 32, ["Competência", "Atual", "Alvo", "Prazo", "Ação concreta no trabalho real", "Como saberemos que fechou", "Status"])
for k in range(1, 4):
    rr = 32 + k
    for col, src in [(1, "C"), (2, "E"), (3, "F"), (4, "J"), (5, "G"), (6, "H"), (7, "L")]:
        um.cell(row=rr, column=col, value=(
            '=IFERROR(INDEX({0}!${1}$5:${1}${2},MATCH({3},{0}!$N$5:$N${2},0)),"")').format(SH_PDI, src, LAST_PDI, k))
    um["D" + str(rr)].number_format = "DD/MM/YYYY"
    um.row_dimensions[rr].height = 34
pinta(um, "A33:G35", "calc")

faixa(um, 37, "4.  ESCUTA — as perguntas que o gestor esquece de fazer", 7)
for i, p in enumerate(["Como está a sua carga? O que está pesando mais do que deveria?",
                       "O que te travou neste ciclo e eu não vi?",
                       "O que você quer estar fazendo daqui a um ano?",
                       "O que eu preciso fazer diferente como gestor?"]):
    rr = 38 + i
    c = um.cell(row=rr, column=1, value=p); c.font = fnt(9, b=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    um.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=7)
    for j in range(2, 8): um.cell(row=rr, column=j).border = BOX
    um.row_dimensions[rr].height = 30

faixa(um, 43, "5.  ANOTAÇÕES DA CONVERSA — escreva à mão, aqui", 7)
for rr in range(44, 49):
    for j in range(1, 8):
        um.cell(row=rr, column=j).border = BOX
    um.row_dimensions[rr].height = 24
um.print_area = "A1:G48"

faixa(um, 50, "6.  REGISTRO DE 1:1 — digite depois da conversa. É esta data que alimenta o alerta de 1:1 atrasada no Painel.", 7)
cab(um, 51, ["Data da 1:1", "Pessoa (ID)", "Gestor", "Compromisso assumido", "Dono", "Prazo", "1:1 realizada?"])
for rr in range(52, LAST_1A1 + 1):
    um["A" + str(rr)].number_format = "DD/MM/YYYY"
    um["F" + str(rr)].number_format = "DD/MM/YYYY"
pinta(um, "A52:G{}".format(LAST_1A1), "in")
for j, v in enumerate([dt.date(2026, 8, 20), "P001", "Rafael Souza",
                       "Incluir teste de caminho de erro nos próximos 3 PRs de rota nova.",
                       "Ana Ribeiro", dt.date(2026, 11, 30), "Sim"]):
    um.cell(row=52, column=1 + j, value=v)
dv(um, "simnao", "G52:G{}".format(LAST_1A1))
um["B51"].comment = Comment(
 "Use o ID da pessoa (P001), não o nome. O ID é a chave estável do modelo:\n"
 "nome muda, ID não. É ele que liga esta linha ao Painel.", "Técnico")
um.page_setup.orientation = "portrait"
um.page_setup.fitToWidth = 1
um.sheet_properties.pageSetUpPr.fitToPage = True
um.freeze_panes = "A6"

# =====================================================================
# PDI
# =====================================================================
pd_ = sh("PDI")
titulo(pd_, "Plano de desenvolvimento", "No máximo 2 focos por pessoa por ciclo. Plano com 6 itens não é plano.")
pd_["A3"] = ('"Melhorar em testes" não é PDI.  ·  "Nos próximos 3 PRs de rota nova, incluir teste de caminho de erro '
             'sem que o review precise pedir — verificado por mim em 30/11" é PDI.')
pd_["A3"].font = fnt(9, it=True, color="C00000")
COLS_PDI = ["Ciclo", "Pessoa", "Competência-alvo", "Nota atual", "Nota alvo", "Gap",
            "Ação concreta no trabalho real", "Como saberemos que fechou", "Apoio", "Prazo",
            "Último check-in", "Status", "ID_PESSOA", "seq"]
larg(pd_, {"A": 10, "B": 20, "C": 24, "D": 10, "E": 10, "F": 8, "G": 46, "H": 46, "I": 22,
           "J": 12, "K": 13, "L": 14, "M": 11, "N": 6})
cab(pd_, 4, COLS_PDI)
PDIEX = [
 (CICLO, "Ana Ribeiro", "Qualidade da entrega", 3,
  "Ser a dona dos testes da nova rota de resgate — escopo real, com o LT como rede de proteção. Não curso.",
  "Nos próximos 3 PRs de rota nova, teste de caminho de erro presente sem o review precisar pedir.",
  "Rafael Souza (LT), pareamento de 1h/semana", dt.date(2026,11,30), dt.date(2026,9,30), "Em andamento"),
 (CICLO, "Ana Ribeiro", "Previsibilidade", 4,
  "Conduzir o refinamento do épico de cupom de ponta a ponta, incluindo a conversa com o PO.",
  "Dois refinamentos conduzidos sem o GP precisar entrar para destravar.",
  "Marcos Lima (GP), observando sem intervir", dt.date(2026,12,15), dt.date(2026,9,30), "Não iniciado"),
]
for i, row in enumerate(PDIEX):
    rr = 5 + i
    for j, v in zip([1, 2, 3, 5, 7, 8, 9, 10, 11, 12], row):
        pd_.cell(row=rr, column=j, value=v)
for rr in range(5, LAST_PDI + 1):
    pd_.cell(row=rr, column=13, value='=IFERROR(INDEX(Pessoas!$A$5:$A${1},MATCH($B{0},Pessoas!$B$5:$B${1},0)),"")'.format(rr, LAST_PES))
    pd_.cell(row=rr, column=4, value=(
        '=IF($B{0}="","",IF(COUNTIFS({1}!$AA$5:$AA${2},$M{0}&"|"&VALUE(LEFT($A{0},4))*4+VALUE(RIGHT($A{0},1))&"|"&$C{0},'
        '{1}!$P$5:$P${2},">0")=0,"",SUMIFS({1}!$P$5:$P${2},{1}!$AA$5:$AA${2},'
        '$M{0}&"|"&VALUE(LEFT($A{0},4))*4+VALUE(RIGHT($A{0},1))&"|"&$C{0})))').format(rr, SH_AV, LAST_AV))
    pd_.cell(row=rr, column=6, value='=IF(OR($D{0}="",$E{0}=""),"",$E{0}-$D{0})'.format(rr))
    pd_.cell(row=rr, column=14, value=(
        '=IF(OR($B{0}="",$B{0}<>{1}!$B$4,$A{0}<>Config!$B$4),"",COUNTIFS($B$5:$B{0},{1}!$B$4,$A$5:$A{0},Config!$B$4))'
        ).format(rr, SH_1A1))
    for col in ("J", "K"): pd_[col + str(rr)].number_format = "DD/MM/YYYY"
for rng in ("A5:C{}", "E5:E{}", "G5:L{}"): pinta(pd_, rng.format(LAST_PDI), "in")
for rng in ("D5:D{}", "F5:F{}", "M5:N{}"): pinta(pd_, rng.format(LAST_PDI), "calc")
for k, col in [("ciclo", "A"), ("pessoa", "B"), ("comp", "C"), ("nota", "E"), ("stpdi", "L")]:
    dv(pd_, k, "{0}5:{0}{1}".format(col, LAST_PDI))
pd_["H4"].comment = Comment(
 "O campo que faz diferença.\n\n"
 "'Melhorar em testes' não se verifica. 'Teste de caminho de erro presente sem o review pedir, "
 "nos próximos 3 PRs, verificado por mim em 30/11' se verifica.", "Liderança Técnica")
pd_.freeze_panes = "C5"

# =====================================================================
# ALERTAS
# =====================================================================
alr = sh("Alertas")
titulo(alr, "Integridade do ciclo — o que precisa estar zerado antes de fechar",
       "Oito alertas, não dezoito. Os que dependem de série histórica ficam suprimidos com '—' até haver ciclos suficientes.")
larg(alr, {"A": 4, "B": 44, "C": 12, "D": 10, "E": 86})
cab(alr, 4, ["#", "Verificação", "Ocorrências", "Ciclos", "Por que importa / o que fazer"])
P, O_, A_ = SH_PA, SH_OC, SH_AV
CHECKS = [
 ("Pessoas ativas sem avaliação no ciclo",
  '=COUNTIF({0}!$V$7:$V${1},"Sem avaliação no ciclo")'.format(P, LAST_PA), 1,
  "É a visão mais valiosa da planilha. Ciclo com metade da squad sem nota não vale para mérito nem para PDI."),
 ("Notas 1, 2 ou 5 sem ocorrência vinculada",
  '=SUM({0}!$T$7:$T${1})'.format(P, LAST_PA), 1,
  "Nota sem lastro é impressão, não avaliação. Registre a ocorrência que sustenta a nota, ou mude a nota."),
 ("Nota sustentada por omissão (regra 9)",
  '=SUM({0}!$U$7:$U${1})'.format(P, LAST_PA), 1,
  "Há ocorrência negativa não conversada E nota 3 ou mais. O GP que não deu o feedback foi obrigado a dar 3 — "
  "a omissão virou nota melhor para a pessoa. Endereçado ao gestor do GP, nunca à pessoa avaliada."),
 ("Ocorrências negativas nunca conversadas",
  '=SUM({0}!$Q$7:$Q${1})'.format(P, LAST_PA), 1,
  "Não entram no cálculo do ciclo e não viajam no handover. Surpresa na devolutiva é falha de gestão."),
 ("Divergências GP × LT de 2+ pontos sem justificativa",
  '=COUNTIFS({0}!$N$5:$N${1},">=2",{0}!$S$5:$S${1},"")'.format(A_, LAST_AV), 1,
  "Divergência não é erro, é sinal. Cada lado traz uma ocorrência; se só um tem evidência, a nota dele prevalece. Nunca a média."),
 ("Registros de ocorrência incompletos",
  '=COUNTIF({0}!$W$5:$W${1},"Faltam*")'.format(O_, LAST_OC), 1,
  "A obrigatoriedade aqui é detectiva, não preventiva: ninguém é impedido de salvar. O que o modelo garante é que a linha "
  "incompleta não conta — não entra no lastro, não viaja no handover."),
 ("1:1 atrasada há mais de 30 dias",
  '=COUNTIF({0}!$V$7:$V${1},"1:1 atrasada")'.format(P, LAST_PA), 1,
  "Planilha que não vira conversa morre. Esta é a falha que mata a ferramenta — as outras são consertáveis."),
 ("Carga de gestão acima de 15 liderados",
  '=IF(COUNTIF(Pessoas!$I$5:$I${0},"Ativo")>15,COUNTIF(Pessoas!$I$5:$I${0},"Ativo")-15,0)'.format(LAST_PES), 1,
  "Acima de 15, a nota final vira média disfarçada e ninguém percebe. É um alerta sobre a estrutura, não sobre as pessoas."),
]
for i, (nome, form, minciclos, porq) in enumerate(CHECKS):
    rr = 5 + i
    alr.cell(row=rr, column=1, value=i + 1).font = fnt(9, b=True)
    alr.cell(row=rr, column=2, value=nome).font = fnt(9, b=True)
    c = alr.cell(row=rr, column=3, value='=IF(Config!$B$9<{0},"—",{1})'.format(minciclos, form[1:]))
    c.font = fnt(12, b=True); c.alignment = Alignment(horizontal="center")
    alr.cell(row=rr, column=4, value=minciclos).alignment = Alignment(horizontal="center")
    alr.cell(row=rr, column=4).font = fnt(9, color="A6A6A6")
    alr.cell(row=rr, column=5, value=porq).font = fnt(9)
    for j in range(1, 6):
        alr.cell(row=rr, column=j).border = BOX
        alr.cell(row=rr, column=j).alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="center" if j in (1, 3, 4) else "left")
    alr.row_dimensions[rr].height = 40
alr.conditional_formatting.add("C5:C12", CellIsRule(operator="greaterThan", formula=["0"],
    fill=PatternFill("solid", fgColor=C_BAD), font=Font(name=F, size=12, bold=True, color="9C0006")))
alr.conditional_formatting.add("C5:C12", CellIsRule(operator="equal", formula=["0"],
    fill=PatternFill("solid", fgColor=C_GOOD)))
alr["D4"].comment = Comment(
 "Ciclos mínimos para o alerta valer.\n\n"
 "Sem carência, no primeiro ciclo quase todos disparam de uma vez e a aba vira ruído permanente. "
 "Uma aba que grita 40 linhas na primeira abertura nunca mais é aberta.", "Qualidade")
alr["B15"] = "Alertas deliberadamente FORA desta lista"
alr["B15"].font = fnt(10, b=True, color=C_TIT)
alr["B16"] = ("Nota chapada · concentração de registros no último mês · sem ocorrência há 45 dias · aniversário de tempo no nível · "
              "PDI sem movimento há 60 dias · 3+ negativas sem nenhuma positiva · 3 ciclos com gap ≥ 0 sem caso de promoção montado. "
              "Todos legítimos, nenhum urgente — e juntos afogam os oito que importam.")
alr["B17"] = ("Premissas desta versão: expectativas por senioridade e pesos por eixo vieram do modelo, não de medição. "
              "Os limiares (30 dias sem 1:1, divergência ≥ 2, gap −0,3, 15 liderados) são convenções editáveis. "
              "Ao mudar um limiar, mude também na fórmula do Painel que o usa.")
for r_ in (16, 17):
    alr.merge_cells(start_row=r_, start_column=2, end_row=r_, end_column=5)
    alr.cell(row=r_, column=2).font = fnt(9, it=True, color="595959")
    alr.cell(row=r_, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    alr.row_dimensions[r_].height = 40

# =====================================================================
# INSTRUÇÕES
# =====================================================================
ins = sh("Instruções")
titulo(ins, "Avaliação de Desempenho por Squad — GP + Líder Técnico + Gestor de Pessoas",
       "Versão 3 · construída com as visões de RH, Gerência de Projetos, Liderança Técnica e Gestão de Pessoas, e revisada por um revisor de qualidade e um revisor técnico")
larg(ins, {"A": 40, "B": 96})
BLOCOS = [
 ("faixa", "O QUE ESTA PLANILHA É"),
 ("p", "", "O registro contínuo é o produto; a nota é subproduto. Se só uma aba sobreviver, que seja 'Ocorrências'. "
           "Uma nota sem ocorrência vinculada não é avaliação, é impressão."),
 ("faixa", "RITMO DE USO"),
 ("kv", "Quando acontecer — 45 segundos", "Lance a ocorrência: bloco rápido, colunas A a J, um único campo de texto. "
                                          "Meta: 2 a 4 por pessoa por mês, com ao menos 1 positiva a cada 2 registros."),
 ("kv", "Sexta — 15 min pela squad", "Abra o 'Painel'. Leia a primeira linha vermelha: de quem você não sabe nada. Não avalie nada."),
 ("kv", "Fim do ciclo — 20 min por pessoa", "Preencha 'Avaliação'. Você não escreve do zero: lê o que já foi acumulado."),
 ("kv", "Antes da devolutiva", "Calibração GP × LT só nos itens divergentes. Depois, a 1:1 com a aba '1a1' impressa."),
 ("faixa", "QUEM AVALIA O QUÊ"),
 ("kv", "Líder Técnico decide", "Qualidade da entrega · Domínio técnico · Design de solução."),
 ("kv", "GP decide", "Previsibilidade."),
 ("kv", "Compartilhado", "Comunicação e colaboração · Autonomia e ownership. Quem não tem evidência deixa em branco — "
                         "em branco é resposta legítima e melhor que chute."),
 ("faixa", "AS NOVE REGRAS DURAS"),
 ("n", "1", "3 é 'atende plenamente PARA A SENIORIDADE'. Não é nota ruim. Um Júnior 3 e um Sênior 3 são comportamentos diferentes."),
 ("n", "2", "Nota 1, 2 ou 5 exige ocorrência vinculada no ciclo. Sem lastro, a célula acusa e o ciclo não fecha."),
 ("n", "3", "Nunca a média entre avaliadores — nem GP × LT, nem GP × GP em quem tem dois contratos. A divergência é a informação."),
 ("n", "4", "Ocorrência negativa nunca conversada com a pessoa NÃO entra no ciclo e não viaja no handover."),
 ("n", "5", "'Fato observável' não aceita adjetivo de personalidade. O juízo vai em 'Consequência', e tem de ser defensável a partir do fato ao lado."),
 ("n", "6", "Comparação e ranking usam SEMPRE o gap contra a expectativa da senioridade. Nunca a nota bruta, nunca entre squads."),
 ("n", "7", "Sem curva forçada. Distribuição imposta em squad de 6 pessoas é matematicamente absurda."),
 ("n", "8", "Conduta, assédio, denúncia e saúde NÃO entram aqui, em nenhuma aba. Registre 'encaminhado ao RH em <data>' e a apuração corre em processo próprio."),
 ("n", "9", "Nota 3 ou mais COM ocorrência negativa não conversada acende 'Nota sustentada por omissão' — endereçado ao gestor do GP, nunca à pessoa avaliada."),
 ("faixa", "DUAS COISAS QUE ESTA PLANILHA NÃO FAZ, E VOCÊ PRECISA SABER"),
 ("p", "", "1) A obrigatoriedade é DETECTIVA, não preventiva. Ninguém é impedido de salvar uma ocorrência incompleta — validação de dados "
           "não torna campo obrigatório, célula em branco nunca dispara validação. O que o modelo garante é que a linha incompleta NÃO CONTA: "
           "não entra no lastro, não viaja no handover, e acende alerta."),
 ("p", "", "2) Ocorrência antiga não se esconde sozinha. Use o filtro na coluna 'Exibição' (Ativa / Histórico / Arquivada). "
           "Erro tratado não se cobra três vezes — mas quem esconde é o filtro, não a fórmula."),
 ("faixa", "LGPD — LEIA ANTES DE ESCREVER"),
 ("p", "", "Não registre dado sensível: saúde, diagnóstico, religião, orientação sexual, filiação sindical, origem racial, vida familiar. "
           "Ausência registra o FATO, nunca o motivo."),
 ("p", "", "A pessoa pode pedir acesso ao que se registrou sobre ela. Regra prática: não escreva nada que você não sustentaria lendo em voz alta para ela."),
 ("p", "", "Arquivo único no drive corporativo, com coautoria. Nunca anexo de e-mail, nunca cópia local. Senha de planilha do Excel não é controle de acesso — "
           "por isso retenção, caso de promoção e registro privado do gestor vivem em ARQUIVO SEPARADO, não aqui."),
 ("faixa", "LEGENDA"),
 ("leg_in", "Fundo amarelo, texto azul", "Célula que você preenche."),
 ("leg_calc", "Fundo cinza", "Célula calculada. Não digite por cima — você quebra o Painel."),
 ("p", "", "As abas 'Pessoas', 'Alocações' e 'Config' são cadastro: ajuste no começo e não mexa no dia a dia. "
           "No dia a dia você convive com quatro abas: Painel, Ocorrências, Avaliação e 1a1."),
 ("faixa", "LIMITE HONESTO"),
 ("p", "", "Excel não tem controle de acesso por linha nem trilha de auditoria. Use isto para provar o modelo em 2 ou 3 ciclos. Se funcionar, migre para um sistema."),
 ("p", "", "As expectativas por senioridade e os pesos por eixo são premissa desta versão, não benchmark de mercado. Calibre ao fim do primeiro ciclo real "
           "e registre quem confirmou, na aba Config."),
]
r = 4
for b in BLOCOS:
    if b[0] == "faixa":
        faixa(ins, r, b[1], 2)
    elif b[0] == "kv":
        ins.cell(row=r, column=1, value=b[1]).font = fnt(9, b=True)
        c = ins.cell(row=r, column=2, value=b[2]); c.font = fnt(9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    elif b[0] == "n":
        ins.cell(row=r, column=1, value="Regra " + b[1]).font = fnt(9, b=True, color=C_TIT)
        c = ins.cell(row=r, column=2, value=b[2]); c.font = fnt(9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    elif b[0] == "p":
        c = ins.cell(row=r, column=2, value=b[2]); c.font = fnt(9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    elif b[0].startswith("leg"):
        c1 = ins.cell(row=r, column=1, value=b[1])
        c1.fill = PatternFill("solid", fgColor=C_IN if b[0] == "leg_in" else C_CALC)
        c1.font = fnt(9, color=BLUE if b[0] == "leg_in" else "000000"); c1.border = BOX
        ins.cell(row=r, column=2, value=b[2]).font = fnt(9)
    ins.row_dimensions[r].height = 30
    r += 1

del wb["Sheet"]
ORDEM = ["Instruções", "Painel", "Ocorrências", "Avaliação", "1a1", "PDI", "Alertas",
         "Pessoas", "Alocações", "Config"]
wb._sheets = [wb[n] for n in ORDEM]
wb.active = 1

# =====================================================================
# VALIDADOR ESTÁTICO — falha o build. Regra que não é verificada não existe.
# =====================================================================
PROIBIDAS = ["XLOOKUP", "FILTER(", "UNIQUE(", "SORT(", "SEQUENCE(", "SUMPRODUCT("]
SEM_PREFIXO = ["MAXIFS(", "MINIFS(", "TEXTJOIN(", "CONCAT(", "IFS(", "SWITCH("]
erros = []
abas = set(wb.sheetnames)
n_today = 0
n_formulas = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            f = c.value
            if not isinstance(f, str) or not f.startswith("="):
                continue
            n_formulas += 1
            loc = "{}!{}".format(ws.title, c.coordinate)
            # 1. referência de coluna inteira
            for m in re.finditer(r'(?:^|[^0-9A-Z$])\$?([A-Z]{1,3}):\$?([A-Z]{1,3})(?![0-9])', f):
                erros.append("{}: referência de coluna inteira '{}:{}'".format(loc, m.group(1), m.group(2)))
            # 2. funções proibidas
            for p in PROIBIDAS:
                if p in f.upper():
                    erros.append("{}: função proibida {}".format(loc, p))
            # 3. INDEX sobre expressão / fórmula matricial disfarçada
            if re.search(r'INDEX\(\s*\(', f):
                erros.append("{}: INDEX sobre expressão (matricial disfarçada)".format(loc))
            # 4. _xlfn ausente ou duplicado
            up = f.upper()
            for fn in SEM_PREFIXO:
                for m in re.finditer(re.escape(fn), up):
                    # so conta se for inicio real do nome da funcao (nao dentro de COUNTIFS/SUMIFS)
                    ant = up[m.start() - 1] if m.start() > 0 else ""
                    if ant.isalpha():
                        continue
                    if not up[max(0, m.start() - 6):m.start()].endswith("_XLFN."):
                        erros.append("{}: {} sem prefixo _xlfn.".format(loc, fn))
            if "_XLFN._XLFN." in f.upper():
                erros.append("{}: prefixo _xlfn duplicado".format(loc))
            # 5. TODAY volátil
            n_today += len(re.findall(r'\bTODAY\(\)', f))
            # 6. aba referenciada existe
            for m in re.finditer(r"'([^']+)'!", f):
                if m.group(1) not in abas:
                    erros.append("{}: aba inexistente '{}'".format(loc, m.group(1)))
            for m in re.finditer(r"(?:^|[^A-Za-z0-9_'!])([A-Za-z][A-Za-zÀ-ÿ0-9_]*)!", f):
                nome = m.group(1)
                if nome not in abas and nome not in ("TRUE", "FALSE"):
                    erros.append("{}: aba inexistente '{}'".format(loc, nome))
if n_today != 1:
    erros.append("TODAY() aparece {} vezes; o orçamento permite exatamente 1 (Config!B6).".format(n_today))

if erros:
    print("BUILD REPROVADO — {} problema(s):".format(len(erros)))
    for e in sorted(set(erros))[:40]:
        print("  ·", e)
    sys.exit(1)

wb.save(OUT)
print("OK  {}".format(OUT))
print("    {} fórmulas · {} abas · validador estático: 0 problemas".format(n_formulas, len(wb.sheetnames)))

# =====================================================================
# MODELO-SOMBRA — o que as células-chave DEVEM valer, calculado em Python
# =====================================================================
notas = {c: EX[c][2] for c in COMP_NOMES}
por_eixo = {}
for cid, eixo, nome, dono, _ in COMP:
    por_eixo.setdefault(eixo, []).append(notas[nome])
med = {e: sum(v) / len(v) for e, v in por_eixo.items()}
ponderada = sum(med[e] * PESO[e] for e in EIXOS) / sum(PESO[e] for e in EIXOS)
exp_pleno = dict(SENIOR)["Pleno"]
ocorr_ana = sum(1 for o in OCORR if o[1] == "Ana Ribeiro" and o[0].month in (7, 8, 9))
ESPERADO = {
 "Painel!F7":  round(med[EA], 4),
 "Painel!G7":  round(med[EB], 4),
 "Painel!H7":  round(med[EC], 4),
 "Painel!I7":  round(med[ED], 4),
 "Painel!J7":  round(ponderada, 4),
 "Painel!K7":  exp_pleno,
 "Painel!L7":  round(ponderada - exp_pleno, 4),
 "Painel!N7":  ocorr_ana,
 "Painel!T7":  0,
 "Painel!U7":  0,
 "Alertas!C5": 2,
 "Alertas!C6": 0,
 "Alertas!C7": 0,
 "Alertas!C8": 0,
 "Alertas!C11": 0,
}
import json
with open("/home/user/kaneo/docs/gestao-pessoas/.modelo_sombra.json", "w") as fh:
    json.dump(ESPERADO, fh, indent=1, ensure_ascii=False)
print("    modelo-sombra: {} células-chave esperadas gravadas".format(len(ESPERADO)))
for k, v in ESPERADO.items():
    print("      {:14s} = {}".format(k, v))
