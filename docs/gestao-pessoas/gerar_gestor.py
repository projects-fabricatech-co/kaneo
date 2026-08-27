# -*- coding: utf-8 -*-
"""Gerador do arquivo PRIVADO do Gestor de Pessoas — SPEC-v3 secao 10.
Separado do arquivo de squad por PRIVACIDADE: Excel nao tem controle de acesso por linha,
e separacao de arquivo e o unico controle que o formato paga."""
import re, sys, datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.comments import Comment

OUT = "/home/user/kaneo/docs/gestao-pessoas/Avaliacao_Desempenho_Gestor_v3.xlsx"
F = "Arial"
C_TIT, C_HEAD, C_SUB = "7B2E2E", "9C3A3A", "F2DCDB"
C_IN, C_CALC, C_WARN, C_BAD, C_GOOD = "FFF2CC", "F2F2F2", "FDE9D9", "FFC7CE", "C6EFCE"
BLUE = "0000FF"
thin = Side(style="thin", color="BFBFBF"); BOX = Border(thin, thin, thin, thin)
CICLO = "2026-Q3"
LAST_RET, LAST_PRO, LAST_REG = 104, 44, 154

def fnt(sz=10, b=False, color="000000", it=False):
    return Font(name=F, size=sz, bold=b, color=color, italic=it)

wb = openpyxl.Workbook(); wb.calculation.fullCalcOnLoad = True
def sh(n):
    ws = wb.create_sheet(n); ws.sheet_view.showGridLines = False; return ws
def titulo(ws, t, sub=None):
    ws["A1"] = t; ws["A1"].font = Font(name=F, size=14, bold=True, color=C_TIT)
    ws.row_dimensions[1].height = 22
    if sub:
        ws["A2"] = sub; ws["A2"].font = fnt(9, it=True, color="595959")
def cab(ws, row, vals):
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=1 + i, value=v)
        c.font = Font(name=F, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD); c.border = BOX
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 32
def larg(ws, spec):
    for k, v in spec.items(): ws.column_dimensions[k].width = v
def faixa(ws, row, txt, span=10):
    ws.cell(row=row, column=1, value=txt).font = Font(name=F, size=10, bold=True, color=C_TIT)
    for j in range(1, span + 1):
        ws.cell(row=row, column=j).fill = PatternFill("solid", fgColor=C_SUB)
def pinta(ws, rng, kind):
    fill = PatternFill("solid", fgColor={"in": C_IN, "calc": C_CALC}[kind])
    for r in ws[rng]:
        for c in r:
            c.fill = fill; c.border = BOX
            c.font = fnt(9, color=BLUE if kind == "in" else "000000")
            c.alignment = Alignment(vertical="top", wrap_text=True)

NIVEIS = ["Baixo", "Médio", "Alto"]
FATORES = ["Remuneração", "Alocação / projeto", "Carreira travada", "Relação com a liderança",
           "Carga de trabalho", "Fora do trabalho", "Não identificado"]
BASES = ["Declarado pela pessoa", "Observado por mim", "Terceiro reportou"]
STATUS = ["Aberto", "Em ação", "Encerrado — retida", "Encerrado — continua em risco", "Encerrado — saiu"]
TIPOS_REG = ["Aspiração de carreira", "Compromisso que assumi", "Feedback de terceiro",
             "Leitura de padrão entre ciclos"]
ST_REG = ["Em aberto", "Cumprido", "Renegociado", "Não cumprido"]
SENIOR = ["Júnior", "Pleno", "Sênior", "Especialista"]

# =====================================================================
# INSTRUÇÕES (+ configuração mínima)
# =====================================================================
ins = sh("Instruções")
titulo(ins, "Arquivo privado do Gestor de Pessoas",
       "NÃO compartilhe com GP nem com Líder Técnico. Este arquivo existe separado por privacidade, não por organização.")
larg(ins, {"A": 34, "B": 96})
ins["A4"] = "Ciclo ativo"; ins["A4"].font = fnt(10, b=True)
ins["B4"] = CICLO
ins["B4"].fill = PatternFill("solid", fgColor=C_IN); ins["B4"].font = fnt(10, b=True, color=BLUE); ins["B4"].border = BOX
ins["A5"] = "Hoje"; ins["A5"].font = fnt(10, b=True)
ins["B5"] = "=TODAY()"; ins["B5"].number_format = "DD/MM/YYYY"
ins["A6"] = "Índice do ciclo ativo"; ins["A6"].font = fnt(10, b=True)
ins["B6"] = '=VALUE(LEFT($B$4,4))*4+VALUE(RIGHT($B$4,1))'
for r in (5, 6):
    ins.cell(row=r, column=2).fill = PatternFill("solid", fgColor=C_CALC)
    ins.cell(row=r, column=2).border = BOX; ins.cell(row=r, column=2).font = fnt(10, b=True)

BL = [
 ("faixa", "POR QUE ESTE ARQUIVO É SEPARADO"),
 ("p", "", "Excel não tem controle de acesso por linha. Prometer 'só o Gestor vê' dentro do arquivo que o GP abre todo dia "
           "seria falso. Separação de arquivo é o único controle de acesso que o formato paga — por isso retenção, caso de "
           "promoção e registro privado vivem aqui, e não na planilha da squad."),
 ("p", "", "Se o GP enxergar a avaliação de risco de retenção, ele muda como aloca a pessoa. Vira profecia autorrealizável."),
 ("faixa", "COMO OS DADOS CHEGAM AQUI"),
 ("p", "", "Por cópia manual, no fechamento do ciclo. NÃO crie link externo para a planilha da squad: link externo quebra, "
           "ninguém conserta, e o Excel passa a mostrar o último valor em cache como se fosse atual."),
 ("p", "", "Use sempre o ID da pessoa (P001), nunca só o nome. Nome muda; ID não."),
 ("faixa", "RISCO DE RETENÇÃO — AS DUAS REGRAS QUE IMPORTAM"),
 ("n", "1", "Risco é uma AVALIAÇÃO DATADA, nunca um atributo da pessoa. Toda linha tem data e validade de 2 ciclos. "
            "Risco não reafirmado esvazia sozinho — é a única defesa contra 'risco alto' de dois anos atrás virar cicatriz "
            "que o próximo gestor lê como fato."),
 ("n", "2", "Alto desempenho × alto risco é o único quadrante com PRAZO: ação em 15 dias, com fator identificado e "
            "contrapartida concreta. Conversa vaga não conta como ação. E registre o desfecho — quem não fecha o loop "
            "nunca aprende por que perde gente."),
 ("faixa", "CASO DE PROMOÇÃO"),
 ("p", "", "Promoção se sustenta em trajetória, não em um trimestre bom. No primeiro ciclo de uso o veredito é 'Sem base' — "
           "nunca 'Caso frágil'. 'Frágil' é um juízo sobre a pessoa; emiti-lo por ausência de dados que ainda não podiam "
           "existir é injusto e queima a ferramenta na primeira semana."),
 ("p", "", "Diversidade de evidência é o número mais subestimado da tabela: notas altas vindas todas do mesmo avaliador "
           "não são um caso, são uma opinião. Pegue o valor na coluna 'novo avaliador' da aba Ocorrências da squad."),
 ("faixa", "REGISTRO DO GESTOR"),
 ("p", "", "Compromisso que você assumiu e não registrou é dívida invisível — e é a principal causa de pedido de demissão "
           "surpresa. Os que vencem em 7 dias ficam destacados no topo da aba."),
 ("faixa", "O QUE NÃO ENTRA AQUI, EM NENHUMA ABA"),
 ("p", "", "Conduta, assédio, denúncia, saúde, diagnóstico, religião, orientação sexual, filiação sindical, origem racial e "
           "vida familiar. Isso corre em processo próprio do RH. 'Contexto pessoal' significa carga e disponibilidade, "
           "não intimidade."),
 ("p", "", "Regra prática que vale para as três abas: não escreva nada que você não sustentaria lendo em voz alta para a pessoa."),
]
r = 8
for b in BL:
    if b[0] == "faixa":
        faixa(ins, r, b[1], 2)
    elif b[0] == "n":
        ins.cell(row=r, column=1, value="Regra " + b[1]).font = fnt(9, b=True, color=C_TIT)
        c = ins.cell(row=r, column=2, value=b[2]); c.font = fnt(9); c.alignment = Alignment(wrap_text=True, vertical="top")
    else:
        c = ins.cell(row=r, column=2, value=b[2]); c.font = fnt(9); c.alignment = Alignment(wrap_text=True, vertical="top")
    ins.row_dimensions[r].height = 32
    r += 1

# listas
ins["D4"] = "LISTAS"; ins["D4"].font = fnt(10, b=True, color=C_TIT)
LISTAS = [("D", "Nível", NIVEIS), ("E", "Fator", FATORES), ("F", "Base do sinal", BASES),
          ("G", "Status retenção", STATUS), ("H", "Tipo registro", TIPOS_REG),
          ("I", "Status registro", ST_REG), ("J", "Senioridade", SENIOR)]
for col, head, vals in LISTAS:
    c = ins[col + "5"]; c.value = head
    c.font = Font(name=F, size=9, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=C_HEAD)
    c.border = BOX; c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, v in enumerate(vals):
        cc = ins[col + str(6 + i)]; cc.value = v
        cc.fill = PatternFill("solid", fgColor=C_IN); cc.font = fnt(9, color=BLUE); cc.border = BOX
    ins.column_dimensions[col].width = 24
def LST(col, n): return "=Instruções!${0}$6:${0}${1}".format(col, 5 + n)
DV = {"nivel": LST("D", len(NIVEIS)), "fator": LST("E", len(FATORES)), "base": LST("F", len(BASES)),
      "stret": LST("G", len(STATUS)), "tipo": LST("H", len(TIPOS_REG)), "streg": LST("I", len(ST_REG)),
      "senior": LST("J", len(SENIOR))}
def dv(ws, key, rng):
    d = DataValidation(type="list", formula1=DV[key], allow_blank=True)
    ws.add_data_validation(d); d.add(rng)

# =====================================================================
# RETENÇÃO
# =====================================================================
re_ = sh("Retenção")
titulo(re_, "Risco de retenção — uma linha por AVALIAÇÃO, com data",
       "Nunca um atributo permanente da pessoa. A validade de 2 ciclos é o mecanismo central desta aba.")
re_["A3"] = ("Risco não reafirmado esvazia sozinho. Sem isso, 'risco alto' de dois anos atrás vira cicatriz que o próximo "
             "gestor lê como fato — e a pessoa paga por uma conversa que ninguém lembra.")
re_["A3"].font = fnt(9, it=True, color="C00000")
COLS = ["Data", "ID_PESSOA", "Nome", "Nível", "Fator", "Base do sinal", "Ação concreta", "Prazo",
        "Status", "Desfecho registrado", "Idx ciclo", "Válido até (idx)", "Vigente?", "Dias sem ação"]
larg(re_, {"A": 12, "B": 11, "C": 20, "D": 10, "E": 20, "F": 22, "G": 46, "H": 12, "I": 26,
           "J": 30, "K": 10, "L": 13, "M": 11, "N": 13})
cab(re_, 4, COLS)
EXR = [(dt.date(2026,8,26), "P003", "Carla Nunes", "Alto", "Carreira travada", "Declarado pela pessoa",
        "Assumir a trilha técnica do Contrato B como referência de arquitetura, com escopo formalizado até 15/09.",
        dt.date(2026,9,10), "Em ação", "")]
for i, row in enumerate(EXR):
    for j, v in enumerate(row): re_.cell(row=5 + i, column=1 + j, value=v)
for rr in range(5, LAST_RET + 1):
    re_.cell(row=rr, column=11, value='=IF($A{0}="","",YEAR($A{0})*4+ROUNDUP(MONTH($A{0})/3,0))'.format(rr))
    re_.cell(row=rr, column=12, value='=IF($K{0}="","",$K{0}+2)'.format(rr))
    re_.cell(row=rr, column=13, value=(
        '=IF($A{0}="","",IF(AND(LEFT($I{0},9)<>"Encerrado",$L{0}>=Instruções!$B$6),"Sim","Não"))').format(rr))
    re_.cell(row=rr, column=14, value=(
        '=IF(OR($A{0}="",$M{0}<>"Sim"),"",IF($D{0}<>"Alto","",IF($I{0}="Em ação","em ação",'
        'Instruções!$B$5-$A{0})))').format(rr))
    for col in ("A", "H"): re_[col + str(rr)].number_format = "DD/MM/YYYY"
pinta(re_, "A5:J{}".format(LAST_RET), "in")
pinta(re_, "K5:N{}".format(LAST_RET), "calc")
for k, col in [("nivel", "D"), ("fator", "E"), ("base", "F"), ("stret", "I")]:
    dv(re_, k, "{0}5:{0}{1}".format(col, LAST_RET))
re_.conditional_formatting.add("A5:N{}".format(LAST_RET),
    FormulaRule(formula=['AND($M5="Sim",$D5="Alto")'], fill=PatternFill("solid", fgColor=C_WARN)))
re_.conditional_formatting.add("A5:N{}".format(LAST_RET),
    FormulaRule(formula=['$M5="Não"'], font=Font(name=F, size=9, color="A6A6A6")))
re_.conditional_formatting.add("N5:N{}".format(LAST_RET),
    CellIsRule(operator="greaterThan", formula=["15"], fill=PatternFill("solid", fgColor=C_BAD),
               font=Font(name=F, size=9, bold=True, color="9C0006")))
re_["N4"].comment = Comment(
 "Dias desde a avaliação, para risco ALTO ainda vigente e sem ação registrada.\n\n"
 "Acima de 15 dias fica vermelho. Alto desempenho × alto risco é o único quadrante com prazo — "
 "e conversa vaga não conta como ação.", "Gestão de Pessoas")
re_["F4"].comment = Comment(
 "'Declarado pela pessoa' vale mais que 'Observado por mim', e muito mais que 'Terceiro reportou'.\n\n"
 "Sem esta coluna, boato de corredor e conversa franca entram na planilha com o mesmo peso.", "Gestão de Pessoas")
re_["J4"].comment = Comment(
 "Retida · Continua em risco · Saiu.\n\nQuem não fecha o loop nunca aprende por que perde gente.", "Gestão de Pessoas")
re_.freeze_panes = "C5"
re_.auto_filter.ref = "A4:N{}".format(LAST_RET)

# =====================================================================
# CASO DE PROMOÇÃO
# =====================================================================
pr = sh("Caso de Promoção")
titulo(pr, "Caso de promoção — trajetória, não ciclo",
       "Cole os gaps do Painel da squad ao fim de cada ciclo. Promoção se sustenta em desempenho sustentado, não em um trimestre bom.")
pr["A3"] = ("Notas altas vindas todas do mesmo avaliador não são um caso, são uma opinião. "
            "Diversidade de evidência é o número mais subestimado desta tabela.")
pr["A3"].font = fnt(9, it=True, color="C00000")
COLS_P = ["ID_PESSOA", "Nome", "Senioridade", "Meses no nível",
          "gap -5", "gap -4", "gap -3", "gap -2", "gap -1", "gap ciclo atual",
          "Gap médio (3 últimos)", "Ciclos consecutivos com gap ≥ 0", "Ocorrências vinculadas",
          "Avaliadores distintos", "Evidência de escopo ampliado", "Caso montado?", "Veredito",
          "c1", "c2", "c3", "c4", "c5", "c6"]
larg(pr, {"A": 11, "B": 20, "C": 13, "D": 13, "E": 9, "F": 9, "G": 9, "H": 9, "I": 9, "J": 11,
          "K": 13, "L": 14, "M": 12, "N": 12, "O": 46, "P": 13, "Q": 22,
          "R": 5, "S": 5, "T": 5, "U": 5, "V": 5, "W": 5})
# rotulos de ciclo calculados a partir do indice ativo
for i in range(6):
    off = i - 5
    pr.cell(row=3, column=5 + i, value=(
        '=TEXT(INT((Instruções!$B$6{0}-1)/4),"0000")&"-Q"&(Instruções!$B$6{0}-INT((Instruções!$B$6{0}-1)/4)*4)'
        ).format("" if off == 0 else "{0:+d}".format(off)))
    pr.cell(row=3, column=5 + i).font = fnt(8, b=True, color="595959")
    pr.cell(row=3, column=5 + i).alignment = Alignment(horizontal="center")
cab(pr, 4, COLS_P)
for i, lab in enumerate(["c1", "c2", "c3", "c4", "c5", "c6"]):
    c = pr.cell(row=4, column=18 + i)
    c.fill = PatternFill("solid", fgColor="A6A6A6"); c.font = Font(name=F, size=8, italic=True, color="FFFFFF")
EXP = [("P001", "Ana Ribeiro", "Pleno", None, None, None, None, -0.10, 0.20, 0.375, None, None, 2, 2,
        "Assumiu os testes da rota de resgate como dona, com o LT como rede de proteção.")]
for i, row in enumerate(EXP):
    for j, v in enumerate(row):
        if v is not None: pr.cell(row=5 + i, column=1 + j, value=v)
for rr in range(5, LAST_PRO + 1):
    # cadeia de consecutivos, do ciclo mais antigo para o mais novo
    pr.cell(row=rr, column=18, value='=IF($E{0}="",0,IF($E{0}>=0,1,0))'.format(rr))
    for i in range(1, 6):
        col_gap = chr(ord("E") + i)
        pr.cell(row=rr, column=18 + i, value=(
            '=IF(${1}{0}="",0,IF(${1}{0}>=0,${2}{0}+1,0))').format(rr, col_gap, chr(ord("R") + i - 1)))
    pr.cell(row=rr, column=11, value=(
        '=IF($B{0}="","",IF(COUNT($H{0}:$J{0})=0,"",AVERAGE($H{0}:$J{0})))').format(rr))
    pr.cell(row=rr, column=12, value=(
        '=IF($B{0}="","",IF($J{0}<>"",$W{0},IF($I{0}<>"",$V{0},IF($H{0}<>"",$U{0},'
        'IF($G{0}<>"",$T{0},IF($F{0}<>"",$S{0},$R{0}))))))').format(rr))
    pr.cell(row=rr, column=16, value=(
        '=IF($B{0}="","",IF(AND($M{0}>0,$N{0}>1,LEN($O{0})>20),"Sim","Não"))').format(rr))
    pr.cell(row=rr, column=17, value=(
        '=IF($B{0}="","",'
        'IF(COUNT($E{0}:$J{0})<3,"Sem base — poucos ciclos",'
        'IF($N{0}<=1,"Caso frágil — um avaliador só",'
        'IF($M{0}=0,"Caso frágil — sem evidência",'
        'IF($L{0}<2,"Caso frágil — não sustentado",'
        'IF(AND($L{0}>=3,$K{0}>=0.2,$P{0}="Sim"),"Caso sustentado","Caso em formação"))))))').format(rr))
    for col in ("E", "F", "G", "H", "I", "J", "K"):
        pr[col + str(rr)].number_format = "+0.00;-0.00;0.00"
        pr[col + str(rr)].alignment = Alignment(horizontal="center")
    for col in ("D", "L", "M", "N"):
        pr[col + str(rr)].alignment = Alignment(horizontal="center")
pinta(pr, "A5:J{}".format(LAST_PRO), "in")
pinta(pr, "M5:O{}".format(LAST_PRO), "in")
pinta(pr, "K5:L{}".format(LAST_PRO), "calc")
pinta(pr, "P5:W{}".format(LAST_PRO), "calc")
dv(pr, "senior", "C5:C{}".format(LAST_PRO))
for txt, cor in [("Caso sustentado", C_GOOD), ("Caso em formação", C_WARN)]:
    pr.conditional_formatting.add("Q5:Q{}".format(LAST_PRO),
        CellIsRule(operator="equal", formula=['"{}"'.format(txt)], fill=PatternFill("solid", fgColor=cor)))
pr.conditional_formatting.add("Q5:Q{}".format(LAST_PRO),
    FormulaRule(formula=['LEFT($Q5,12)="Caso frágil"'], fill=PatternFill("solid", fgColor=C_BAD)))
pr.conditional_formatting.add("Q5:Q{}".format(LAST_PRO),
    FormulaRule(formula=['LEFT($Q5,8)="Sem base"'], font=Font(name=F, size=9, italic=True, color="A6A6A6")))
pr["Q4"].comment = Comment(
 "No primeiro ciclo de uso o veredito é 'Sem base', NUNCA 'Caso frágil'.\n\n"
 "'Frágil' é um juízo sobre a pessoa. Emiti-lo por ausência de dados que ainda não podiam existir "
 "é injusto e queima a ferramenta na primeira semana.", "Qualidade")
pr["N4"].comment = Comment(
 "Quantos avaliadores DIFERENTES registraram ocorrência sobre esta pessoa no período.\n\n"
 "Pegue na coluna 'novo avaliador' da aba Ocorrências da planilha da squad: some as marcas de 1 da pessoa.\n\n"
 "1 avaliador = opinião. 2 ou mais = caso.", "Técnico")
pr["O4"].comment = Comment(
 "O que ela fez ALÉM da própria entrega: mentoria, incidente fora do escopo, ADR que outros times adotaram, "
 "automação que economizou tempo do time.\n\n"
 "É o que separa 'faz bem o trabalho' de 'merece o próximo nível' — e quase nunca aparece em métrica de entrega.",
 "Liderança Técnica")
pr.freeze_panes = "C5"

# =====================================================================
# REGISTRO DO GESTOR
# =====================================================================
rg = sh("Registro do Gestor")
titulo(rg, "Registro privado do gestor",
       "Aspirações, compromissos que você assumiu, feedback de terceiros e padrões que só você vê.")
rg["A3"] = ('="Compromissos vencendo em até 7 dias: "&COUNTIFS($D$6:$D${0},"Compromisso que assumi",'
            '$G$6:$G${0},"Em aberto",$F$6:$F${0},"<="&Instruções!$B$5+7,$F$6:$F${0},">="&Instruções!$B$5)'
            '&"   ·   já vencidos: "&COUNTIFS($D$6:$D${0},"Compromisso que assumi",$G$6:$G${0},"Em aberto",'
            '$F$6:$F${0},"<"&Instruções!$B$5)').format(LAST_REG)
rg["A3"].font = Font(name=F, size=11, bold=True, color="9C0006")
rg["A3"].comment = Comment(
 "Compromisso que você assumiu e não cumpriu é dívida invisível — e é a principal causa de pedido "
 "de demissão surpresa.\n\nRegistrar sem prazo visível só documenta a dívida; não paga.", "Gestão de Pessoas")
COLS_R = ["Data", "ID_PESSOA", "Nome", "Tipo", "Registro", "Prazo", "Status", "Vence em (dias)"]
larg(rg, {"A": 12, "B": 11, "C": 20, "D": 24, "E": 72, "F": 12, "G": 14, "H": 14})
cab(rg, 5, COLS_R)
EXG = [
 (dt.date(2026,8,20), "P001", "Ana Ribeiro", "Compromisso que assumi",
  "Prometi levar o caso dela para o comitê de calibração de outubro com a evidência do PR #812.",
  dt.date(2026,10,15), "Em aberto"),
 (dt.date(2026,8,26), "P003", "Carla Nunes", "Aspiração de carreira",
  "Quer trilha técnica, não gestão. Disse que aceitou o papel de LT por falta de alternativa, não por escolha.",
  None, "Em aberto"),
 (dt.date(2026,8,12), "P002", "Bruno Tavares", "Leitura de padrão entre ciclos",
  "Comunica bem quando o problema é técnico e trava quando o problema é político. Vale observar mais um ciclo antes de virar feedback.",
  None, "Em aberto"),
]
for i, row in enumerate(EXG):
    for j, v in enumerate(row):
        if v is not None: rg.cell(row=6 + i, column=1 + j, value=v)
for rr in range(6, LAST_REG + 1):
    rg.cell(row=rr, column=8, value=(
        '=IF(OR($F{0}="",$G{0}<>"Em aberto"),"",$F{0}-Instruções!$B$5)').format(rr))
    for col in ("A", "F"): rg[col + str(rr)].number_format = "DD/MM/YYYY"
    rg["H" + str(rr)].alignment = Alignment(horizontal="center")
pinta(rg, "A6:G{}".format(LAST_REG), "in")
pinta(rg, "H6:H{}".format(LAST_REG), "calc")
dv(rg, "tipo", "D6:D{}".format(LAST_REG))
dv(rg, "streg", "G6:G{}".format(LAST_REG))
rg.conditional_formatting.add("A6:H{}".format(LAST_REG),
    FormulaRule(formula=['AND($H6<>"",$H6<0)'], fill=PatternFill("solid", fgColor=C_BAD)))
rg.conditional_formatting.add("A6:H{}".format(LAST_REG),
    FormulaRule(formula=['AND($H6<>"",$H6>=0,$H6<=7)'], fill=PatternFill("solid", fgColor=C_WARN)))
rg["E5"].comment = Comment(
 "Nada de saúde, diagnóstico, vida familiar, religião ou conduta. 'Contexto pessoal' aqui significa "
 "carga e disponibilidade, não intimidade.\n\n"
 "Regra prática: não escreva nada que você não sustentaria lendo em voz alta para a pessoa.", "RH")
rg.freeze_panes = "C6"
rg.auto_filter.ref = "A5:H{}".format(LAST_REG)

del wb["Sheet"]
wb._sheets = [wb[n] for n in ["Instruções", "Retenção", "Caso de Promoção", "Registro do Gestor"]]
wb.active = 1

# =====================================================================
# VALIDADOR ESTÁTICO — mesmo portão do arquivo de squad
# =====================================================================

def desbalanceada(f):
    """Parenteses fora de literal de texto. O motor de formulas so acusa isso no calculo;
    o build tem de pegar antes."""
    d, dentro = 0, False
    for ch in f:
        if ch == '"':
            dentro = not dentro
        elif not dentro:
            if ch == "(":
                d += 1
            elif ch == ")":
                d -= 1
            if d < 0:
                return True
    return d != 0

PROIBIDAS = ["XLOOKUP", "FILTER(", "UNIQUE(", "SORT(", "SEQUENCE(", "SUMPRODUCT("]
SEM_PREFIXO = ["MAXIFS(", "MINIFS(", "TEXTJOIN(", "CONCAT(", "IFS(", "SWITCH("]
erros, abas, n_today, n_f = [], set(wb.sheetnames), 0, 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            f = c.value
            if not isinstance(f, str) or not f.startswith("="): continue
            n_f += 1; up = f.upper(); loc = "{}!{}".format(ws.title, c.coordinate)
            if desbalanceada(f):
                erros.append("{}: parênteses desbalanceados".format(loc))
            for m in re.finditer(r'(?:^|[^0-9A-Z$])\$?([A-Z]{1,3}):\$?([A-Z]{1,3})(?![0-9])', f):
                erros.append("{}: coluna inteira {}:{}".format(loc, m.group(1), m.group(2)))
            for p in PROIBIDAS:
                if p in up: erros.append("{}: função proibida {}".format(loc, p))
            if re.search(r'INDEX\(\s*\(', f): erros.append("{}: INDEX sobre expressão".format(loc))
            for fn in SEM_PREFIXO:
                for m in re.finditer(re.escape(fn), up):
                    ant = up[m.start() - 1] if m.start() > 0 else ""
                    if ant.isalpha(): continue
                    if not up[max(0, m.start() - 6):m.start()].endswith("_XLFN."):
                        erros.append("{}: {} sem prefixo _xlfn.".format(loc, fn))
            if "_XLFN._XLFN." in up: erros.append("{}: prefixo duplicado".format(loc))
            n_today += len(re.findall(r'\bTODAY\(\)', f))
            for m in re.finditer(r"'([^']+)'!", f):
                if m.group(1) not in abas: erros.append("{}: aba inexistente '{}'".format(loc, m.group(1)))
            for m in re.finditer(r"(?:^|[^A-Za-z0-9_'!])([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ0-9_]*)!", f):
                if m.group(1) not in abas: erros.append("{}: aba inexistente '{}'".format(loc, m.group(1)))
if n_today != 1:
    erros.append("TODAY() aparece {} vezes; permitido exatamente 1.".format(n_today))
if erros:
    print("BUILD REPROVADO — {} problema(s):".format(len(erros)))
    for e in sorted(set(erros))[:30]: print("  ·", e)
    sys.exit(1)
wb.save(OUT)
print("OK  {}".format(OUT))
print("    {} fórmulas · {} abas · validador estático: 0 problemas".format(n_f, len(wb.sheetnames)))
